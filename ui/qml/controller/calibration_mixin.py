from __future__ import annotations

import csv
import math
from pathlib import Path
import re
import time
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

from j1939.j1939_can_identifier import J1939CanIdentifier
from colors import RowColor
from uds.data_identifiers import UdsData
from uds.uds_identifiers import UdsIdentifiers

from .contract import AppControllerContract

class AppControllerCalibrationMixin(AppControllerContract):
    _FUEL_TEMP_COMP_REF_X10 = 200
    _INT16_MIN = -32768
    _INT16_MAX = 32767
    _TEMP_COMP_CHART_POINT_LIMIT = 8000
    _TEMP_COMP_SEGMENT_COUNT = 5
    _TEMP_COMP_SEGMENT_BORDERS_COUNT = 4
    _TEMP_COMP_DEFAULT_SEG_BORDERS_X10 = (-200, 0, 200, 400)
    _TEMP_COMP_DIR_HYST_MIN_X10 = 1
    _TEMP_COMP_DIR_HYST_DEFAULT_X10 = 5
    _TEMP_COMP_REC_MIN_BRANCH_POINTS = 8
    _TEMP_COMP_REC_MIN_SPAN_MODE2_X10 = 80
    _TEMP_COMP_REC_MIN_SPAN_MODE1_X10 = 40
    _TEMP_COMP_REC_MIN_SEG_POINTS = 24
    _TEMP_COMP_REC_MIN_SEG_SPAN_X10 = 20
    _TEMP_COMP_REC_MODE1_MIN_GAIN = 0.04
    _TEMP_COMP_REC_MODE2_MIN_GAIN = 0.08
    _TEMP_COMP_MODE_NAMES = {
        0: "single (линейная K1)",
        1: "segmented (K1 по сегментам)",
        2: "segmented heat/cool (раздельно нагрев/охлаждение)",
    }
    _TEMP_COMP_MODE_SINGLE_LINEAR = 0
    _TEMP_COMP_MODE_SEGMENTED = 1
    _TEMP_COMP_MODE_SEGMENTED_HEAT_COOL = 2
    _TEMP_COMP_TREND_COOLING = -1
    _TEMP_COMP_TREND_STABLE = 0
    _TEMP_COMP_TREND_HEATING = 1
    _TEMP_COMP_ADV_SEGMENT_BORDER_KEYS = (
        "seg_t1_x10",
        "seg_t2_x10",
        "seg_t3_x10",
        "seg_t4_x10",
    )
    _TEMP_COMP_ADV_COOLING_KEYS = (
        "k1_cool_seg1_x100",
        "k1_cool_seg2_x100",
        "k1_cool_seg3_x100",
        "k1_cool_seg4_x100",
        "k1_cool_seg5_x100",
    )
    _TEMP_COMP_ADV_HEATING_KEYS = (
        "k1_heat_seg1_x100",
        "k1_heat_seg2_x100",
        "k1_heat_seg3_x100",
        "k1_heat_seg4_x100",
        "k1_heat_seg5_x100",
    )
    _TEMP_COMP_BASE_READ_FIELDS: tuple[dict[str, object], ...] = (
        {
            "key": "base_k1",
            "var": UdsData.fuel_temp_comp_k1_x100,
            "label": "Коэффициент K1",
        },
        {
            "key": "base_k0",
            "var": UdsData.fuel_temp_comp_k0_count,
            "label": "Коэффициент K0",
        },
        {
            "key": "base_zero_trim",
            "var": UdsData.fuel_zero_trim_count,
            "label": "Смещение 0% (zero trim)",
        },
    )
    _TEMP_COMP_BASE_READ_KEYS = tuple(str(field.get("key", "")) for field in _TEMP_COMP_BASE_READ_FIELDS)
    _TEMP_COMP_ADVANCED_FIELDS = (
        {
            "key": "mode",
            "var": UdsData.fuel_temp_comp_mode,
            "label": "Режим компенсации",
            "unit": "",
            "signed": False,
            "min": 0,
            "max": 2,
            "placeholder": "Режим: 0=single, 1=segmented, 2=heat/cool",
        },
        {
            "key": "dir_hyst_x10",
            "var": UdsData.fuel_temp_comp_dir_hyst_x10,
            "label": "Гистерезис смены направления",
            "unit": "0.1 °C",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "Единицы 0.1 °C: например 5 = 0.5 °C",
        },
        {
            "key": "seg_t1_x10",
            "var": UdsData.fuel_temp_comp_seg_t1_x10,
            "label": "Граница T1 (S1/S2)",
            "unit": "0.1 °C",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "seg_t2_x10",
            "var": UdsData.fuel_temp_comp_seg_t2_x10,
            "label": "Граница T2 (S2/S3)",
            "unit": "0.1 °C",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "seg_t3_x10",
            "var": UdsData.fuel_temp_comp_seg_t3_x10,
            "label": "Граница T3 (S3/S4)",
            "unit": "0.1 °C",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "seg_t4_x10",
            "var": UdsData.fuel_temp_comp_seg_t4_x10,
            "label": "Граница T4 (S4/S5)",
            "unit": "0.1 °C",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_cool_seg1_x100",
            "var": UdsData.fuel_temp_comp_k1_cool_seg1_x100,
            "label": "K1 охлаждение S1",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_cool_seg2_x100",
            "var": UdsData.fuel_temp_comp_k1_cool_seg2_x100,
            "label": "K1 охлаждение S2",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_cool_seg3_x100",
            "var": UdsData.fuel_temp_comp_k1_cool_seg3_x100,
            "label": "K1 охлаждение S3",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_cool_seg4_x100",
            "var": UdsData.fuel_temp_comp_k1_cool_seg4_x100,
            "label": "K1 охлаждение S4",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_cool_seg5_x100",
            "var": UdsData.fuel_temp_comp_k1_cool_seg5_x100,
            "label": "K1 охлаждение S5",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_heat_seg1_x100",
            "var": UdsData.fuel_temp_comp_k1_heat_seg1_x100,
            "label": "K1 нагрев S1",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_heat_seg2_x100",
            "var": UdsData.fuel_temp_comp_k1_heat_seg2_x100,
            "label": "K1 нагрев S2",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_heat_seg3_x100",
            "var": UdsData.fuel_temp_comp_k1_heat_seg3_x100,
            "label": "K1 нагрев S3",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_heat_seg4_x100",
            "var": UdsData.fuel_temp_comp_k1_heat_seg4_x100,
            "label": "K1 нагрев S4",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
        {
            "key": "k1_heat_seg5_x100",
            "var": UdsData.fuel_temp_comp_k1_heat_seg5_x100,
            "label": "K1 нагрев S5",
            "unit": "count/°C x100",
            "signed": True,
            "min": -32768,
            "max": 32767,
            "placeholder": "signed int16 (dec/0xHEX)",
        },
    )

    @classmethod
    def _temp_comp_advanced_fields(cls) -> tuple[dict[str, object], ...]:
        """Цель функции в выдаче метаданных расширенной компенсации, затем она возвращает единый список полей для UI и UDS."""
        return cls._TEMP_COMP_ADVANCED_FIELDS

    @classmethod
    def _temp_comp_advanced_field_by_key(cls, key: str) -> dict[str, object] | None:
        """Цель функции в поиске поля по ключу, затем она возвращает его описание для операций чтения/записи."""
        target_key = str(key or "").strip()
        if not target_key:
            return None
        for field in cls._TEMP_COMP_ADVANCED_FIELDS:
            if str(field.get("key")) == target_key:
                return field
        return None

    @classmethod
    def _temp_comp_advanced_field_by_did(cls, did: int) -> dict[str, object] | None:
        """Цель функции в поиске поля по DID, затем она возвращает описание для разбора ответов UDS."""
        target_did = int(did) & 0xFFFF
        for field in cls._TEMP_COMP_ADVANCED_FIELDS:
            field_var = field.get("var")
            if field_var is None:
                continue
            if (int(field_var.pid) & 0xFFFF) == target_did:
                return field
        return None

    @classmethod
    def _temp_comp_read_field_by_key(cls, key: str) -> dict[str, object] | None:
        """Цель функции в поиске поля в общей карте чтения, затем она учитывает базовые DID и расширенные параметры."""
        target_key = str(key or "").strip()
        if not target_key:
            return None
        for field in cls._TEMP_COMP_BASE_READ_FIELDS:
            if str(field.get("key", "")) == target_key:
                return field
        return cls._temp_comp_advanced_field_by_key(target_key)

    @classmethod
    def _temp_comp_mode_text(cls, raw_mode: int) -> str:
        """Цель функции в понятном отображении режима, затем она возвращает код и имя режима для UI и логов."""
        mode_value = int(raw_mode)
        mode_name = cls._TEMP_COMP_MODE_NAMES.get(mode_value, "reserved")
        return f"{mode_value} ({mode_name})"

    @staticmethod
    def _temp_comp_field_display_name(field: dict[str, object]) -> str:
        """Цель функции в формировании короткого имени поля, затем она добавляет DID к подписи для читаемых логов."""
        field_var = field.get("var")
        did_text = "----"
        if field_var is not None:
            did_text = f"{int(field_var.pid) & 0xFFFF:04X}"
        return f"{str(field.get('label', 'Параметр'))} (DID 0x{did_text})"

    @staticmethod
    def _temp_comp_field_ui_value_text(field: dict[str, object], value: int | None) -> str:
        """Цель функции в подготовке текста значения для UI, затем она добавляет единицы измерения и формат режима."""
        if value is None:
            field_var = field.get("var")
            if field_var is None:
                return "не считан"
            return f"не считан (DID 0x{int(field_var.pid) & 0xFFFF:04X})"

        field_key = str(field.get("key", ""))
        if field_key == "mode":
            return AppControllerCalibrationMixin._temp_comp_mode_text(int(value))
        if field_key == "dir_hyst_x10":
            hysteresis_c = float(int(value)) / 10.0
            return f"{int(value)} (={hysteresis_c:.1f} °C)"

        unit = str(field.get("unit", "") or "").strip()
        if unit:
            return f"{int(value)} {unit}"
        return str(int(value))

    @staticmethod
    def _parse_calibration_response_identifier(identifier: int) -> tuple[int, int, int] | None:
        """Цель функции в безопасном разборе CAN-ID ответа, затем она возвращает pgn/dst/src для UDS-фильтрации."""
        try:
            parsed = J1939CanIdentifier(int(identifier))
        except Exception:
            return None
        parsed_pgn = int(parsed.pgn) & 0x3FFFF
        parsed_dst = int(parsed.dst) & 0xFF
        parsed_src = int(parsed.src) & 0xFF
        return parsed_pgn, parsed_dst, parsed_src

    def _try_bind_calibration_runtime_target_from_session_response(self, identifier: int, payload: list[int]):
        """Цель функции в привязке целевого SA в авто-режиме, затем она фиксирует узел по реальному ответу на 0x10."""
        if self._calibration_target_node_sa is not None:
            return
        if not self._calibration_waiting_session:
            return
        if str(self._calibration_sequence_waiting_action or "") != "activate_session":
            return
        if len(payload) < 2:
            return

        sid = int(payload[1]) & 0xFF
        is_session_answer = sid == 0x50
        if not is_session_answer:
            is_session_answer = sid == 0x7F and len(payload) >= 4 and (int(payload[2]) & 0xFF) == 0x10
        if not is_session_answer:
            return

        parsed = self._parse_calibration_response_identifier(identifier)
        if parsed is None:
            return
        parsed_pgn, parsed_dst, parsed_src = parsed
        expected_pgn = int(UdsIdentifiers.rx.pgn) & 0x3FFFF
        expected_dst = int(UdsIdentifiers.rx.dst) & 0xFF
        if parsed_pgn != expected_pgn or parsed_dst != expected_dst:
            return

        if self._calibration_runtime_target_sa == parsed_src:
            return
        self._calibration_runtime_target_sa = int(parsed_src) & 0xFF

        configured_auto_sa = int(UdsIdentifiers.rx.src) & 0xFF
        if int(parsed_src) != configured_auto_sa:
            self._append_log(
                (
                    f"Калибровка: авто-режим подтвердил ответ 0x10 от узла 0x{int(parsed_src) & 0xFF:02X}. "
                    "Для текущей сессии используется этот узел."
                ),
                RowColor.blue,
            )

    def _is_calibration_response_identifier(self, identifier: int) -> bool:
        parsed = self._parse_calibration_response_identifier(identifier)
        if parsed is None:
            return False
        parsed_pgn, parsed_dst, parsed_src = parsed

        expected_pgn = int(UdsIdentifiers.rx.pgn) & 0x3FFFF
        if parsed_pgn != expected_pgn:
            return False

        expected_dst = int(UdsIdentifiers.rx.dst) & 0xFF
        if parsed_dst != expected_dst:
            return False

        expected_src = int(self._resolve_calibration_target_sa()) & 0xFF
        return parsed_src == expected_src

    def _build_calibration_tx_identifier(self) -> int:
        try:
            tx = J1939CanIdentifier(int(UdsIdentifiers.tx.identifier))
            tx.dst = int(self._resolve_calibration_target_sa()) & 0xFF
            return int(tx.identifier)
        except Exception:
            return int(UdsIdentifiers.tx.identifier)

    def _resolve_calibration_target_sa(self) -> int:
        if self._calibration_target_node_sa is not None:
            return int(self._calibration_target_node_sa) & 0xFF
        if self._calibration_runtime_target_sa is not None:
            return int(self._calibration_runtime_target_sa) & 0xFF
        if 0 <= int(self._observed_candidate_index) < len(self._observed_candidate_values):
            return int(self._observed_candidate_values[int(self._observed_candidate_index)]) & 0xFF
        return int(UdsIdentifiers.rx.src) & 0xFF

    def _configure_calibration_uds_services(self):
        # Для этого МК DID в 0x22/0x2E всегда идут в стандартном UDS big-endian порядке.
        self._calibration_read_service.set_byte_order("big")
        self._calibration_write_service.set_byte_order("big")

    @staticmethod
    def _calibration_did_label(did: int) -> str:
        if int(did) == int(UdsData.empty_fuel_tank.pid):
            return "уровня 0%"
        if int(did) == int(UdsData.full_fuel_tank.pid):
            return "уровня 100%"
        if int(did) == int(UdsData.fuel_temp_comp_k1_x100.pid):
            return "коэффициента K1"
        if int(did) == int(UdsData.fuel_temp_comp_k0_count.pid):
            return "коэффициента K0"
        if int(did) == int(UdsData.fuel_zero_trim_count.pid):
            return "коррекции zero trim"
        field = AppControllerCalibrationMixin._temp_comp_advanced_field_by_did(int(did))
        if field is not None:
            return AppControllerCalibrationMixin._temp_comp_field_display_name(field)
        return f"DID 0x{int(did) & 0xFFFF:04X}"

    def _pending_calibration_write_did(self) -> int | None:
        if self._calibration_restore_current_did is not None:
            return int(self._calibration_restore_current_did)

        check_dids = [
            int(UdsData.empty_fuel_tank.pid),
            int(UdsData.full_fuel_tank.pid),
            int(UdsData.fuel_temp_comp_k1_x100.pid),
            int(UdsData.fuel_temp_comp_k0_count.pid),
            int(UdsData.fuel_zero_trim_count.pid),
        ]
        for field in self._temp_comp_advanced_fields():
            field_var = field.get("var")
            if field_var is None:
                continue
            check_dids.append(int(field_var.pid))

        for did in check_dids:
            if did in self._calibration_write_verify_pending:
                return did

        if len(self._calibration_write_verify_pending) == 1:
            return int(next(iter(self._calibration_write_verify_pending.keys())))

        return None

    def _is_calibration_security_ready(self) -> bool:
        if (not self._service_security_unlocked) or (self._service_access_target_sa is None):
            return False
        return (int(self._service_access_target_sa) & 0xFF) == self._resolve_calibration_target_sa()

    def _ensure_calibration_write_ready(self, operation_label: str) -> bool:
        if not self._can.is_connect:
            self.infoMessage.emit("Калибровка", "Сначала подключите CAN-адаптер.")
            return False

        if not self._can.is_trace:
            self.infoMessage.emit("Калибровка", "Сначала включите трассировку CAN.")
            return False

        if not self._calibration_session_ready:
            message = "Сначала запустите калибровку и дождитесь подтверждения extended-сессии UDS."
            self.infoMessage.emit("Калибровка", message)
            self._append_log(f"Калибровка: {operation_label} не отправлена. {message}", RowColor.yellow)
            return False

        target_sa = self._resolve_calibration_target_sa()
        if self._service_access_target_sa is None or not self._service_security_unlocked:
            message = f"Для записи откройте Security Access 0x27 для узла 0x{target_sa:02X}."
            self.infoMessage.emit("Калибровка", message)
            self._append_log(f"Калибровка: {operation_label} не отправлена. {message}", RowColor.yellow)
            return False

        access_target_sa = int(self._service_access_target_sa) & 0xFF
        if access_target_sa != target_sa:
            message = (
                f"Security Access открыт для узла 0x{access_target_sa:02X}, "
                f"а калибровка настроена на узел 0x{target_sa:02X}."
            )
            self.infoMessage.emit("Калибровка", message)
            self._append_log(f"Калибровка: {operation_label} не отправлена. {message}", RowColor.yellow)
            return False

        self._configure_calibration_uds_services()
        return True

    def _reset_calibration_sequence_state(self):
        if self._calibration_sequence_delay_timer.isActive():
            self._calibration_sequence_delay_timer.stop()
        if self._calibration_sequence_timeout_timer.isActive():
            self._calibration_sequence_timeout_timer.stop()
        self._calibration_sequence_next_action = ""
        self._calibration_sequence_waiting_action = ""

    def _schedule_calibration_sequence_action(self, action: str, delay_ms: int | None = None):
        self._calibration_sequence_next_action = str(action or "")
        delay = self._calibration_sequence_delay_ms if delay_ms is None else delay_ms
        self._calibration_sequence_delay_timer.start(max(0, int(delay)))

    def _start_calibration_sequence_wait(self, action: str):
        self._calibration_sequence_waiting_action = str(action or "")
        if self._calibration_sequence_timeout_timer.isActive():
            self._calibration_sequence_timeout_timer.stop()
        self._calibration_sequence_timeout_timer.start(self._calibration_sequence_timeout_ms)

    def _finish_calibration_sequence_wait(self, expected_action: str | None = None) -> bool:
        current_action = str(self._calibration_sequence_waiting_action or "")
        if expected_action is not None and current_action != str(expected_action):
            return False
        if self._calibration_sequence_timeout_timer.isActive():
            self._calibration_sequence_timeout_timer.stop()
        self._calibration_sequence_waiting_action = ""
        return True

    def _set_calibration_service_access_state(
        self,
        *,
        busy: bool,
        pending_action: str = "",
        unlocked: bool = False,
        target_sa: int | None = None,
        status: str | None = None,
    ):
        self._service_access_busy = bool(busy)
        self._service_access_pending_action = str(pending_action or "")
        self._service_security_unlocked = bool(unlocked)
        self._service_access_target_sa = None if target_sa is None else (int(target_sa) & 0xFF)
        if status is not None:
            self._service_access_status = str(status)
        self.serviceAccessChanged.emit()

    def _fail_calibration_activation(self, message: str):
        self._reset_calibration_sequence_state()
        self._stop_calibration_temp_comp_advanced_read_sequence()
        self._reset_calibration_temp_comp_recommendation_apply_queue()
        self._calibration_runtime_target_sa = None
        self._set_calibration_service_access_state(
            busy=False,
            pending_action="",
            unlocked=False,
            target_sa=None,
            status=message,
        )
        self._calibration_waiting_session = False
        self._calibration_session_ready = False
        self._stop_calibration_poll_timer()
        self._calibration_write_verify_pending = {}
        self.calibrationVerificationChanged.emit()
        if self._calibration_active:
            self._calibration_active = False
            self.calibrationStateChanged.emit()
        self._append_log(message, RowColor.red)
        self._recompute_calibration_wizard_state()

    def _invalidate_calibration_session_after_nrc(self, message: str):
        """Цель функции в сбросе локального состояния калибровки после потери сессии на МК, затем она переводит UI в безопасный режим до повторного запуска."""
        self._reset_calibration_sequence_state()
        self._stop_calibration_temp_comp_advanced_read_sequence()
        self._reset_calibration_temp_comp_recommendation_apply_queue()
        self._reset_calibration_temp_comp_k0_air_zero_adjust_state()
        self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
        self._reset_calibration_temp_comp_zero_trim_verify_state()
        self._calibration_runtime_target_sa = None
        self._calibration_waiting_session = False
        self._calibration_session_ready = False
        self._stop_calibration_poll_timer()
        self._calibration_write_verify_pending = {}
        self.calibrationVerificationChanged.emit()
        self._set_calibration_service_access_state(
            busy=False,
            pending_action="",
            unlocked=False,
            target_sa=None,
            status=message,
        )
        if self._calibration_active:
            self._calibration_active = False
            self.calibrationStateChanged.emit()
        self._append_log(message, RowColor.yellow)
        self._recompute_calibration_wizard_state()

    def _finish_calibration_deactivation(self, message: str):
        self._reset_calibration_sequence_state()
        self._stop_calibration_temp_comp_advanced_read_sequence()
        self._reset_calibration_temp_comp_recommendation_apply_queue()
        self._calibration_runtime_target_sa = None
        self._set_calibration_service_access_state(
            busy=False,
            pending_action="",
            unlocked=False,
            target_sa=None,
            status=message,
        )
        self._calibration_waiting_session = False
        self._calibration_session_ready = False
        self._calibration_write_verify_pending = {}
        self._calibration_restore_active = False
        self._calibration_restore_current_did = None
        self.calibrationVerificationChanged.emit()
        self._recompute_calibration_wizard_state()

    def _send_calibration_security_seed_request(self):
        target_sa = self._resolve_calibration_target_sa()
        self._set_calibration_service_access_state(
            busy=True,
            pending_action="calibration_security_seed",
            unlocked=False,
            target_sa=target_sa,
            status=f"Калибровка: запрос seed 0x27 для SA 0x{target_sa:02X}...",
        )
        self._append_log(
            f"Калибровка: автоматический запрос Security Access seed для узла 0x{target_sa:02X}.",
            RowColor.blue,
        )
        self._service_security_access_service.request_seed(self._build_calibration_tx_identifier())
        self._start_calibration_sequence_wait("security_seed")

    def _send_calibration_security_key_request(self):
        target_sa = self._resolve_calibration_target_sa()
        self._set_calibration_service_access_state(
            busy=True,
            pending_action="calibration_security_key",
            unlocked=False,
            target_sa=target_sa,
            status=f"Калибровка: отправка key 0x27 для SA 0x{target_sa:02X}...",
        )
        self._append_log(
            f"Калибровка: отправка Security Access key для узла 0x{target_sa:02X}.",
            RowColor.blue,
        )
        self._service_security_access_service.request_check_key(self._build_calibration_tx_identifier())
        self._start_calibration_sequence_wait("security_key")

    def _send_calibration_initial_read(self, did: int):
        self._configure_calibration_uds_services()
        target_var = UdsData.empty_fuel_tank if int(did) == int(UdsData.empty_fuel_tank.pid) else UdsData.full_fuel_tank
        label = "0%" if int(did) == int(UdsData.empty_fuel_tank.pid) else "100%"
        if not self._calibration_read_service.read_data_by_identifier(self._build_calibration_tx_identifier(), target_var):
            self._fail_calibration_activation(f"Калибровка: не удалось отправить чтение уровня {label}.")
            return
        self._append_log(f"Калибровка: автоматическое чтение сохраненного уровня {label}.", RowColor.blue)
        wait_action = "read_level_0" if int(did) == int(UdsData.empty_fuel_tank.pid) else "read_level_100"
        self._start_calibration_sequence_wait(wait_action)

    def _on_calibration_sequence_delay_timeout(self):
        action = str(self._calibration_sequence_next_action or "")
        self._calibration_sequence_next_action = ""
        if not action:
            return

        if action == "request_security_seed":
            self._send_calibration_security_seed_request()
            return

        if action == "send_security_key":
            self._send_calibration_security_key_request()
            return

        if action == "read_level_0":
            self._send_calibration_initial_read(int(UdsData.empty_fuel_tank.pid))
            return

        if action == "read_level_100":
            self._send_calibration_initial_read(int(UdsData.full_fuel_tank.pid))

    def _on_calibration_sequence_timeout(self):
        action = str(self._calibration_sequence_waiting_action or "")
        if action == "activate_session":
            self._fail_calibration_activation("Калибровка: таймаут ожидания ответа на Extended Session 0x10.")
            return
        if action == "security_seed":
            self._fail_calibration_activation("Калибровка: таймаут ожидания seed на Security Access 0x27.")
            return
        if action == "security_key":
            self._fail_calibration_activation("Калибровка: таймаут ожидания подтверждения key на Security Access 0x27.")
            return
        if action == "read_level_0":
            self._fail_calibration_activation("Калибровка: таймаут чтения сохраненного уровня 0%.")
            return
        if action == "read_level_100":
            self._fail_calibration_activation("Калибровка: таймаут чтения сохраненного уровня 100%.")
            return
        if action == "deactivate_session":
            self._finish_calibration_deactivation("Калибровка: таймаут возврата в default-сессию, локальное состояние сброшено.")

    def _add_calibration_recent_sample(self, value: int):
        now_monotonic = time.monotonic()
        self._calibration_recent_samples.append((now_monotonic, int(value)))
        min_ts = now_monotonic - float(self._calibration_recent_window_sec)
        self._calibration_recent_samples = [
            (sample_ts, sample_value)
            for (sample_ts, sample_value) in self._calibration_recent_samples
            if float(sample_ts) >= min_ts
        ]
        if len(self._calibration_recent_samples) > 100:
            self._calibration_recent_samples = self._calibration_recent_samples[-100:]
        # Автообновление стабильного значения на каждом новом семпле.
        self._recompute_calibration_stable_capture()

    def _recompute_calibration_stable_capture(self) -> tuple[int | None, int]:
        now_monotonic = time.monotonic()
        valid_samples = [
            int(sample_value)
            for (sample_ts, sample_value) in self._calibration_recent_samples
            if (now_monotonic - float(sample_ts)) <= float(self._calibration_recent_window_sec)
        ]

        if len(valid_samples) < 2:
            if self._calibration_captured_available:
                self._calibration_captured_available = False
                self.calibrationValuesChanged.emit()
            return None, len(valid_samples)

        avg_value = sum(valid_samples) / float(len(valid_samples))
        captured = int(round(avg_value))
        changed = (not self._calibration_captured_available) or (self._calibration_captured_level != captured)
        self._calibration_captured_level = captured
        self._calibration_captured_available = True
        if changed:
            self.calibrationValuesChanged.emit()
        return captured, len(valid_samples)

    @staticmethod
    def _decode_signed_value(raw_value: int, bits: int) -> int:
        """Цель функции в корректной интерпретации знаковых DID, затем она выполняет sign-extension по числу бит."""
        width = max(1, int(bits))
        mask = (1 << width) - 1
        sign_bit = 1 << (width - 1)
        value = int(raw_value) & mask
        if value & sign_bit:
            return value - (1 << width)
        return value

    @staticmethod
    def _c_trunc_div(numerator: int, denominator: int) -> int:
        """Цель функции в повторении поведения C-деления, затем она делит с усечением к нулю."""
        if int(denominator) == 0:
            raise ZeroDivisionError("Деление на ноль недопустимо.")

        quotient = abs(int(numerator)) // abs(int(denominator))
        if (int(numerator) < 0) ^ (int(denominator) < 0):
            return -quotient
        return quotient

    def _set_calibration_temp_comp_zero_trim_last_report(
        self,
        *,
        old_zero_trim: int | None,
        new_zero_trim: int | None,
        residual_x10: int | None,
        status_text: str,
        write_csv: bool = False,
    ):
        """Цель функции в формировании краткой сводки подгонки zero trim, затем она сохраняет человеко-понятный отчет для оператора."""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        node_sa = int(self._resolve_calibration_target_sa()) & 0xFF
        temperature_x10 = self._calibration_temp_comp_last_temperature_x10
        if temperature_x10 is None:
            temperature_text = "н/д"
        else:
            temperature_text = f"{float(int(temperature_x10)) / 10.0:+.1f}°C"

        old_text = "н/д" if old_zero_trim is None else str(int(old_zero_trim))
        new_text = "н/д" if new_zero_trim is None else str(int(new_zero_trim))
        residual_text = "н/д" if residual_x10 is None else f"{float(int(residual_x10)) / 10.0:+.1f}%"
        status_value = str(status_text or "").strip()
        if not status_value:
            status_value = "выполнено"

        self._calibration_temp_comp_zero_trim_last_report = (
            f"{timestamp} | SA 0x{node_sa:02X} | T={temperature_text} | "
            f"trim {old_text}->{new_text} | остаток {residual_text} | {status_value}."
        )
        if bool(write_csv):
            try:
                csv_path = self._append_calibration_temp_comp_zero_trim_report_csv_row(
                    timestamp_text=timestamp,
                    node_sa=node_sa,
                    temperature_text=temperature_text,
                    old_zero_trim_text=old_text,
                    new_zero_trim_text=new_text,
                    residual_text=residual_text,
                    status_text=status_value,
                )
                if str(self._calibration_temp_comp_zero_trim_csv_log_path) != str(csv_path):
                    self._calibration_temp_comp_zero_trim_csv_log_path = str(csv_path)
                    self._append_log(
                        f"Калибровка: CSV-лог операций zero trim записывается в {csv_path}.",
                        RowColor.blue,
                    )
            except Exception as exc:
                self._append_log(
                    f"Калибровка: не удалось записать CSV-лог zero trim: {str(exc)}",
                    RowColor.yellow,
                )

    def _resolve_calibration_temp_comp_zero_trim_report_directory(self) -> Path:
        """Цель функции в выборе каталога логов подгонки zero trim, затем она возвращает рабочий путь для CSV-отчета операций."""
        session_dir = self._collector_session_dir
        if session_dir is not None:
            try:
                resolved_session_dir = Path(session_dir).expanduser().resolve()
                resolved_session_dir.mkdir(parents=True, exist_ok=True)
                return resolved_session_dir
            except Exception:
                pass

        try:
            output_dir = Path(str(self._collector_output_directory)).expanduser().resolve()
            output_dir.mkdir(parents=True, exist_ok=True)
            return output_dir
        except Exception:
            fallback_dir = Path(self._project_root_directory) / "logs"
            fallback_dir.mkdir(parents=True, exist_ok=True)
            return fallback_dir

    def _append_calibration_temp_comp_zero_trim_report_csv_row(
        self,
        *,
        timestamp_text: str,
        node_sa: int,
        temperature_text: str,
        old_zero_trim_text: str,
        new_zero_trim_text: str,
        residual_text: str,
        status_text: str,
    ) -> str:
        """Цель функции в накоплении CSV-истории подгонки zero trim, затем она добавляет строку операции в UTF-8 файл сессии."""
        report_dir = self._resolve_calibration_temp_comp_zero_trim_report_directory()
        csv_path = report_dir / "calibration_zero_trim_session.csv"
        need_header = not csv_path.exists()
        with csv_path.open("a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file, delimiter=";")
            if need_header:
                writer.writerow(
                    (
                        "Время",
                        "Узел",
                        "Температура",
                        "Zero trim старый",
                        "Zero trim новый",
                        "Остаток",
                        "Статус",
                    )
                )
            writer.writerow(
                (
                    str(timestamp_text),
                    f"0x{int(node_sa) & 0xFF:02X}",
                    str(temperature_text),
                    str(old_zero_trim_text),
                    str(new_zero_trim_text),
                    str(residual_text),
                    str(status_text),
                )
            )
        return str(csv_path)

    @classmethod
    def _saturate_int16(cls, value: int) -> int:
        """Цель функции в защите от переполнения DID int16, затем она ограничивает значение диапазоном -32768..32767."""
        return max(cls._INT16_MIN, min(cls._INT16_MAX, int(value)))

    @classmethod
    def _calculate_zero_trim_adjustment(
        cls,
        *,
        span_count: int,
        current_level_x10: int,
        current_zero_trim: int,
    ) -> tuple[int, int, int]:
        """Цель функции в вычислении коррекции 0%-смещения, затем она возвращает delta/target/residual в формате int16 + x10%."""
        span = int(span_count)
        if span <= 0:
            raise ValueError("Некорректный span для расчета zero trim.")

        normalized_level_x10 = max(-1000, min(1000, int(current_level_x10)))
        delta_zero_trim = cls._c_trunc_div((-normalized_level_x10) * span, 1000)
        target_zero_trim = cls._saturate_int16(int(current_zero_trim) + int(delta_zero_trim))
        applied_delta = int(target_zero_trim) - int(current_zero_trim)
        residual_x10 = int(normalized_level_x10) + cls._c_trunc_div(int(applied_delta) * 1000, span)
        return int(delta_zero_trim), int(target_zero_trim), int(residual_x10)

    @staticmethod
    def _classify_zero_trim_verification_result(
        *,
        residual_x10: int,
        tolerance_x10: int,
        repeat_threshold_x10: int,
    ) -> str:
        """Цель функции в единой классификации автопроверки zero trim, затем она возвращает статус success/repeat/mechanics."""
        abs_level_x10 = abs(int(residual_x10))
        normalized_tolerance_x10 = max(0, int(tolerance_x10))
        normalized_repeat_threshold_x10 = max(normalized_tolerance_x10, int(repeat_threshold_x10))
        if abs_level_x10 <= normalized_tolerance_x10:
            return "success"
        if abs_level_x10 <= normalized_repeat_threshold_x10:
            return "repeat"
        return "mechanics"

    @classmethod
    def _apply_temperature_compensation_model(
        cls,
        raw_period: int,
        temperature_x10: int,
        k1_x100: int,
        k0_count: int = 0,
    ) -> int:
        """Цель функции в точном повторении алгоритма МК, затем она считает Pcomp по формуле K1+K0 из fuel.c."""
        compensated_period = int(raw_period)
        k1_coefficient = int(k1_x100)
        if k1_coefficient != 0:
            d_temperature_x10 = int(temperature_x10) - int(cls._FUEL_TEMP_COMP_REF_X10)
            delta_period = cls._c_trunc_div(k1_coefficient * d_temperature_x10, 1000)
            compensated_period -= delta_period

        compensated_period += int(k0_count)

        if compensated_period < 0:
            compensated_period = 0
        if compensated_period > 0xFFFF:
            compensated_period = 0xFFFF

        return int(compensated_period)

    @classmethod
    def _apply_temperature_compensation_model_precise(
        cls,
        raw_period: int | float,
        temperature_x10: int,
        k1_x100: int,
        k0_count: int = 0,
    ) -> float:
        """Цель функции в плавном расчете компенсации для аналитики, затем она считает Pcomp без целочисленного усечения для наглядного графика."""
        compensated_period = float(raw_period)
        k1_coefficient = float(k1_x100)
        if k1_coefficient != 0.0:
            d_temperature_x10 = float(int(temperature_x10) - int(cls._FUEL_TEMP_COMP_REF_X10))
            delta_period = (k1_coefficient * d_temperature_x10) / 1000.0
            compensated_period -= delta_period

        compensated_period += float(k0_count)

        if compensated_period < 0.0:
            compensated_period = 0.0
        if compensated_period > 65535.0:
            compensated_period = 65535.0

        return float(compensated_period)

    @staticmethod
    def _temp_comp_segment_table_is_zero(table_values: list[int]) -> bool:
        """Цель функции в проверке заполненности таблицы K1, затем она возвращает True только если все значения нулевые."""
        return all(int(value) == 0 for value in table_values)

    @staticmethod
    def _temp_comp_segment_table_is_linear(table_values: list[int], linear_k1_x100: int) -> bool:
        """Цель функции в определении «линейной» таблицы, затем она проверяет равенство всех сегментов базовому K1."""
        target_value = int(linear_k1_x100)
        if len(table_values) <= 0:
            return True
        return all(int(value) == target_value for value in table_values)

    @classmethod
    def _temp_comp_values_have_informative_segments(
        cls,
        values: dict[str, int | None],
        *,
        mode_value: int,
        linear_k1_x100: int,
    ) -> bool:
        """Цель функции в оценке полезности таблиц для preview, затем она подтверждает, что mode 1/2 задан не вырожденными сегментами."""
        normalized_mode = int(mode_value)
        cooling_table, heating_table = cls._temp_comp_build_segment_tables_from_values(values)
        cooling_zero = cls._temp_comp_segment_table_is_zero(cooling_table)
        heating_zero = cls._temp_comp_segment_table_is_zero(heating_table)
        cooling_linear = cls._temp_comp_segment_table_is_linear(cooling_table, int(linear_k1_x100))
        heating_linear = cls._temp_comp_segment_table_is_linear(heating_table, int(linear_k1_x100))

        if normalized_mode == int(cls._TEMP_COMP_MODE_SEGMENTED):
            return (not cooling_zero) and (not cooling_linear)
        if normalized_mode == int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
            if cooling_zero and heating_zero:
                return False
            if cooling_linear and heating_linear:
                return False
            return True
        return True

    @classmethod
    def _temp_comp_get_mode_from_values(cls, advanced_values: dict[str, int | None]) -> int:
        """Цель функции в чтении рабочего режима, затем она ограничивает значение диапазоном 0..2."""
        raw_mode = advanced_values.get("mode")
        if raw_mode is None:
            return int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)
        mode_value = int(raw_mode)
        if mode_value < int(cls._TEMP_COMP_MODE_SINGLE_LINEAR) or mode_value > int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
            return int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)
        return int(mode_value)

    @classmethod
    def _temp_comp_get_hysteresis_x10_from_values(cls, advanced_values: dict[str, int | None]) -> int:
        """Цель функции в чтении гистерезиса ветки, затем она подставляет безопасный default при некорректном значении."""
        raw_hysteresis = advanced_values.get("dir_hyst_x10")
        if raw_hysteresis is None:
            return int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)
        hysteresis_x10 = int(raw_hysteresis)
        if hysteresis_x10 < int(cls._TEMP_COMP_DIR_HYST_MIN_X10):
            return int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)
        return int(hysteresis_x10)

    @classmethod
    def _temp_comp_get_borders_x10_from_values(cls, advanced_values: dict[str, int | None]) -> tuple[int, int, int, int]:
        """Цель функции в чтении T1..T4 из параметров, затем она гарантирует строго возрастающие границы сегментов."""
        defaults = tuple(int(value) for value in cls._TEMP_COMP_DEFAULT_SEG_BORDERS_X10)
        keys = ("seg_t1_x10", "seg_t2_x10", "seg_t3_x10", "seg_t4_x10")
        values: list[int] = []
        for index, key in enumerate(keys):
            raw_value = advanced_values.get(key)
            if raw_value is None:
                return defaults
            values.append(int(raw_value))

        if not (values[0] < values[1] < values[2] < values[3]):
            return defaults

        return (int(values[0]), int(values[1]), int(values[2]), int(values[3]))

    @staticmethod
    def _temp_comp_build_segment_tables_from_values(advanced_values: dict[str, int | None]) -> tuple[list[int], list[int]]:
        """Цель функции в сборке таблиц K1 по сегментам, затем она возвращает списки для охлаждения и нагрева."""
        cooling_table: list[int] = []
        heating_table: list[int] = []
        for segment_index in range(1, 6):
            cooling_key = f"k1_cool_seg{segment_index}_x100"
            heating_key = f"k1_heat_seg{segment_index}_x100"
            cooling_raw = advanced_values.get(cooling_key)
            heating_raw = advanced_values.get(heating_key)
            cooling_table.append(0 if cooling_raw is None else int(cooling_raw))
            heating_table.append(0 if heating_raw is None else int(heating_raw))
        return cooling_table, heating_table

    @classmethod
    def _apply_temp_comp_mode_sequence(
        cls,
        samples_in_time_order: list[dict[str, object]],
        *,
        linear_k1_x100: int,
        k0_count: int,
        advanced_values: dict[str, int | None],
    ) -> dict[int, float]:
        """Цель функции в повторении выбора K1 как на МК, затем она считает Pcomp с учетом mode/segment/heat-cool."""
        result_by_sample_id: dict[int, float] = {}
        if len(samples_in_time_order) <= 0:
            return result_by_sample_id

        mode = cls._temp_comp_get_mode_from_values(advanced_values)
        borders_x10 = cls._temp_comp_get_borders_x10_from_values(advanced_values)
        hysteresis_x10 = cls._temp_comp_get_hysteresis_x10_from_values(advanced_values)
        cooling_table, heating_table = cls._temp_comp_build_segment_tables_from_values(advanced_values)
        cooling_table_zero = cls._temp_comp_segment_table_is_zero(cooling_table)
        heating_table_zero = cls._temp_comp_segment_table_is_zero(heating_table)

        last_temperature_x10 = int(samples_in_time_order[0].get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10))
        trend_state = int(cls._TEMP_COMP_TREND_STABLE)

        for sample in samples_in_time_order:
            raw_period = float(sample.get("period", 0.0))
            temperature_x10 = int(sample.get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10))
            selected_k1_x100 = int(linear_k1_x100)

            if mode == int(cls._TEMP_COMP_MODE_SEGMENTED):
                if not cooling_table_zero:
                    segment_index = cls._temp_comp_segment_index(temperature_x10, borders_x10)
                    selected_k1_x100 = int(cooling_table[int(segment_index)])
            elif mode == int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
                if temperature_x10 > (last_temperature_x10 + hysteresis_x10):
                    trend_state = int(cls._TEMP_COMP_TREND_HEATING)
                elif temperature_x10 < (last_temperature_x10 - hysteresis_x10):
                    trend_state = int(cls._TEMP_COMP_TREND_COOLING)
                last_temperature_x10 = int(temperature_x10)

                heating_branch = int(trend_state) == int(cls._TEMP_COMP_TREND_HEATING)
                active_table_zero = bool(heating_table_zero) if heating_branch else bool(cooling_table_zero)
                opposite_table_zero = bool(cooling_table_zero) if heating_branch else bool(heating_table_zero)
                if active_table_zero and opposite_table_zero:
                    selected_k1_x100 = int(linear_k1_x100)
                else:
                    if active_table_zero:
                        heating_branch = not heating_branch
                    segment_index = cls._temp_comp_segment_index(temperature_x10, borders_x10)
                    if heating_branch:
                        selected_k1_x100 = int(heating_table[int(segment_index)])
                    else:
                        selected_k1_x100 = int(cooling_table[int(segment_index)])
            else:
                selected_k1_x100 = int(linear_k1_x100)

            if mode != int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
                last_temperature_x10 = int(temperature_x10)
                trend_state = int(cls._TEMP_COMP_TREND_STABLE)

            compensated_period = cls._apply_temperature_compensation_model_precise(
                raw_period,
                int(temperature_x10),
                int(selected_k1_x100),
                int(k0_count),
            )
            result_by_sample_id[id(sample)] = float(compensated_period)

        return result_by_sample_id

    @classmethod
    def _build_effective_temp_comp_values(
        cls,
        current_values: dict[str, int | None],
        recommended_values: dict[str, int | None],
    ) -> dict[str, int | None]:
        """Цель функции в формировании рабочего набора параметров, затем она дополняет отсутствующие текущие значения рекомендациями."""
        effective_values: dict[str, int | None] = {}
        for field in cls._temp_comp_advanced_fields():
            key = str(field.get("key", ""))
            if not key:
                continue
            current_value = current_values.get(key)
            if current_value is not None:
                effective_values[key] = int(current_value)
                continue
            recommended_value = recommended_values.get(key)
            if recommended_value is not None:
                effective_values[key] = int(recommended_value)
            else:
                effective_values[key] = None
        return effective_values

    def _period_to_level_percent(self, period: int | float) -> float | None:
        """Цель функции в пересчете периода в проценты, затем она применяет текущие границы empty/full из калибровки."""
        empty_period = int(self._calibration_level_0)
        full_period = int(self._calibration_level_100)
        span = full_period - empty_period
        if span <= 0:
            return None
        return ((float(period) - float(empty_period)) * 100.0) / float(span)

    @staticmethod
    def _linear_regression_slope(x_values: list[float], y_values: list[float]) -> float | None:
        """Цель функции в оценке температурного дрейфа, затем она возвращает наклон линейной регрессии dy/dx."""
        if len(x_values) != len(y_values):
            return None
        if len(x_values) < 2:
            return None

        x_mean = sum(float(x) for x in x_values) / float(len(x_values))
        y_mean = sum(float(y) for y in y_values) / float(len(y_values))

        ss_x = sum((float(x) - x_mean) ** 2 for x in x_values)
        if ss_x <= 0.0:
            return None

        cov_xy = sum((float(x) - x_mean) * (float(y) - y_mean) for (x, y) in zip(x_values, y_values))
        return float(cov_xy / ss_x)

    @staticmethod
    def _calc_reduction_percent(before: float | None, after: float | None) -> float | None:
        """Цель функции в расчете эффекта компенсации, затем она возвращает процент снижения модуля дрейфа."""
        if before is None or after is None:
            return None
        if float(before) == 0.0:
            return 0.0
        return (1.0 - abs(float(after)) / abs(float(before))) * 100.0

    @staticmethod
    def _calc_percentile_abs(values: list[float], percentile: float) -> float | None:
        """Цель функции в расчете устойчивой оценки ошибки, затем она возвращает перцентиль абсолютного отклонения."""
        if len(values) <= 0:
            return None
        sorted_abs = sorted(abs(float(value)) for value in values)
        if len(sorted_abs) <= 0:
            return None
        ratio = max(0.0, min(1.0, float(percentile)))
        index = int(round(ratio * float(len(sorted_abs) - 1)))
        index = max(0, min(index, len(sorted_abs) - 1))
        return float(sorted_abs[index])

    @staticmethod
    def _calc_level_error_metrics(level_values: list[float]) -> dict[str, float | tuple[float, float]] | None:
        """Цель функции в человеко-понятной оценке компенсации, затем она считает коридор, максимум и P95 ошибки уровня."""
        if len(level_values) <= 0:
            return None

        min_level = min(float(value) for value in level_values)
        max_level = max(float(value) for value in level_values)
        max_abs = max(abs(float(value)) for value in level_values)
        p95_abs = AppControllerCalibrationMixin._calc_percentile_abs(level_values, 0.95)
        if p95_abs is None:
            return None

        return {
            "range": (float(min_level), float(max_level)),
            "max_abs": float(max_abs),
            "p95_abs": float(p95_abs),
        }

    @staticmethod
    def _find_max_abs_level_error(level_values: list[float]) -> tuple[int | None, float | None]:
        """Цель функции в поиске точки наихудшей ошибки, затем она возвращает индекс и signed-значение уровня."""
        if len(level_values) <= 0:
            return None, None
        max_index = max(range(len(level_values)), key=lambda idx: abs(float(level_values[idx])))
        return int(max_index), float(level_values[max_index])

    @staticmethod
    def _calc_quantile(values: list[float], ratio: float) -> float | None:
        """Цель функции в вычислении квантили набора, затем она возвращает интерполированное значение для заданной доли."""
        if len(values) <= 0:
            return None
        ordered = sorted(float(value) for value in values)
        if len(ordered) == 1:
            return float(ordered[0])

        bounded_ratio = max(0.0, min(1.0, float(ratio)))
        position = bounded_ratio * float(len(ordered) - 1)
        left_index = int(math.floor(position))
        right_index = int(math.ceil(position))
        if left_index == right_index:
            return float(ordered[left_index])

        weight = float(position - float(left_index))
        left_value = float(ordered[left_index])
        right_value = float(ordered[right_index])
        return (left_value * (1.0 - weight)) + (right_value * weight)

    @staticmethod
    def _temp_comp_segment_index(temperature_x10: int, borders_x10: tuple[int, int, int, int]) -> int:
        """Цель функции в определении номера сегмента S1..S5, затем она выбирает интервал по границам T1..T4."""
        value = int(temperature_x10)
        if value < int(borders_x10[0]):
            return 0
        if value < int(borders_x10[1]):
            return 1
        if value < int(borders_x10[2]):
            return 2
        if value < int(borders_x10[3]):
            return 3
        return 4

    @classmethod
    def _build_temp_comp_segment_borders_x10(cls, temperature_x10_values: list[int]) -> tuple[int, int, int, int]:
        """Цель функции в построении валидных T1..T4, затем она подбирает строгие границы по распределению температур."""
        if len(temperature_x10_values) < int(cls._TEMP_COMP_SEGMENT_COUNT):
            return tuple(int(value) for value in cls._TEMP_COMP_DEFAULT_SEG_BORDERS_X10)

        ordered = sorted(int(value) for value in temperature_x10_values)
        q20 = cls._calc_quantile([float(value) for value in ordered], 0.2)
        q40 = cls._calc_quantile([float(value) for value in ordered], 0.4)
        q60 = cls._calc_quantile([float(value) for value in ordered], 0.6)
        q80 = cls._calc_quantile([float(value) for value in ordered], 0.8)

        if q20 is None or q40 is None or q60 is None or q80 is None:
            return tuple(int(value) for value in cls._TEMP_COMP_DEFAULT_SEG_BORDERS_X10)

        borders = [int(round(q20)), int(round(q40)), int(round(q60)), int(round(q80))]
        if not (borders[0] < borders[1] < borders[2] < borders[3]):
            min_temp = int(ordered[0])
            max_temp = int(ordered[-1])
            span = int(max_temp - min_temp)
            if span < int(cls._TEMP_COMP_SEGMENT_COUNT):
                return tuple(int(value) for value in cls._TEMP_COMP_DEFAULT_SEG_BORDERS_X10)
            step = max(1, int(round(float(span) / float(cls._TEMP_COMP_SEGMENT_COUNT))))
            borders = [min_temp + step, min_temp + (2 * step), min_temp + (3 * step), min_temp + (4 * step)]

        for index in range(1, len(borders)):
            if int(borders[index]) <= int(borders[index - 1]):
                borders[index] = int(borders[index - 1]) + 1

        return (int(borders[0]), int(borders[1]), int(borders[2]), int(borders[3]))

    @classmethod
    def _recommend_temp_comp_dir_hyst_x10(cls, samples: list[dict[str, object]]) -> int:
        """Цель функции в рекомендации гистерезиса ветки, затем она подбирает порог на основе реального шага температуры."""
        if len(samples) < 2:
            return int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)

        deltas_x10: list[int] = []
        previous_temp_x10 = int(samples[0].get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10))
        for sample in samples[1:]:
            current_temp_x10 = int(sample.get("temperature_x10", previous_temp_x10))
            delta_abs_x10 = abs(int(current_temp_x10) - int(previous_temp_x10))
            if delta_abs_x10 > 0:
                deltas_x10.append(int(delta_abs_x10))
            previous_temp_x10 = int(current_temp_x10)

        if len(deltas_x10) <= 0:
            return int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)

        median_step_x10 = cls._calc_quantile([float(value) for value in deltas_x10], 0.5)
        if median_step_x10 is None:
            return int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)

        suggested_hyst_x10 = int(round(float(median_step_x10) * 2.0))
        suggested_hyst_x10 = max(int(cls._TEMP_COMP_DIR_HYST_MIN_X10), min(50, int(suggested_hyst_x10)))
        return int(suggested_hyst_x10)

    @classmethod
    def _build_temp_comp_trend_sequence(
        cls,
        samples: list[dict[str, object]],
        hysteresis_x10: int,
    ) -> list[int]:
        """Цель функции в повторении логики МК heat/cool, затем она определяет тренд температуры для каждой точки."""
        if len(samples) <= 0:
            return []

        active_hysteresis_x10 = int(hysteresis_x10)
        if active_hysteresis_x10 < int(cls._TEMP_COMP_DIR_HYST_MIN_X10):
            active_hysteresis_x10 = int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)

        trend = 0
        previous_temp_x10 = int(samples[0].get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10))
        sequence: list[int] = []
        for sample in samples:
            current_temp_x10 = int(sample.get("temperature_x10", previous_temp_x10))
            if current_temp_x10 > (previous_temp_x10 + active_hysteresis_x10):
                trend = 1
            elif current_temp_x10 < (previous_temp_x10 - active_hysteresis_x10):
                trend = -1
            previous_temp_x10 = int(current_temp_x10)
            sequence.append(int(trend))
        return sequence

    @classmethod
    def _temp_comp_slope_for_indices(
        cls,
        samples: list[dict[str, object]],
        sample_indices: list[int],
    ) -> float | None:
        """Цель функции в оценке локального K1, затем она считает наклон period(T) по выбранным индексам."""
        if len(sample_indices) < 2:
            return None

        temperatures: list[float] = []
        periods: list[float] = []
        for index in sample_indices:
            if int(index) < 0 or int(index) >= len(samples):
                continue
            sample = samples[int(index)]
            temperatures.append(float(sample.get("temperature_c", float(int(sample.get("temperature_x10", 0))) / 10.0)))
            periods.append(float(sample.get("period", 0.0)))

        return cls._linear_regression_slope(temperatures, periods)

    @classmethod
    def _temp_comp_slope_for_indices_robust(
        cls,
        samples: list[dict[str, object]],
        sample_indices: list[int],
    ) -> float | None:
        """Цель функции в устойчивой оценке локального K1, затем она отбрасывает узкие/короткие выборки и возвращает slope только для валидного сегмента."""
        if len(sample_indices) < int(cls._TEMP_COMP_REC_MIN_SEG_POINTS):
            return None

        temp_values_x10: list[int] = []
        for index in sample_indices:
            if int(index) < 0 or int(index) >= len(samples):
                continue
            sample = samples[int(index)]
            temp_values_x10.append(int(sample.get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10)))

        if len(temp_values_x10) < int(cls._TEMP_COMP_REC_MIN_SEG_POINTS):
            return None
        if (max(temp_values_x10) - min(temp_values_x10)) < int(cls._TEMP_COMP_REC_MIN_SEG_SPAN_X10):
            return None

        return cls._temp_comp_slope_for_indices(samples, sample_indices)

    @staticmethod
    def _temp_comp_smooth_segment_slopes(slope_values: list[float]) -> list[float]:
        """Цель функции в сглаживании соседних сегментов, затем она уменьшает случайные перепады K1 между S1..S5."""
        if len(slope_values) <= 1:
            return list(slope_values)

        smoothed: list[float] = []
        for index, value in enumerate(slope_values):
            left = float(slope_values[index - 1]) if index > 0 else float(value)
            right = float(slope_values[index + 1]) if (index + 1) < len(slope_values) else float(value)
            center = float(value)
            smoothed_value = (left * 0.25) + (center * 0.5) + (right * 0.25)
            smoothed.append(float(smoothed_value))
        return smoothed

    @classmethod
    def _temp_comp_candidate_mode_score(
        cls,
        samples_in_time_order: list[dict[str, object]],
        *,
        linear_k1_x100: int,
        advanced_values: dict[str, int],
    ) -> float:
        """Цель функции в сравнении режимов компенсации, затем она рассчитывает score по остаточному дрейфу и штрафу за сложность режима."""
        if len(samples_in_time_order) < 2:
            return float("inf")

        period_by_sample_id = cls._apply_temp_comp_mode_sequence(
            samples_in_time_order,
            linear_k1_x100=int(linear_k1_x100),
            k0_count=0,
            advanced_values={str(key): int(value) for key, value in advanced_values.items()},
        )
        temperatures: list[float] = []
        compensated_periods: list[float] = []
        for sample in samples_in_time_order:
            temperatures.append(float(sample.get("temperature_c", 0.0)))
            compensated_periods.append(float(period_by_sample_id.get(id(sample), float(sample.get("period", 0.0)))))

        global_slope = cls._linear_regression_slope(temperatures, compensated_periods)
        if global_slope is None:
            return float("inf")

        mode_value = cls._temp_comp_get_mode_from_values({str(key): int(value) for key, value in advanced_values.items()})
        complexity_penalty = 0.0
        if mode_value == int(cls._TEMP_COMP_MODE_SEGMENTED):
            complexity_penalty = 0.05
        elif mode_value == int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
            complexity_penalty = 0.12

        return abs(float(global_slope)) + float(complexity_penalty)

    @classmethod
    def _build_temp_comp_advanced_recommendations(
        cls,
        samples: list[dict[str, object]],
        fallback_linear_k1_x100: int,
        forced_mode: int | None = None,
    ) -> dict[str, int]:
        """Цель функции в автоподборе DID 0x001D..0x002C, затем она возвращает рекомендованные mode/hyst/T/K1-сегменты."""
        recommendations: dict[str, int] = {}
        forced_mode_value: int | None = None
        if forced_mode is not None:
            candidate_mode = int(forced_mode)
            if candidate_mode in (
                int(cls._TEMP_COMP_MODE_SINGLE_LINEAR),
                int(cls._TEMP_COMP_MODE_SEGMENTED),
                int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL),
            ):
                forced_mode_value = int(candidate_mode)

        safe_linear_k1_x100 = cls._saturate_int16(int(fallback_linear_k1_x100))
        recommendations["mode"] = 0
        recommendations["dir_hyst_x10"] = int(cls._TEMP_COMP_DIR_HYST_DEFAULT_X10)
        default_borders = tuple(int(value) for value in cls._TEMP_COMP_DEFAULT_SEG_BORDERS_X10)
        recommendations["seg_t1_x10"] = int(default_borders[0])
        recommendations["seg_t2_x10"] = int(default_borders[1])
        recommendations["seg_t3_x10"] = int(default_borders[2])
        recommendations["seg_t4_x10"] = int(default_borders[3])
        for segment_index in range(1, int(cls._TEMP_COMP_SEGMENT_COUNT) + 1):
            recommendations[f"k1_cool_seg{segment_index}_x100"] = int(safe_linear_k1_x100)
            recommendations[f"k1_heat_seg{segment_index}_x100"] = int(safe_linear_k1_x100)

        if len(samples) < 2:
            return recommendations

        ordered_by_time = sorted(
            samples,
            key=lambda item: float(item.get("timestamp", 0.0)),
        )
        temperatures_x10 = [int(sample.get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10)) for sample in ordered_by_time]
        if len(temperatures_x10) < 2:
            return recommendations

        borders_x10 = cls._build_temp_comp_segment_borders_x10(temperatures_x10)
        recommendations["seg_t1_x10"] = int(borders_x10[0])
        recommendations["seg_t2_x10"] = int(borders_x10[1])
        recommendations["seg_t3_x10"] = int(borders_x10[2])
        recommendations["seg_t4_x10"] = int(borders_x10[3])

        hysteresis_x10 = cls._recommend_temp_comp_dir_hyst_x10(ordered_by_time)
        recommendations["dir_hyst_x10"] = int(hysteresis_x10)
        trends = cls._build_temp_comp_trend_sequence(ordered_by_time, int(hysteresis_x10))

        cool_segment_indices: list[list[int]] = [[] for _ in range(int(cls._TEMP_COMP_SEGMENT_COUNT))]
        heat_segment_indices: list[list[int]] = [[] for _ in range(int(cls._TEMP_COMP_SEGMENT_COUNT))]
        cool_all_indices: list[int] = []
        heat_all_indices: list[int] = []

        for index, sample in enumerate(ordered_by_time):
            segment_index = cls._temp_comp_segment_index(
                int(sample.get("temperature_x10", cls._FUEL_TEMP_COMP_REF_X10)),
                borders_x10,
            )
            trend_value = int(trends[index]) if index < len(trends) else 0
            if trend_value < 0:
                cool_segment_indices[int(segment_index)].append(int(index))
                cool_all_indices.append(int(index))
            elif trend_value > 0:
                heat_segment_indices[int(segment_index)].append(int(index))
                heat_all_indices.append(int(index))

        cool_global_slope = cls._temp_comp_slope_for_indices(ordered_by_time, cool_all_indices)
        heat_global_slope = cls._temp_comp_slope_for_indices(ordered_by_time, heat_all_indices)
        all_indices = [int(index) for index in range(len(ordered_by_time))]
        common_global_slope = cls._temp_comp_slope_for_indices(ordered_by_time, all_indices)
        if common_global_slope is None:
            common_global_slope = float(safe_linear_k1_x100) / 100.0

        cooling_slopes: list[float] = []
        heating_slopes: list[float] = []
        for segment_index in range(1, int(cls._TEMP_COMP_SEGMENT_COUNT) + 1):
            segment_zero_based = int(segment_index - 1)

            cooling_slope = cls._temp_comp_slope_for_indices_robust(ordered_by_time, cool_segment_indices[segment_zero_based])
            if cooling_slope is None:
                cooling_slope = cool_global_slope
            if cooling_slope is None:
                cooling_slope = heat_global_slope
            if cooling_slope is None:
                cooling_slope = common_global_slope
            if cool_global_slope is not None:
                cooling_slope = (float(cooling_slope) * 0.75) + (float(cool_global_slope) * 0.25)
            else:
                cooling_slope = float(cooling_slope)
            cooling_slopes.append(float(cooling_slope))

            heating_slope = cls._temp_comp_slope_for_indices_robust(ordered_by_time, heat_segment_indices[segment_zero_based])
            if heating_slope is None:
                heating_slope = heat_global_slope
            if heating_slope is None:
                heating_slope = cool_global_slope
            if heating_slope is None:
                heating_slope = common_global_slope
            if heat_global_slope is not None:
                heating_slope = (float(heating_slope) * 0.75) + (float(heat_global_slope) * 0.25)
            else:
                heating_slope = float(heating_slope)
            heating_slopes.append(float(heating_slope))

        cooling_slopes = cls._temp_comp_smooth_segment_slopes(cooling_slopes)
        heating_slopes = cls._temp_comp_smooth_segment_slopes(heating_slopes)
        for segment_index in range(1, int(cls._TEMP_COMP_SEGMENT_COUNT) + 1):
            segment_zero_based = int(segment_index - 1)
            recommendations[f"k1_cool_seg{segment_index}_x100"] = cls._saturate_int16(
                int(round(float(cooling_slopes[segment_zero_based]) * 100.0))
            )
            recommendations[f"k1_heat_seg{segment_index}_x100"] = cls._saturate_int16(
                int(round(float(heating_slopes[segment_zero_based]) * 100.0))
            )

        temperature_span_x10 = int(max(temperatures_x10) - min(temperatures_x10))
        mode_candidate_values: dict[int, dict[str, int]] = {}
        mode_candidate_scores: dict[int, float] = {}

        mode0_values = {str(key): int(value) for key, value in recommendations.items()}
        mode0_values["mode"] = int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)
        for segment_index in range(1, int(cls._TEMP_COMP_SEGMENT_COUNT) + 1):
            mode0_values[f"k1_cool_seg{segment_index}_x100"] = int(safe_linear_k1_x100)
            mode0_values[f"k1_heat_seg{segment_index}_x100"] = int(safe_linear_k1_x100)
        mode_candidate_values[int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)] = mode0_values
        mode_candidate_scores[int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)] = cls._temp_comp_candidate_mode_score(
            ordered_by_time,
            linear_k1_x100=int(safe_linear_k1_x100),
            advanced_values=mode0_values,
        )

        enable_mode1 = (
            temperature_span_x10 >= int(cls._TEMP_COMP_REC_MIN_SPAN_MODE1_X10)
            or forced_mode_value == int(cls._TEMP_COMP_MODE_SEGMENTED)
        )
        if enable_mode1:
            mode1_values = {str(key): int(value) for key, value in recommendations.items()}
            mode1_values["mode"] = int(cls._TEMP_COMP_MODE_SEGMENTED)
            for segment_index in range(1, int(cls._TEMP_COMP_SEGMENT_COUNT) + 1):
                mode1_values[f"k1_heat_seg{segment_index}_x100"] = int(mode1_values.get(f"k1_cool_seg{segment_index}_x100", int(safe_linear_k1_x100)))
            mode_candidate_values[int(cls._TEMP_COMP_MODE_SEGMENTED)] = mode1_values
            mode_candidate_scores[int(cls._TEMP_COMP_MODE_SEGMENTED)] = cls._temp_comp_candidate_mode_score(
                ordered_by_time,
                linear_k1_x100=int(safe_linear_k1_x100),
                advanced_values=mode1_values,
            )

        enable_mode2 = (
            (
                temperature_span_x10 >= int(cls._TEMP_COMP_REC_MIN_SPAN_MODE2_X10)
                and len(cool_all_indices) >= int(cls._TEMP_COMP_REC_MIN_BRANCH_POINTS)
                and len(heat_all_indices) >= int(cls._TEMP_COMP_REC_MIN_BRANCH_POINTS)
            )
            or forced_mode_value == int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)
        )
        if enable_mode2:
            mode2_values = {str(key): int(value) for key, value in recommendations.items()}
            mode2_values["mode"] = int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)
            mode_candidate_values[int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)] = mode2_values
            mode_candidate_scores[int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)] = cls._temp_comp_candidate_mode_score(
                ordered_by_time,
                linear_k1_x100=int(safe_linear_k1_x100),
                advanced_values=mode2_values,
            )

        selected_mode = int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)
        score_mode0 = float(mode_candidate_scores.get(int(cls._TEMP_COMP_MODE_SINGLE_LINEAR), float("inf")))
        score_mode1 = float(mode_candidate_scores.get(int(cls._TEMP_COMP_MODE_SEGMENTED), float("inf")))
        score_mode2 = float(mode_candidate_scores.get(int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL), float("inf")))

        if forced_mode_value is not None and forced_mode_value in mode_candidate_values:
            selected_mode = int(forced_mode_value)
        else:
            if (
                int(cls._TEMP_COMP_MODE_SEGMENTED) in mode_candidate_scores
                and score_mode1 < (score_mode0 * (1.0 - float(cls._TEMP_COMP_REC_MODE1_MIN_GAIN)))
            ):
                selected_mode = int(cls._TEMP_COMP_MODE_SEGMENTED)

            baseline_for_mode2 = min(
                score_mode0,
                score_mode1 if int(cls._TEMP_COMP_MODE_SEGMENTED) in mode_candidate_scores else float("inf"),
            )
            if (
                int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL) in mode_candidate_scores
                and baseline_for_mode2 < float("inf")
                and score_mode2 < (baseline_for_mode2 * (1.0 - float(cls._TEMP_COMP_REC_MODE2_MIN_GAIN)))
            ):
                selected_mode = int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)

        selected_values = mode_candidate_values.get(selected_mode, mode_candidate_values[int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)])
        recommendations = {str(key): int(value) for key, value in selected_values.items()}
        recommendations["mode"] = int(selected_mode)

        return recommendations

    @staticmethod
    def _parse_csv_float(value: object) -> float | None:
        """Цель функции в чтении чисел из CSV, затем она преобразует строку с запятой или точкой в float."""
        raw = str(value or "").strip().replace(" ", "")
        if not raw:
            return None
        normalized = raw.replace(",", ".")
        try:
            return float(normalized)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _parse_csv_int(value: object) -> int | None:
        """Цель функции в разборе целого значения периода, затем она обрабатывает как integer, так и float-представление."""
        parsed_float = AppControllerCalibrationMixin._parse_csv_float(value)
        if parsed_float is None:
            return None
        try:
            return int(round(float(parsed_float)))
        except Exception:
            return None

    @staticmethod
    def _extract_node_sa_from_text(text: object) -> int | None:
        """Цель функции в извлечении адреса узла, затем она ищет SA формата 0xNN в произвольной строке."""
        raw = str(text or "")
        match = re.search(r"0x([0-9a-fA-F]{1,2})", raw)
        if match is None:
            return None
        try:
            return int(match.group(1), 16) & 0xFF
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _extract_empty_full_from_text(text: object) -> tuple[int | None, int | None]:
        """Цель функции в разборе калибровки из метаданных CSV, затем она извлекает empty/full через регулярные выражения."""
        raw = str(text or "")
        empty_match = re.search(r"empty\s*=\s*(-?\d+)", raw, flags=re.IGNORECASE)
        full_match = re.search(r"full\s*=\s*(-?\d+)", raw, flags=re.IGNORECASE)
        empty_value = int(empty_match.group(1)) if empty_match is not None else None
        full_value = int(full_match.group(1)) if full_match is not None else None
        return empty_value, full_value

    def _extract_temp_comp_csv_node_candidates(self, rows: list[list[str]]) -> set[int]:
        """Цель функции в получении списка узлов из CSV, затем она ищет SA в структурных и обычных заголовках."""
        node_candidates: set[int] = set()
        if len(rows) == 0:
            return node_candidates

        for row in rows[:3]:
            for cell in row:
                text = str(cell or "")
                lowered = text.casefold()
                if ("узел" not in lowered) and ("node" not in lowered):
                    continue
                parsed_sa = self._extract_node_sa_from_text(text)
                if parsed_sa is None:
                    continue
                node_candidates.add(int(parsed_sa) & 0xFF)

        if len(node_candidates) > 0:
            return node_candidates

        for row in rows[:8]:
            joined = ";".join(str(cell or "").strip() for cell in row).casefold()
            if not (("период" in joined or "period" in joined) and ("температ" in joined or "temp" in joined)):
                continue

            for cell in row:
                text = str(cell or "")
                lowered = text.casefold()
                if not (
                    ("период" in lowered)
                    or ("period" in lowered)
                    or ("температ" in lowered)
                    or ("temp" in lowered)
                ):
                    continue
                parsed_sa = self._extract_node_sa_from_text(text)
                if parsed_sa is None:
                    continue
                node_candidates.add(int(parsed_sa) & 0xFF)
            break

        return node_candidates

    @staticmethod
    def _is_csv_header_like(row: list[str]) -> bool:
        """Цель функции в пропуске служебных строк, затем она определяет заголовки/метаданные по ключевым словам."""
        if len(row) == 0:
            return True
        joined = ";".join(str(cell or "").strip() for cell in row).casefold()
        if not joined:
            return True
        markers = ("время", "узел", "калибровк", "формул", "period", "период", "температ", "temp", "топлив", "fuel")
        return any(marker in joined for marker in markers)

    def _build_temp_comp_samples_from_columns(
        self,
        rows: list[list[str]],
        *,
        data_start: int,
        period_index: int,
        temperature_index: int,
    ) -> list[dict[str, object]]:
        """Цель функции в сборке точек анализа, затем она конвертирует строки CSV в формат внутренних семплов."""
        samples: list[dict[str, object]] = []
        for row in rows[data_start:]:
            if len(row) <= max(period_index, temperature_index):
                continue

            period = self._parse_csv_int(row[period_index])
            temperature_c = self._parse_csv_float(row[temperature_index])
            if period is None or temperature_c is None:
                continue

            temperature_x10 = int(round(float(temperature_c) * 10.0))
            samples.append(
                {
                    "period": int(period),
                    "temperature_x10": int(temperature_x10),
                    "temperature_c": float(temperature_x10) / 10.0,
                    "timestamp": float(len(samples)),
                }
            )

        return samples

    def _parse_temp_comp_structured_csv_all_nodes(
        self,
        rows: list[list[str]],
    ) -> dict[int, dict[str, object]]:
        """Цель функции в разборе all_nodes CSV по всем узлам, затем она возвращает выборки и калибровку для каждого SA."""
        if len(rows) < 3:
            return {}

        node_row = rows[0]
        second_row = rows[1] if len(rows) > 1 else []
        third_row = rows[2] if len(rows) > 2 else []

        second_joined = ";".join(str(cell or "").strip() for cell in second_row).casefold()
        has_calibration_meta = (
            ("empty=" in second_joined)
            or ("full=" in second_joined)
            or ("калибровк" in second_joined)
        )

        calibration_row = second_row if has_calibration_meta else []
        labels_row = third_row if has_calibration_meta else second_row
        data_start = 3 if has_calibration_meta else 2
        node_data_by_sa: dict[int, dict[str, object]] = {}

        for index, cell in enumerate(node_row):
            text = str(cell or "")
            lowered = text.casefold()
            if ("узел" not in lowered) and ("node" not in lowered):
                continue

            node_sa = self._extract_node_sa_from_text(text)
            if node_sa is None:
                continue
            node_sa_key = int(node_sa) & 0xFF

            period_index = None
            temperature_index = None
            search_end = min(len(labels_row), index + 6)
            for column in range(index, search_end):
                label_text = str(labels_row[column] if column < len(labels_row) else "").casefold()
                if period_index is None and (("период" in label_text) or ("period" in label_text)):
                    period_index = column
                if temperature_index is None and (("температ" in label_text) or ("temp" in label_text)):
                    temperature_index = column

            if period_index is None:
                period_index = index
            if temperature_index is None:
                temperature_index = index + 2

            samples = self._build_temp_comp_samples_from_columns(
                rows,
                data_start=data_start,
                period_index=int(period_index),
                temperature_index=int(temperature_index),
            )
            if len(samples) <= 0:
                continue

            calibration_cells: list[str] = []
            if index < len(calibration_row):
                calibration_cells.append(str(calibration_row[index]))
            if (index + 1) < len(calibration_row):
                calibration_cells.append(str(calibration_row[index + 1]))
            empty_period, full_period = self._extract_empty_full_from_text(";".join(calibration_cells))

            existing_entry = node_data_by_sa.get(node_sa_key)
            if existing_entry is None:
                node_data_by_sa[node_sa_key] = {
                    "samples": list(samples),
                    "empty": None if empty_period is None else int(empty_period),
                    "full": None if full_period is None else int(full_period),
                }
                continue

            merged_samples = list(existing_entry.get("samples", []))
            merged_samples.extend(samples)
            existing_entry["samples"] = merged_samples
            if empty_period is not None:
                existing_entry["empty"] = int(empty_period)
            if full_period is not None:
                existing_entry["full"] = int(full_period)

        return node_data_by_sa

    def _parse_temp_comp_structured_csv(
        self,
        rows: list[list[str]],
        *,
        selected_sa: int | None,
    ) -> tuple[list[dict[str, object]], int | None, int | None, int | None]:
        """Цель функции в выборе одного узла из all_nodes CSV, затем она возвращает выборку для текущего SA."""
        node_data_by_sa = self._parse_temp_comp_structured_csv_all_nodes(rows)
        if len(node_data_by_sa) <= 0:
            return [], None, None, None

        chosen_sa: int | None = None
        if selected_sa is not None:
            wanted_sa = int(selected_sa) & 0xFF
            if wanted_sa not in node_data_by_sa:
                return [], None, None, None
            chosen_sa = wanted_sa

        if chosen_sa is None:
            chosen_sa = next(iter(node_data_by_sa.keys()))

        chosen_payload = node_data_by_sa.get(int(chosen_sa) & 0xFF, {})
        samples = list(chosen_payload.get("samples", []))
        empty_value = chosen_payload.get("empty")
        full_value = chosen_payload.get("full")
        return (
            samples,
            int(chosen_sa) & 0xFF,
            None if empty_value is None else int(empty_value),
            None if full_value is None else int(full_value),
        )

    def _parse_temp_comp_generic_csv(
        self,
        rows: list[list[str]],
        *,
        selected_sa: int | None,
    ) -> tuple[list[dict[str, object]], int | None, int | None, int | None]:
        """Цель функции в поддержке старых CSV форматов, затем она ищет колонки периода/температуры и строит выборку."""
        header_index = -1
        header_row: list[str] = []
        for index, row in enumerate(rows):
            joined = ";".join(str(cell or "").strip() for cell in row).casefold()
            if ("период" in joined or "period" in joined) and ("температ" in joined or "temp" in joined):
                header_index = index
                header_row = row
                break

        if header_index < 0:
            return [], None, None, None

        period_by_sa: dict[int, int] = {}
        temp_by_sa: dict[int, int] = {}
        period_no_sa: list[int] = []
        temp_no_sa: list[int] = []

        for index, cell in enumerate(header_row):
            text = str(cell or "")
            lowered = text.casefold()
            node_sa = self._extract_node_sa_from_text(text)
            if "период" in lowered or "period" in lowered:
                if node_sa is None:
                    period_no_sa.append(index)
                else:
                    period_by_sa[int(node_sa)] = index
            if "температ" in lowered or "temp" in lowered:
                if node_sa is None:
                    temp_no_sa.append(index)
                else:
                    temp_by_sa[int(node_sa)] = index

        node_sa_used = None
        period_index = None
        temperature_index = None

        if selected_sa is not None:
            selected_key = int(selected_sa) & 0xFF
            if selected_key in period_by_sa and selected_key in temp_by_sa:
                node_sa_used = selected_key
                period_index = int(period_by_sa[selected_key])
                temperature_index = int(temp_by_sa[selected_key])
            elif len(period_no_sa) > 0 and len(temp_no_sa) > 0:
                # CSV без SA-меток в заголовке: используем первую пару колонок.
                node_sa_used = selected_key
                period_index = int(period_no_sa[0])
                temperature_index = int(temp_no_sa[0])
            else:
                return [], None, None, None
        elif len(period_no_sa) > 0 and len(temp_no_sa) > 0:
            period_index = int(period_no_sa[0])
            temperature_index = int(temp_no_sa[0])
        else:
            common_nodes = sorted(set(period_by_sa.keys()) & set(temp_by_sa.keys()))
            if len(common_nodes) == 0:
                return [], None, None, None
            node_sa_used = int(common_nodes[0]) & 0xFF
            period_index = int(period_by_sa[node_sa_used])
            temperature_index = int(temp_by_sa[node_sa_used])

        empty_period = None
        full_period = None
        for row in rows[:header_index + 1]:
            empty_candidate, full_candidate = self._extract_empty_full_from_text(";".join(str(cell) for cell in row))
            if empty_candidate is not None:
                empty_period = int(empty_candidate)
            if full_candidate is not None:
                full_period = int(full_candidate)

        samples = self._build_temp_comp_samples_from_columns(
            rows,
            data_start=header_index + 1,
            period_index=int(period_index),
            temperature_index=int(temperature_index),
        )
        return samples, node_sa_used, empty_period, full_period

    def _parse_temp_comp_generic_csv_all_nodes(
        self,
        rows: list[list[str]],
    ) -> dict[int, dict[str, object]]:
        """Цель функции в разборе SA-колонок generic CSV, затем она формирует выборки по каждому найденному узлу."""
        header_index = -1
        header_row: list[str] = []
        for index, row in enumerate(rows):
            joined = ";".join(str(cell or "").strip() for cell in row).casefold()
            if ("период" in joined or "period" in joined) and ("температ" in joined or "temp" in joined):
                header_index = index
                header_row = row
                break

        if header_index < 0:
            return {}

        period_by_sa: dict[int, int] = {}
        temp_by_sa: dict[int, int] = {}
        for index, cell in enumerate(header_row):
            text = str(cell or "")
            lowered = text.casefold()
            node_sa = self._extract_node_sa_from_text(text)
            if node_sa is None:
                continue
            node_key = int(node_sa) & 0xFF
            if "период" in lowered or "period" in lowered:
                period_by_sa[node_key] = int(index)
            if "температ" in lowered or "temp" in lowered:
                temp_by_sa[node_key] = int(index)

        common_nodes = sorted(set(period_by_sa.keys()) & set(temp_by_sa.keys()))
        if len(common_nodes) <= 0:
            return {}

        empty_period = None
        full_period = None
        for row in rows[:header_index + 1]:
            empty_candidate, full_candidate = self._extract_empty_full_from_text(";".join(str(cell) for cell in row))
            if empty_candidate is not None:
                empty_period = int(empty_candidate)
            if full_candidate is not None:
                full_period = int(full_candidate)

        node_data_by_sa: dict[int, dict[str, object]] = {}
        for node_sa in common_nodes:
            samples = self._build_temp_comp_samples_from_columns(
                rows,
                data_start=header_index + 1,
                period_index=int(period_by_sa[node_sa]),
                temperature_index=int(temp_by_sa[node_sa]),
            )
            if len(samples) <= 0:
                continue
            node_data_by_sa[int(node_sa) & 0xFF] = {
                "samples": list(samples),
                "empty": None if empty_period is None else int(empty_period),
                "full": None if full_period is None else int(full_period),
            }

        return node_data_by_sa

    @staticmethod
    def _normalize_temp_comp_table_cell(cell_value: object) -> str:
        """Цель функции в приведении ячейки к строке, затем она возвращает безопасное текстовое значение для общего парсера."""
        if cell_value is None:
            return ""
        return str(cell_value).strip()

    def _read_calibration_temp_comp_table_rows(
        self,
        table_path: Path,
    ) -> tuple[list[list[str]], str | None]:
        """Цель функции в чтении CSV/XLSX, затем она возвращает строки таблицы в едином формате или текст ошибки."""
        try:
            resolved_path = Path(table_path).expanduser().resolve()
        except Exception:
            resolved_path = Path(table_path)

        if not resolved_path.exists() or not resolved_path.is_file():
            return [], "Файл не найден."

        suffix = str(resolved_path.suffix or "").casefold()
        if suffix == ".csv":
            try:
                with resolved_path.open("r", encoding="utf-8-sig", newline="") as file:
                    rows = list(csv.reader(file, delimiter=";"))
            except Exception as exc:
                return [], f"Не удалось прочитать CSV: {str(exc)}"
            return rows, None

        if suffix == ".xlsx":
            if load_workbook is None:
                return [], "Для загрузки XLSX установите пакет openpyxl."
            try:
                workbook = load_workbook(filename=str(resolved_path), read_only=True, data_only=True)
                worksheet = workbook.active
                rows: list[list[str]] = []
                for row in worksheet.iter_rows(values_only=True):
                    normalized_row = [self._normalize_temp_comp_table_cell(cell) for cell in row]
                    while len(normalized_row) > 0 and normalized_row[-1] == "":
                        normalized_row.pop()
                    rows.append(normalized_row)
                workbook.close()
            except Exception as exc:
                return [], f"Не удалось прочитать XLSX: {str(exc)}"
            return rows, None

        return [], "Поддерживаются только файлы CSV и XLSX."

    def _parse_calibration_temp_comp_csv_file(
        self,
        csv_path: Path,
        *,
        selected_sa: int | None,
    ) -> tuple[list[dict[str, object]], int | None, int | None, int | None, set[int]]:
        """Цель функции в разборе одного CSV/XLSX лога, затем она извлекает точки period+temperature для анализа K1."""
        try:
            resolved_path = Path(csv_path).expanduser().resolve()
        except Exception:
            resolved_path = Path(csv_path)

        if not resolved_path.exists() or not resolved_path.is_file():
            return [], None, None, None, set()

        rows, _ = self._read_calibration_temp_comp_table_rows(resolved_path)
        if len(rows) == 0:
            return [], None, None, None, set()

        node_candidates = self._extract_temp_comp_csv_node_candidates(rows)

        structured_samples, structured_sa, structured_empty, structured_full = self._parse_temp_comp_structured_csv(
            rows,
            selected_sa=selected_sa,
        )
        if len(structured_samples) > 0:
            return structured_samples, structured_sa, structured_empty, structured_full, node_candidates

        # Если это структурированный all_nodes (есть групповые колонки "узел"/"node"),
        # но выбранный SA в файле отсутствует, не подменяем выбор generic-парсером.
        if selected_sa is not None and len(rows) > 0:
            first_row_joined = ";".join(str(cell or "").strip() for cell in rows[0]).casefold()
            if ("узел" in first_row_joined) or ("node" in first_row_joined):
                return [], None, None, None, node_candidates

        parsed_samples, parsed_sa, parsed_empty, parsed_full = self._parse_temp_comp_generic_csv(
            rows,
            selected_sa=selected_sa,
        )
        return parsed_samples, parsed_sa, parsed_empty, parsed_full, node_candidates

    def _parse_calibration_temp_comp_csv_file_all_nodes(
        self,
        csv_path: Path,
    ) -> tuple[dict[int, dict[str, object]], set[int]]:
        """Цель функции в извлечении выборок всех узлов из одного CSV/XLSX, затем она подготавливает карту SA->данные."""
        try:
            resolved_path = Path(csv_path).expanduser().resolve()
        except Exception:
            resolved_path = Path(csv_path)

        if not resolved_path.exists() or not resolved_path.is_file():
            return {}, set()

        rows, _ = self._read_calibration_temp_comp_table_rows(resolved_path)
        if len(rows) == 0:
            return {}, set()

        node_candidates = self._extract_temp_comp_csv_node_candidates(rows)
        structured_data = self._parse_temp_comp_structured_csv_all_nodes(rows)
        if len(structured_data) > 0:
            node_candidates.update(int(value) & 0xFF for value in structured_data.keys())
            return structured_data, node_candidates

        generic_data = self._parse_temp_comp_generic_csv_all_nodes(rows)
        if len(generic_data) > 0:
            node_candidates.update(int(value) & 0xFF for value in generic_data.keys())
            return generic_data, node_candidates

        return {}, node_candidates

    def _load_calibration_temp_comp_csv_files(self, paths: list[Path]) -> tuple[int, int]:
        """Цель функции в пакетной загрузке CSV/XLSX, затем она агрегирует выборки по узлам и обновляет анализ K1."""
        selected_sa = self._calibration_target_node_sa
        selected_key = None if selected_sa is None else (int(selected_sa) & 0xFF)
        preferred_dataset_sa = self._selected_calibration_temp_comp_dataset_sa()
        if preferred_dataset_sa is None:
            if selected_key is not None:
                preferred_dataset_sa = int(selected_key) & 0xFF
            else:
                preferred_dataset_sa = int(self._resolve_calibration_target_sa()) & 0xFF
        requested_files = len(paths)
        loaded_files = 0
        loaded_points = 0
        skipped_files = 0
        csv_node_candidates: set[int] = set()
        aggregated_by_sa: dict[int, dict[str, object]] = {}

        # Перед новой загрузкой очищаем прошлый набор офлайн-выборок по узлам.
        self._calibration_temp_comp_samples_by_node = {}

        for file_index, source_path in enumerate(paths, start=1):
            file_name = ""
            try:
                file_name = Path(source_path).name
            except Exception:
                file_name = str(source_path)
            if not file_name:
                file_name = str(source_path)
            progress_percent = 0
            if int(requested_files) > 0:
                progress_percent = int(
                    round((float(int(file_index) - 1.0) / float(int(requested_files))) * 100.0)
                )
            self._set_calibration_temp_comp_operation_status(
                f"Загрузка CSV/XLSX и пересчет графиков: файл {int(file_index)}/{int(requested_files)} ({file_name}).",
                busy=True,
                progress_percent=progress_percent,
                determinate=True,
            )
            file_suffix = str(Path(source_path).suffix or "").casefold()
            if file_suffix == ".xlsx" and load_workbook is None:
                skipped_files += 1
                self._append_log(
                    "Калибровка: для загрузки XLSX требуется пакет openpyxl. Установите зависимость и повторите попытку.",
                    RowColor.red,
                )
                self._set_calibration_temp_comp_operation_status(
                    "Загрузка XLSX невозможна: не установлен пакет openpyxl.",
                    busy=False,
                    progress_percent=100,
                    determinate=True,
                )
                continue

            file_node_data, file_node_candidates = self._parse_calibration_temp_comp_csv_file_all_nodes(source_path)
            csv_node_candidates.update(int(value) & 0xFF for value in file_node_candidates)

            if len(file_node_data) <= 0:
                samples, node_sa_used, empty_period, full_period, fallback_candidates = self._parse_calibration_temp_comp_csv_file(
                    source_path,
                    selected_sa=selected_key if selected_key is not None else preferred_dataset_sa,
                )
                csv_node_candidates.update(int(value) & 0xFF for value in fallback_candidates)
                if len(samples) <= 0:
                    skipped_files += 1
                    continue

                fallback_sa = node_sa_used
                if fallback_sa is None:
                    if selected_key is not None:
                        fallback_sa = selected_key
                    else:
                        fallback_sa = int(self._resolve_calibration_target_sa()) & 0xFF

                file_node_data = {
                    int(fallback_sa) & 0xFF: {
                        "samples": list(samples),
                        "empty": None if empty_period is None else int(empty_period),
                        "full": None if full_period is None else int(full_period),
                    }
                }

            loaded_files += 1
            for node_sa, payload in file_node_data.items():
                node_key = int(node_sa) & 0xFF
                node_samples = list(payload.get("samples", []))
                if len(node_samples) <= 0:
                    continue

                loaded_points += len(node_samples)
                csv_node_candidates.add(node_key)
                existing = aggregated_by_sa.get(node_key)
                if existing is None:
                    aggregated_by_sa[node_key] = {
                        "samples": list(node_samples),
                        "empty": payload.get("empty"),
                        "full": payload.get("full"),
                    }
                    continue

                merged_samples = list(existing.get("samples", []))
                merged_samples.extend(node_samples)
                existing["samples"] = merged_samples
                if payload.get("empty") is not None:
                    existing["empty"] = payload.get("empty")
                if payload.get("full") is not None:
                    existing["full"] = payload.get("full")

        prepared_by_sa: dict[int, dict[str, object]] = {}
        for node_sa, payload in aggregated_by_sa.items():
            node_key = int(node_sa) & 0xFF
            node_samples = list(payload.get("samples", []))
            if len(node_samples) <= 0:
                continue

            prepared_by_sa[node_key] = {
                "samples": list(node_samples),
                "empty": payload.get("empty"),
                "full": payload.get("full"),
            }

        if len(prepared_by_sa) <= 0:
            self._calibration_temp_comp_samples_by_node = {}
            self._refresh_calibration_temp_comp_dataset_options(preferred_sa=None)
            self._calibration_temp_comp_status = (
                f"CSV/XLSX файлы обработаны ({int(requested_files)}), но данные period/temperature не распознаны."
            )
            if len(csv_node_candidates) > 0:
                available_nodes = ", ".join(f"0x{value:02X}" for value in sorted(csv_node_candidates))
                self._calibration_temp_comp_status = (
                    f"CSV/XLSX файлы обработаны ({int(requested_files)}), но данные period/temperature не распознаны. "
                    f"Доступные узлы в файлах: {available_nodes}."
                )
            self.calibrationTempCompChanged.emit()
            return 0, 0

        self._calibration_temp_comp_samples_by_node = prepared_by_sa
        applied_sa = self._refresh_calibration_temp_comp_dataset_options(preferred_sa=preferred_dataset_sa)
        if applied_sa is None:
            self._reset_calibration_temp_comp_state(
                clear_samples=True,
                clear_coefficients=False,
                clear_cached_nodes=False,
            )
            self._calibration_temp_comp_status = (
                "Не удалось выбрать набор CSV/XLSX для офлайн-анализа. "
                "Проверьте содержимое файла и повторите загрузку."
            )
            self.calibrationTempCompChanged.emit()
            return loaded_files, loaded_points

        if not self._apply_calibration_temp_comp_node_samples(
            applied_sa,
            clear_coefficients=False,
        ):
            self._reset_calibration_temp_comp_state(
                clear_samples=True,
                clear_coefficients=False,
                clear_cached_nodes=False,
            )
            self._calibration_temp_comp_status = (
                f"Не удалось применить набор CSV/XLSX для узла 0x{int(applied_sa) & 0xFF:02X}."
            )
            self.calibrationTempCompChanged.emit()
            return loaded_files, loaded_points

        status_hints: list[str] = []
        if applied_sa is not None:
            status_hints.append(f"Активный узел анализа: 0x{int(applied_sa) & 0xFF:02X}.")
            if len(prepared_by_sa) > 1:
                status_hints.append("Для просмотра другого узла переключите селектор набора CSV/XLSX в этом блоке.")
        if skipped_files > 0:
            status_hints.append(f"Пропущено файлов без подходящих данных: {int(skipped_files)}.")
        if len(prepared_by_sa) > 1:
            sorted_sas = ", ".join(f"0x{value:02X}" for value in sorted(prepared_by_sa.keys()))
            status_hints.append(f"В CSV/XLSX обнаружены узлы: {sorted_sas}.")

        if len(status_hints) > 0:
            self._calibration_temp_comp_status = f"{self._calibration_temp_comp_status} {' '.join(status_hints)}"
            self.calibrationTempCompChanged.emit()

        return loaded_files, loaded_points

    @staticmethod
    def _build_temp_comp_period_accessor(
        samples: list[dict[str, object]],
        compensated_periods: list[float] | None,
    ):
        """Цель функции в выборе источника period для графика, затем она возвращает быстрый accessor по индексу."""
        if compensated_periods is None:
            return lambda idx: float(samples[idx].get("period", 0.0))
        return lambda idx: float(compensated_periods[idx]) if idx < len(compensated_periods) else float(samples[idx].get("period", 0.0))

    @classmethod
    def _select_temp_comp_chart_indices(
        cls,
        samples: list[dict[str, object]],
        *,
        compensated_periods: list[float] | None = None,
        include_indices: list[int] | None = None,
    ) -> list[int]:
        """Цель функции в подготовке индексов графика без перегруза UI, затем она сохраняет пики и края каждого температурного сегмента."""
        total_points = len(samples)
        if total_points <= 0:
            return []

        max_points = max(512, int(cls._TEMP_COMP_CHART_POINT_LIMIT))
        if total_points <= max_points:
            selected_indices = list(range(total_points))
            if include_indices is not None:
                for required_index in include_indices:
                    if not isinstance(required_index, int):
                        continue
                    if required_index < 0 or required_index >= total_points:
                        continue
                    if required_index not in selected_indices:
                        selected_indices.append(required_index)
            selected_indices.sort()
            return selected_indices

        period_at = cls._build_temp_comp_period_accessor(samples, compensated_periods)
        bucket_count = max(1, max_points // 4)
        bucket_size = max(1, int((total_points + bucket_count - 1) // bucket_count))

        selected_indices: list[int] = []
        used_indices: set[int] = set()

        def add_index(index: int):
            if index < 0 or index >= total_points:
                return
            if index in used_indices:
                return
            used_indices.add(index)
            selected_indices.append(index)

        for start in range(0, total_points, bucket_size):
            end = min(total_points, start + bucket_size)
            if end <= start:
                continue

            first_index = start
            last_index = end - 1
            min_index = first_index
            max_index = first_index
            min_period = period_at(first_index)
            max_period = min_period

            for index in range(start + 1, end):
                period_value = period_at(index)
                if period_value < min_period:
                    min_period = period_value
                    min_index = index
                if period_value > max_period:
                    max_period = period_value
                    max_index = index

            for index in sorted({first_index, min_index, max_index, last_index}):
                add_index(index)

        add_index(total_points - 1)
        if include_indices is not None:
            for required_index in include_indices:
                if not isinstance(required_index, int):
                    continue
                add_index(int(required_index))
        selected_indices.sort()
        return selected_indices

    @classmethod
    def _build_calibration_temp_comp_chart_points(
        cls,
        samples: list[dict[str, object]],
        *,
        compensated_periods: list[float] | None = None,
        include_indices: list[int] | None = None,
        highlight_labels_by_index: dict[int, str] | None = None,
    ) -> list[dict[str, object]]:
        """Цель функции в подготовке данных для TrendCanvas, затем она ограничивает payload без потери ключевой формы сигнала."""
        selected_indices = cls._select_temp_comp_chart_indices(
            samples,
            compensated_periods=compensated_periods,
            include_indices=include_indices,
        )
        period_at = cls._build_temp_comp_period_accessor(samples, compensated_periods)

        points: list[dict[str, object]] = []
        for index in selected_indices:
            sample = samples[index]
            point: dict[str, object] = {
                "fuel": float(period_at(index)),
                "temperature": float(sample.get("temperature_c", 0.0)),
                "time": str(int(index) + 1),
            }
            if highlight_labels_by_index is not None and index in highlight_labels_by_index:
                point["isHighlight"] = True
                point["highlightLabel"] = str(highlight_labels_by_index[index])
            points.append(point)
        return points

    def _build_temp_comp_chart_series_entry(
        self,
        *,
        name: str,
        color: str,
        points: list[dict[str, object]],
        max_abs_level: float | None = None,
    ) -> dict[str, object]:
        """Цель функции в единообразном описании серии графика, затем она добавляет метрику max|ошибка| для легенды."""
        series_entry: dict[str, object] = {
            "node": str(name),
            "color": str(color),
            "points": list(points),
        }
        if max_abs_level is not None and math.isfinite(float(max_abs_level)):
            series_entry["maxAbsLevel"] = float(max_abs_level)
            series_entry["maxAbsLevelText"] = f"{float(max_abs_level):.3f} %"
        return series_entry

    def _recompute_calibration_temp_comp_metrics(self):
        """Цель функции в пересчете рекомендаций по K1/K0, затем она обновляет метрики и серии графика для UI."""
        samples = list(self._calibration_temp_comp_samples)
        sample_count = len(samples)
        current_k1 = self._calibration_temp_comp_k1_x100_current
        current_k0 = self._calibration_temp_comp_k0_count_current
        linear_preview_enabled = bool(self._calibration_temp_comp_linear_preview_enabled)
        linear_preview_k1 = self._calibration_temp_comp_linear_preview_k1_x100
        linear_preview_k0 = self._calibration_temp_comp_linear_preview_k0_count
        base_k1: int | None = None
        base_k0: int | None = None
        current_series_k1: int = 0
        current_series_k0: int = 0

        ordered_samples_by_time = sorted(
            samples,
            key=lambda item: float(item.get("timestamp", 0.0)),
        )
        ordered_samples = sorted(
            ordered_samples_by_time,
            key=lambda item: (
                float(item.get("temperature_c", 0.0)),
                float(item.get("timestamp", 0.0)),
            ),
        )
        temperatures = [float(item.get("temperature_c", 0.0)) for item in ordered_samples]
        raw_periods = [float(sample.get("period", 0.0)) for sample in ordered_samples]

        recommended_k1: int | None = None
        delta_k1: int | None = None
        next_k1: int | None = None
        recommended_k0: int | None = None
        delta_k0: int | None = None
        next_k0: int | None = None

        period_slope_before: float | None = None
        period_slope_after: float | None = None
        level_slope_before: float | None = None
        level_slope_after: float | None = None
        period_reduction_percent: float | None = None
        level_reduction_percent: float | None = None

        level_error_range_before: tuple[float, float] | None = None
        level_error_range_after: tuple[float, float] | None = None
        level_error_max_before: float | None = None
        level_error_max_after: float | None = None
        level_error_p95_before: float | None = None
        level_error_p95_after: float | None = None
        raw_level_error_max: float | None = None
        current_level_error_max: float | None = None
        recommended_level_error_max: float | None = None
        raw_level_error_index: int | None = None
        current_level_error_index: int | None = None
        recommended_level_error_index: int | None = None
        raw_level_error_signed: float | None = None
        current_level_error_signed: float | None = None
        recommended_level_error_signed: float | None = None
        recommendation_guard_triggered = False

        current_comp_periods: list[float] = []
        recommended_comp_periods: list[float] = []
        advanced_recommended_values: dict[str, int] = {}
        current_advanced_values: dict[str, int | None] = {}
        current_mode_for_label: int | None = None
        recommended_mode_for_label: int | None = None
        applied_k0_for_recommended: int | None = None
        has_current_comp_context = False
        recommendation_base_k1 = 0
        recommendation_base_k0 = 0

        level_calibration_ready = (
            bool(self._calibration_level_0_known)
            and bool(self._calibration_level_100_known)
            and (int(self._calibration_level_100) > int(self._calibration_level_0))
        )

        if sample_count > 0:
            # Цель в чистом оффлайн-режиме рекомендаций: база K1/K0 всегда 0/0 и не зависит от текущих DID МК.
            base_k1 = int(recommendation_base_k1)
            base_k0 = int(recommendation_base_k0)
            current_series_k1 = int(current_k1) if current_k1 is not None else 0
            current_series_k0 = int(current_k0) if current_k0 is not None else 0
            if linear_preview_enabled:
                if linear_preview_k1 is not None:
                    current_series_k1 = int(linear_preview_k1)
                if linear_preview_k0 is not None:
                    current_series_k0 = int(linear_preview_k0)
            current_advanced_values = {
                str(key): (None if value is None else int(value))
                for key, value in self._calibration_temp_comp_advanced_values.items()
            }
            if linear_preview_enabled:
                current_advanced_values["mode"] = int(self._TEMP_COMP_MODE_SINGLE_LINEAR)
            initial_recommended_values = self._build_temp_comp_advanced_recommendations(
                samples,
                fallback_linear_k1_x100=int(current_series_k1),
            )
            effective_current_values = self._build_effective_temp_comp_values(
                current_advanced_values,
                initial_recommended_values,
            )
            current_mode_for_label = self._temp_comp_get_mode_from_values(effective_current_values)
            current_comp_period_by_sample_id = self._apply_temp_comp_mode_sequence(
                ordered_samples_by_time,
                linear_k1_x100=int(current_series_k1),
                k0_count=int(current_series_k0),
                advanced_values=effective_current_values,
            )
            current_comp_periods = [
                float(current_comp_period_by_sample_id.get(id(sample), float(sample.get("period", 0.0))))
                for sample in ordered_samples
            ]

            advanced_recommended_values = dict(initial_recommended_values)
            has_current_comp_context = (
                (current_k1 is not None)
                or (current_k0 is not None)
                or any(value is not None for value in current_advanced_values.values())
                or bool(linear_preview_enabled)
            )

        if sample_count >= 2 and base_k1 is not None and base_k0 is not None:
            periods_before = list(raw_periods)
            period_slope_before = self._linear_regression_slope(temperatures, periods_before)
            if period_slope_before is None:
                period_slope_before = self._linear_regression_slope(temperatures, raw_periods)

            if period_slope_before is not None:
                recommended_k1 = self._saturate_int16(int(round(float(period_slope_before) * 100.0)))
                delta_k1 = int(recommended_k1) - int(base_k1)
                next_k1 = int(recommended_k1)
                candidate_values = self._build_temp_comp_advanced_recommendations(
                    samples,
                    fallback_linear_k1_x100=int(recommended_k1),
                )
                preferred_mode = int(candidate_values.get("mode", int(self._TEMP_COMP_MODE_SINGLE_LINEAR)))
                if preferred_mode < int(self._TEMP_COMP_MODE_SINGLE_LINEAR) or preferred_mode > int(self._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
                    preferred_mode = int(self._TEMP_COMP_MODE_SINGLE_LINEAR)
                candidate_values["mode"] = int(preferred_mode)
                recommended_mode_for_label = int(preferred_mode)

                candidate_periods_without_k0_by_id = self._apply_temp_comp_mode_sequence(
                    ordered_samples_by_time,
                    linear_k1_x100=int(recommended_k1),
                    k0_count=0,
                    advanced_values=candidate_values,
                )
                candidate_periods_without_k0 = [
                    float(candidate_periods_without_k0_by_id.get(id(sample), float(sample.get("period", 0.0))))
                    for sample in ordered_samples
                ]

                candidate_k0 = int(recommendation_base_k0)
                if level_calibration_ready:
                    candidate_levels_without_k0 = [self._period_to_level_percent(value) for value in candidate_periods_without_k0]
                    if all(level is not None for level in candidate_levels_without_k0):
                        prepared_levels_without_k0 = [
                            float(level)
                            for level in candidate_levels_without_k0
                            if level is not None
                        ]
                        if len(prepared_levels_without_k0) == sample_count:
                            min_level = min(prepared_levels_without_k0)
                            max_level = max(prepared_levels_without_k0)
                            desired_offset_pct = -((min_level + max_level) / 2.0)
                            level_span = int(self._calibration_level_100) - int(self._calibration_level_0)
                            candidate_k0 = self._saturate_int16(
                                int(round((desired_offset_pct * float(level_span)) / 100.0))
                            )

                candidate_periods_by_id = self._apply_temp_comp_mode_sequence(
                    ordered_samples_by_time,
                    linear_k1_x100=int(recommended_k1),
                    k0_count=int(candidate_k0),
                    advanced_values=candidate_values,
                )
                candidate_periods = [
                    float(candidate_periods_by_id.get(id(sample), float(sample.get("period", 0.0))))
                    for sample in ordered_samples
                ]

                if len(candidate_periods) == sample_count:
                    advanced_recommended_values = {str(key): int(value) for key, value in candidate_values.items()}
                    applied_k0_for_recommended = int(candidate_k0)
                    recommended_comp_periods = list(candidate_periods)
                    period_slope_after = self._linear_regression_slope(temperatures, recommended_comp_periods)
                    recommended_k1 = self._saturate_int16(int(recommended_k1))
                    delta_k1 = int(recommended_k1) - int(base_k1)
                    next_k1 = int(recommended_k1)
                    if level_calibration_ready:
                        recommended_k0 = self._saturate_int16(int(candidate_k0))
                        delta_k0 = int(recommended_k0) - int(base_k0)
                        next_k0 = int(recommended_k0)
                    else:
                        recommended_k0 = None
                        delta_k0 = None
                        next_k0 = None

            if level_calibration_ready:
                raw_levels: list[float] = []
                raw_levels_valid = True
                for period_value in raw_periods:
                    converted = self._period_to_level_percent(period_value)
                    if converted is None:
                        raw_levels_valid = False
                        break
                    raw_levels.append(float(converted))

                if raw_levels_valid and len(raw_levels) == sample_count:
                    raw_metrics = self._calc_level_error_metrics(raw_levels)
                    if isinstance(raw_metrics, dict):
                        raw_level_error_max = float(raw_metrics.get("max_abs", 0.0))
                    raw_level_error_index, raw_level_error_signed = self._find_max_abs_level_error(raw_levels)

                    before_levels = list(raw_levels)
                    if has_current_comp_context and len(current_comp_periods) == sample_count:
                        current_levels = [self._period_to_level_percent(value) for value in current_comp_periods]
                        if all(level is not None for level in current_levels):
                            prepared_current_levels = [float(level) for level in current_levels if level is not None]
                            current_metrics = self._calc_level_error_metrics(prepared_current_levels)
                            if isinstance(current_metrics, dict):
                                current_level_error_max = float(current_metrics.get("max_abs", 0.0))
                            current_level_error_index, current_level_error_signed = self._find_max_abs_level_error(prepared_current_levels)
                    if current_level_error_max is None and raw_level_error_max is not None:
                        current_level_error_max = float(raw_level_error_max)
                    if current_level_error_index is None and raw_level_error_index is not None:
                        current_level_error_index = int(raw_level_error_index)
                    if current_level_error_signed is None and raw_level_error_signed is not None:
                        current_level_error_signed = float(raw_level_error_signed)

                    level_slope_before = self._linear_regression_slope(temperatures, before_levels)
                    before_metrics = self._calc_level_error_metrics(before_levels)
                    if isinstance(before_metrics, dict):
                        level_error_range_before = before_metrics.get("range")  # type: ignore[assignment]
                        level_error_max_before = float(before_metrics.get("max_abs", 0.0))
                        level_error_p95_before = float(before_metrics.get("p95_abs", 0.0))

                    if len(recommended_comp_periods) == sample_count:
                        recommended_levels = [self._period_to_level_percent(value) for value in recommended_comp_periods]
                        if all(level is not None for level in recommended_levels):
                            prepared_levels = [float(level) for level in recommended_levels if level is not None]
                            level_slope_after = self._linear_regression_slope(temperatures, prepared_levels)
                            after_metrics = self._calc_level_error_metrics(prepared_levels)
                            if isinstance(after_metrics, dict):
                                level_error_range_after = after_metrics.get("range")  # type: ignore[assignment]
                                level_error_max_after = float(after_metrics.get("max_abs", 0.0))
                                level_error_p95_after = float(after_metrics.get("p95_abs", 0.0))
                                recommended_level_error_max = float(level_error_max_after)
                            recommended_level_error_index, recommended_level_error_signed = self._find_max_abs_level_error(prepared_levels)

            can_compare_quality = (
                len(recommended_comp_periods) == sample_count
                and period_slope_before is not None
                and period_slope_after is not None
            )
            if can_compare_quality:
                candidate_worse = False
                if (
                    level_error_max_before is not None
                    and level_error_max_after is not None
                    and float(level_error_max_after) > (float(level_error_max_before) + 0.0001)
                ):
                    candidate_worse = True
                elif (
                    level_error_p95_before is not None
                    and level_error_p95_after is not None
                    and float(level_error_p95_after) > (float(level_error_p95_before) + 0.0001)
                    and (
                        level_error_max_before is None
                        or level_error_max_after is None
                        or float(level_error_max_after) >= (float(level_error_max_before) - 0.0001)
                    )
                ):
                    candidate_worse = True
                elif abs(float(period_slope_after)) > (abs(float(period_slope_before)) + 0.0001):
                    candidate_worse = True

                if candidate_worse:
                    recommendation_guard_triggered = True
                    recommended_comp_periods = list(raw_periods)
                    period_slope_after = self._linear_regression_slope(temperatures, raw_periods)
                    level_slope_after = level_slope_before
                    level_error_range_after = level_error_range_before
                    level_error_max_after = level_error_max_before
                    level_error_p95_after = level_error_p95_before
                    recommended_level_error_max = raw_level_error_max
                    recommended_level_error_index = raw_level_error_index
                    recommended_level_error_signed = raw_level_error_signed
                    recommended_k1 = int(recommendation_base_k1)
                    delta_k1 = int(recommended_k1) - int(base_k1)
                    next_k1 = int(recommended_k1)
                    recommended_mode_for_label = int(self._TEMP_COMP_MODE_SINGLE_LINEAR)
                    advanced_recommended_values = {
                        "mode": int(self._TEMP_COMP_MODE_SINGLE_LINEAR),
                    }
                    if level_calibration_ready:
                        recommended_k0 = int(recommendation_base_k0)
                        delta_k0 = int(recommended_k0) - int(base_k0)
                        next_k0 = int(recommended_k0)
                elif (
                    recommended_k1 is not None
                    and int(recommended_k1) == int(recommendation_base_k1)
                    and recommended_mode_for_label is None
                ):
                    recommended_mode_for_label = int(self._TEMP_COMP_MODE_SINGLE_LINEAR)
                    if len(advanced_recommended_values) <= 0:
                        if level_calibration_ready:
                            recommended_k0 = int(recommendation_base_k0)
                            delta_k0 = int(recommended_k0) - int(base_k0)
                            next_k0 = int(recommended_k0)

            period_reduction_percent = self._calc_reduction_percent(period_slope_before, period_slope_after)
            level_reduction_percent = self._calc_reduction_percent(level_slope_before, level_slope_after)

        chart_series: list[dict[str, object]] = []
        if sample_count > 0:
            raw_highlight_labels: dict[int, str] | None = None
            if raw_level_error_index is not None:
                raw_signed = 0.0 if raw_level_error_signed is None else float(raw_level_error_signed)
                raw_label = f"max|ошибка|={abs(raw_signed):.3f}% ({raw_signed:+.3f}%)"
                raw_highlight_labels = {int(raw_level_error_index): raw_label}

            raw_points = self._build_calibration_temp_comp_chart_points(
                ordered_samples,
                include_indices=[int(raw_level_error_index)] if raw_level_error_index is not None else None,
                highlight_labels_by_index=raw_highlight_labels,
            )
            chart_series.append(
                self._build_temp_comp_chart_series_entry(
                    name="Сырой период",
                    color="#dc2626",
                    points=raw_points,
                    max_abs_level=raw_level_error_max,
                )
            )

            if (
                len(current_comp_periods) == sample_count
                and has_current_comp_context
            ):
                current_highlight_labels: dict[int, str] | None = None
                if current_level_error_index is not None:
                    current_signed = 0.0 if current_level_error_signed is None else float(current_level_error_signed)
                    current_label = f"max|ошибка|={abs(current_signed):.3f}% ({current_signed:+.3f}%)"
                    current_highlight_labels = {int(current_level_error_index): current_label}
                current_points = self._build_calibration_temp_comp_chart_points(
                    ordered_samples,
                    compensated_periods=current_comp_periods,
                    include_indices=[int(current_level_error_index)] if current_level_error_index is not None else None,
                    highlight_labels_by_index=current_highlight_labels,
                )
                chart_series.append(
                    self._build_temp_comp_chart_series_entry(
                        name=(
                            f"После текущих K1/K0 ({int(current_series_k1)}/{int(current_series_k0)}), mode "
                            f"{int(current_mode_for_label) if current_mode_for_label is not None else 0}"
                            + (" [линейное превью]" if linear_preview_enabled else "")
                        ),
                        color="#2563eb",
                        points=current_points,
                        max_abs_level=current_level_error_max,
                    )
                )

            if len(recommended_comp_periods) == sample_count and recommended_k1 is not None:
                max_shift = max(
                    abs(float(recommended_comp_periods[index]) - float(raw_periods[index]))
                    for index in range(sample_count)
                ) if sample_count > 0 else 0.0
                shift_hint = "" if max_shift >= 0.05 else " (эффект очень мал)"
                recommended_highlight_labels: dict[int, str] | None = None
                if recommended_level_error_index is not None:
                    recommended_signed = 0.0 if recommended_level_error_signed is None else float(recommended_level_error_signed)
                    recommended_label = f"max|ошибка|={abs(recommended_signed):.3f}% ({recommended_signed:+.3f}%)"
                    recommended_highlight_labels = {int(recommended_level_error_index): recommended_label}
                recommended_points = self._build_calibration_temp_comp_chart_points(
                    ordered_samples,
                    compensated_periods=recommended_comp_periods,
                    include_indices=[int(recommended_level_error_index)] if recommended_level_error_index is not None else None,
                    highlight_labels_by_index=recommended_highlight_labels,
                )
                chart_series.append(
                    self._build_temp_comp_chart_series_entry(
                        name=(
                            f"После рекоменд. K1/K0 ({int(recommended_k1)}/{int(next_k0 if next_k0 is not None else base_k0 or 0)}), "
                            f"mode {int(recommended_mode_for_label) if recommended_mode_for_label is not None else int(self._TEMP_COMP_MODE_SINGLE_LINEAR)}"
                            f"{shift_hint}"
                        ),
                        color="#16a34a",
                        points=recommended_points,
                        max_abs_level=recommended_level_error_max,
                    )
                )

        if sample_count <= 0:
            detail_text = "Загрузите CSV/XLSX из коллектора (all_nodes.csv, 0xNN.csv или XLSX), чтобы рассчитать K1/K0."
        elif sample_count < 2:
            detail_text = f"Собрано точек: {sample_count}. Для регрессии нужно минимум 2."
        else:
            temp_min = min(temperatures)
            temp_max = max(temperatures)
            period_min = min(raw_periods)
            period_max = max(raw_periods)
            detail_text = (
                f"Точек: {sample_count}, температура: {temp_min:.1f}..{temp_max:.1f} °C, "
                f"период: {period_min:.1f}..{period_max:.1f} count."
            )
            if base_k1 is not None:
                detail_text += f" Базовый K1={int(base_k1)}."
            if base_k0 is not None:
                detail_text += f" Базовый K0={int(base_k0)}."
            if recommended_k1 is not None:
                detail_text += f" Рекоменд. K1={int(recommended_k1)}."
            if recommended_k0 is not None:
                detail_text += f" Рекоменд. K0={int(recommended_k0)}."
            if delta_k1 is not None:
                detail_text += f" dK1={int(delta_k1):+d}."
            if delta_k0 is not None:
                detail_text += f" dK0={int(delta_k0):+d}."
            detail_text += " Этап «текущее» рассчитан по активному режиму компенсации (mode/segment/heat-cool) и текущим DID узла."
            detail_text += " Рекомендации K1/K0 считаются только по данным выбранного CSV/XLSX набора узла (без влияния текущих коэффициентов МК)."
            detail_text += " Если коэффициенты не считаны, используется офлайн-база K1/K0 = 0/0."
            detail_text += " Рекомендованные DID 0x001D..0x002C отображаются рядом с текущими значениями в расширенной таблице."
            detail_text += " Метрика «текущее» относится к синей серии (если синяя отсутствует, то к красной)."
            detail_text += " Метрика «после рекомендаций» относится к зеленой серии."
            if linear_preview_enabled:
                detail_text += " Для синей серии включено локальное линейное превью K1/K0 (без записи в МК)."
            if recommendation_guard_triggered:
                detail_text += " Защитное правило: новая рекомендация оказалась хуже текущего набора, поэтому предложено сохранить текущие параметры."
            if not level_calibration_ready:
                detail_text += " Для расчета K0 и метрик ошибки нужны калибровки 0% и 100%."

        if sample_count <= 0:
            self._calibration_temp_comp_status = detail_text
        else:
            self._calibration_temp_comp_status = f"Офлайн-анализ CSV/XLSX. {detail_text}"

        self._calibration_temp_comp_k1_x100_base = base_k1
        self._calibration_temp_comp_k1_x100_recommended = recommended_k1
        self._calibration_temp_comp_k1_x100_delta = delta_k1
        self._calibration_temp_comp_k1_x100_next = next_k1
        self._calibration_temp_comp_k0_count_base = base_k0
        self._calibration_temp_comp_k0_count_recommended = recommended_k0
        self._calibration_temp_comp_k0_count_delta = delta_k0
        self._calibration_temp_comp_k0_count_next = next_k0
        self._calibration_temp_comp_period_slope_before = period_slope_before
        self._calibration_temp_comp_period_slope_after = period_slope_after
        self._calibration_temp_comp_level_slope_before = level_slope_before
        self._calibration_temp_comp_level_slope_after = level_slope_after
        self._calibration_temp_comp_period_reduction_percent = period_reduction_percent
        self._calibration_temp_comp_level_reduction_percent = level_reduction_percent
        self._calibration_temp_comp_level_error_range_before = level_error_range_before
        self._calibration_temp_comp_level_error_range_after = level_error_range_after
        self._calibration_temp_comp_level_error_max_before = level_error_max_before
        self._calibration_temp_comp_level_error_max_after = level_error_max_after
        self._calibration_temp_comp_level_error_p95_before = level_error_p95_before
        self._calibration_temp_comp_level_error_p95_after = level_error_p95_after
        self._calibration_temp_comp_chart_series = chart_series
        self._calibration_temp_comp_chart_revision = int(self._calibration_temp_comp_chart_revision) + 1
        self._calibration_temp_comp_advanced_recommended_values = dict(advanced_recommended_values)
        self.calibrationTempCompChanged.emit()

    def _apply_calibration_temp_comp_node_samples(self, node_sa: int, *, clear_coefficients: bool) -> bool:
        """Цель функции в активации выборки конкретного узла, затем она обновляет метрики и связанные поля UI."""
        node_key = int(node_sa) & 0xFF
        payload = self._calibration_temp_comp_samples_by_node.get(node_key)
        if not isinstance(payload, dict):
            return False

        node_samples = list(payload.get("samples", []))
        if len(node_samples) <= 0:
            return False

        self._calibration_temp_comp_samples = list(node_samples)
        last_sample = self._calibration_temp_comp_samples[-1]
        self._calibration_temp_comp_last_period = int(last_sample.get("period", 0))
        self._calibration_temp_comp_last_temperature_x10 = int(last_sample.get("temperature_x10", 0))
        self._calibration_temp_comp_last_temperature_c = float(last_sample.get("temperature_c", 0.0))
        self._calibration_temp_comp_linear_preview_enabled = False
        self._calibration_temp_comp_linear_preview_k1_x100 = None
        self._calibration_temp_comp_linear_preview_k0_count = None
        self._set_calibration_temp_comp_preview_status(
            "Ожидание превью.",
            busy=False,
            progress_percent=0,
            determinate=True,
        )

        if clear_coefficients:
            self._calibration_temp_comp_k1_x100_current = None
            self._calibration_temp_comp_k0_count_current = None
            self._calibration_temp_comp_zero_trim_count_current = None
            self._calibration_temp_comp_zero_trim_count_recommended = None
            self._calibration_temp_comp_zero_trim_count_delta = None
            self._calibration_temp_comp_zero_trim_count_next = None
            self._calibration_temp_comp_zero_trim_residual_x10 = None
            self._calibration_temp_comp_advanced_values = {}
            self._calibration_temp_comp_advanced_recommended_values = {}

        empty_value = payload.get("empty")
        full_value = payload.get("full")
        levels_changed = False
        if empty_value is not None and full_value is not None:
            empty_period = int(empty_value)
            full_period = int(full_value)
            if full_period > empty_period:
                if (
                    int(self._calibration_level_0) != empty_period
                    or int(self._calibration_level_100) != full_period
                    or not bool(self._calibration_level_0_known)
                    or not bool(self._calibration_level_100_known)
                ):
                    self._calibration_level_0 = empty_period
                    self._calibration_level_100 = full_period
                    self._calibration_level_0_known = True
                    self._calibration_level_100_known = True
                    levels_changed = True
        if levels_changed:
            self.calibrationValuesChanged.emit()

        self._recompute_calibration_temp_comp_metrics()
        return True

    def _selected_calibration_temp_comp_dataset_sa(self) -> int | None:
        """Цель функции в чтении текущего выбора набора CSV/XLSX, затем она возвращает SA активного офлайн-узла."""
        selected_index = int(self._selected_calibration_temp_comp_dataset_index)
        if selected_index < 0 or selected_index >= len(self._calibration_temp_comp_dataset_values):
            return None
        return int(self._calibration_temp_comp_dataset_values[selected_index]) & 0xFF

    def _refresh_calibration_temp_comp_dataset_options(self, *, preferred_sa: int | None) -> int | None:
        """Цель функции в обновлении списка наборов CSV/XLSX, затем она сохраняет согласованный индекс выбранного узла анализа."""
        previous_options = list(self._calibration_temp_comp_dataset_options)
        previous_values = list(self._calibration_temp_comp_dataset_values)
        previous_index = int(self._selected_calibration_temp_comp_dataset_index)
        previous_selected_sa = self._selected_calibration_temp_comp_dataset_sa()

        dataset_values = sorted(int(value) & 0xFF for value in self._calibration_temp_comp_samples_by_node.keys())
        dataset_options = [f"Набор CSV/XLSX: 0x{int(value) & 0xFF:02X}" for value in dataset_values]

        preferred_key = None if preferred_sa is None else (int(preferred_sa) & 0xFF)
        selected_sa = None
        if preferred_key is not None and preferred_key in dataset_values:
            selected_sa = preferred_key
        elif previous_selected_sa is not None and previous_selected_sa in dataset_values:
            selected_sa = previous_selected_sa
        elif len(dataset_values) > 0:
            selected_sa = int(dataset_values[0]) & 0xFF

        selected_index = -1
        if selected_sa is not None:
            selected_index = dataset_values.index(int(selected_sa) & 0xFF)

        self._calibration_temp_comp_dataset_values = list(dataset_values)
        self._calibration_temp_comp_dataset_options = list(dataset_options)
        self._selected_calibration_temp_comp_dataset_index = int(selected_index)

        if (
            previous_options != self._calibration_temp_comp_dataset_options
            or previous_values != self._calibration_temp_comp_dataset_values
            or previous_index != self._selected_calibration_temp_comp_dataset_index
        ):
            self.calibrationTempCompChanged.emit()
        return selected_sa

    def _select_calibration_temp_comp_cached_node(
        self,
        *,
        selected_sa: int | None,
        clear_coefficients: bool,
    ) -> int | None:
        """Цель функции в выборе активного узла из загруженного кэша, затем она применяет выборку этого узла."""
        if len(self._calibration_temp_comp_samples_by_node) <= 0:
            return None

        chosen_sa = None
        if selected_sa is not None:
            selected_key = int(selected_sa) & 0xFF
            if selected_key in self._calibration_temp_comp_samples_by_node:
                chosen_sa = selected_key
        else:
            auto_sa = int(self._resolve_calibration_target_sa()) & 0xFF
            if auto_sa in self._calibration_temp_comp_samples_by_node:
                chosen_sa = auto_sa
            else:
                chosen_sa = sorted(self._calibration_temp_comp_samples_by_node.keys())[0]

        if chosen_sa is None:
            return None

        if not self._apply_calibration_temp_comp_node_samples(
            int(chosen_sa) & 0xFF,
            clear_coefficients=clear_coefficients,
        ):
            return None
        return int(chosen_sa) & 0xFF

    def _reset_calibration_temp_comp_state(
        self,
        *,
        clear_samples: bool,
        clear_coefficients: bool,
        clear_cached_nodes: bool = False,
    ):
        """Цель функции в безопасном сбросе сценария компенсации, затем она очищает runtime и при необходимости кэш узлов."""
        self._stop_calibration_temp_comp_advanced_read_sequence()
        self._reset_calibration_temp_comp_recommendation_apply_queue()
        self._calibration_temp_comp_last_period = None
        self._calibration_temp_comp_last_temperature_x10 = None
        self._calibration_temp_comp_last_temperature_c = None

        if clear_samples:
            self._calibration_temp_comp_samples = []
            if clear_cached_nodes:
                self._calibration_temp_comp_samples_by_node = {}
            self._refresh_calibration_temp_comp_dataset_options(preferred_sa=None)

        if clear_coefficients:
            self._calibration_temp_comp_k1_x100_current = None
            self._calibration_temp_comp_k0_count_current = None
            self._calibration_temp_comp_zero_trim_count_current = None
            self._calibration_temp_comp_zero_trim_count_recommended = None
            self._calibration_temp_comp_zero_trim_count_delta = None
            self._calibration_temp_comp_zero_trim_count_next = None
            self._calibration_temp_comp_zero_trim_residual_x10 = None
            self._calibration_temp_comp_advanced_values = {}
            self._calibration_temp_comp_advanced_recommended_values = {}
        self._calibration_temp_comp_linear_preview_enabled = False
        self._calibration_temp_comp_linear_preview_k1_x100 = None
        self._calibration_temp_comp_linear_preview_k0_count = None
        self._set_calibration_temp_comp_preview_status(
            "Ожидание превью.",
            busy=False,
            progress_percent=0,
            determinate=True,
        )

        self._calibration_temp_comp_k1_x100_base = None
        self._calibration_temp_comp_k1_x100_recommended = None
        self._calibration_temp_comp_k1_x100_delta = None
        self._calibration_temp_comp_k1_x100_next = None
        self._reset_calibration_temp_comp_k0_air_zero_adjust_state()
        self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
        self._reset_calibration_temp_comp_zero_trim_verify_state()
        self._calibration_temp_comp_k0_count_base = None
        self._calibration_temp_comp_k0_count_recommended = None
        self._calibration_temp_comp_k0_count_delta = None
        self._calibration_temp_comp_k0_count_next = None
        self._calibration_temp_comp_zero_trim_count_recommended = None
        self._calibration_temp_comp_zero_trim_count_delta = None
        self._calibration_temp_comp_zero_trim_count_next = None
        self._calibration_temp_comp_zero_trim_residual_x10 = None
        self._calibration_temp_comp_period_slope_before = None
        self._calibration_temp_comp_period_slope_after = None
        self._calibration_temp_comp_level_slope_before = None
        self._calibration_temp_comp_level_slope_after = None
        self._calibration_temp_comp_period_reduction_percent = None
        self._calibration_temp_comp_level_reduction_percent = None
        self._calibration_temp_comp_level_error_range_before = None
        self._calibration_temp_comp_level_error_range_after = None
        self._calibration_temp_comp_level_error_max_before = None
        self._calibration_temp_comp_level_error_max_after = None
        self._calibration_temp_comp_level_error_p95_before = None
        self._calibration_temp_comp_level_error_p95_after = None
        self._calibration_temp_comp_chart_series = []

        self._recompute_calibration_temp_comp_metrics()

    def _request_calibration_runtime_snapshot(self):
        """Цель функции в опросе рабочего периода калибровки, затем она читает DID 0x0014 для отображения текущего значения."""
        if not self._can.is_connect or not self._can.is_trace:
            return
        self._configure_calibration_uds_services()
        tx_identifier = self._build_calibration_tx_identifier()
        self._calibration_read_service.read_data_by_identifier(tx_identifier, UdsData.curr_fuel_tank)

    def _reset_calibration_temp_comp_k0_air_zero_adjust_state(self):
        """Цель функции в сбросе автоподстройки K0, затем она очищает временные значения DID 0x0012/0x0013/0x0018/0x001C."""
        if self._calibration_temp_comp_k0_air_zero_adjust_timeout_timer.isActive():
            self._calibration_temp_comp_k0_air_zero_adjust_timeout_timer.stop()
        self._calibration_temp_comp_k0_air_zero_adjust_active = False
        self._calibration_temp_comp_k0_air_zero_adjust_empty_period = None
        self._calibration_temp_comp_k0_air_zero_adjust_full_period = None
        self._calibration_temp_comp_k0_air_zero_adjust_level_x10 = None
        self._calibration_temp_comp_k0_air_zero_adjust_current_k0 = None

    def _reset_calibration_temp_comp_zero_trim_air_zero_adjust_state(self):
        """Цель функции в сбросе автоподстройки zero trim, затем она очищает временные значения DID 0x0012/0x0013/0x0018/0x002D."""
        if self._calibration_temp_comp_zero_trim_air_zero_adjust_timeout_timer.isActive():
            self._calibration_temp_comp_zero_trim_air_zero_adjust_timeout_timer.stop()
        self._calibration_temp_comp_zero_trim_air_zero_adjust_active = False
        self._calibration_temp_comp_zero_trim_air_zero_adjust_empty_period = None
        self._calibration_temp_comp_zero_trim_air_zero_adjust_full_period = None
        self._calibration_temp_comp_zero_trim_air_zero_adjust_level_x10 = None
        self._calibration_temp_comp_zero_trim_air_zero_adjust_level_samples = []
        self._calibration_temp_comp_zero_trim_air_zero_adjust_current_zero_trim = None

    def _reset_calibration_temp_comp_zero_trim_verify_state(self):
        """Цель функции в безопасном завершении автопроверки zero trim, затем она сбрасывает флаг и таймер ожидания DID 0x0018."""
        if self._calibration_temp_comp_zero_trim_verify_timeout_timer.isActive():
            self._calibration_temp_comp_zero_trim_verify_timeout_timer.stop()
        self._calibration_temp_comp_zero_trim_verify_pending = False
        self._calibration_temp_comp_zero_trim_verify_retries_left = 0

    def _request_next_calibration_temp_comp_k0_air_zero_adjust_did(self) -> bool:
        """Цель функции в пошаговом чтении DID для автоподстройки K0, затем она отправляет только следующий необходимый запрос."""
        if not bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
            return False

        next_var = None
        if self._calibration_temp_comp_k0_air_zero_adjust_empty_period is None:
            next_var = UdsData.empty_fuel_tank
        elif self._calibration_temp_comp_k0_air_zero_adjust_full_period is None:
            next_var = UdsData.full_fuel_tank
        elif self._calibration_temp_comp_k0_air_zero_adjust_level_x10 is None:
            next_var = UdsData.raw_fuel_level
        elif self._calibration_temp_comp_k0_air_zero_adjust_current_k0 is None:
            next_var = UdsData.fuel_temp_comp_k0_count
        else:
            return False

        self._configure_calibration_uds_services()
        sent = bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                next_var,
            )
        )
        if not sent:
            return False

        self._calibration_temp_comp_k0_air_zero_adjust_timeout_timer.start(
            max(300, int(self._calibration_temp_comp_k0_air_zero_adjust_timeout_ms))
        )
        return True

    def _continue_calibration_temp_comp_k0_air_zero_adjust(self):
        """Цель функции в продолжении автоподстройки K0 после ответа DID, затем она запрашивает следующий DID или завершает расчет."""
        if not bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
            return

        if self._calibration_temp_comp_k0_air_zero_adjust_timeout_timer.isActive():
            self._calibration_temp_comp_k0_air_zero_adjust_timeout_timer.stop()

        missing_data = (
            self._calibration_temp_comp_k0_air_zero_adjust_empty_period is None
            or self._calibration_temp_comp_k0_air_zero_adjust_full_period is None
            or self._calibration_temp_comp_k0_air_zero_adjust_level_x10 is None
            or self._calibration_temp_comp_k0_air_zero_adjust_current_k0 is None
        )
        if missing_data:
            if self._request_next_calibration_temp_comp_k0_air_zero_adjust_did():
                return
            self._reset_calibration_temp_comp_k0_air_zero_adjust_state()
            self._set_calibration_temp_comp_operation_status(
                "Автоподстройка K0 остановлена: не удалось отправить следующий DID-запрос.",
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        self._try_apply_calibration_temp_comp_k0_air_zero_adjust()

    def _request_next_calibration_temp_comp_zero_trim_air_zero_adjust_did(self) -> bool:
        """Цель функции в пошаговом чтении DID для автоподстройки zero trim, затем она отправляет только следующий необходимый запрос."""
        if not bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
            return False

        next_var = None
        required_samples = max(1, int(self._calibration_temp_comp_zero_trim_air_zero_adjust_required_samples))
        level_samples = list(self._calibration_temp_comp_zero_trim_air_zero_adjust_level_samples)
        if self._calibration_temp_comp_zero_trim_air_zero_adjust_empty_period is None:
            next_var = UdsData.empty_fuel_tank
        elif self._calibration_temp_comp_zero_trim_air_zero_adjust_full_period is None:
            next_var = UdsData.full_fuel_tank
        elif len(level_samples) < required_samples:
            next_var = UdsData.raw_fuel_level
        elif self._calibration_temp_comp_zero_trim_air_zero_adjust_current_zero_trim is None:
            next_var = UdsData.fuel_zero_trim_count
        else:
            return False

        self._configure_calibration_uds_services()
        sent = bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                next_var,
            )
        )
        if not sent:
            return False

        self._calibration_temp_comp_zero_trim_air_zero_adjust_timeout_timer.start(
            max(300, int(self._calibration_temp_comp_zero_trim_air_zero_adjust_timeout_ms))
        )
        return True

    def _continue_calibration_temp_comp_zero_trim_air_zero_adjust(self):
        """Цель функции в продолжении автоподстройки zero trim после ответа DID, затем она запрашивает следующий DID или завершает расчет."""
        if not bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
            return

        if self._calibration_temp_comp_zero_trim_air_zero_adjust_timeout_timer.isActive():
            self._calibration_temp_comp_zero_trim_air_zero_adjust_timeout_timer.stop()

        required_samples = max(1, int(self._calibration_temp_comp_zero_trim_air_zero_adjust_required_samples))
        sample_count = len(self._calibration_temp_comp_zero_trim_air_zero_adjust_level_samples)
        missing_data = (
            self._calibration_temp_comp_zero_trim_air_zero_adjust_empty_period is None
            or self._calibration_temp_comp_zero_trim_air_zero_adjust_full_period is None
            or sample_count < required_samples
            or self._calibration_temp_comp_zero_trim_air_zero_adjust_current_zero_trim is None
        )
        if missing_data:
            if self._request_next_calibration_temp_comp_zero_trim_air_zero_adjust_did():
                return
            self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
            self._set_calibration_temp_comp_operation_status(
                "Автоподстройка zero trim остановлена: не удалось отправить следующий DID-запрос.",
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        self._try_apply_calibration_temp_comp_zero_trim_air_zero_adjust()

    def _on_calibration_temp_comp_zero_trim_air_zero_adjust_timeout(self):
        """Цель функции в защите от зависания автоподстройки zero trim, затем она завершает операцию по таймауту шага чтения DID."""
        if not bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
            return
        self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
        self._set_calibration_temp_comp_operation_status(
            "Автоподстройка zero trim остановлена по таймауту чтения DID. Повторите операцию.",
            busy=False,
            progress_percent=100,
            determinate=True,
        )
        self._append_log(
            "Калибровка: автоподстройка zero trim остановлена по таймауту (ожидание DID 0x0012/0x0013/0x0018/0x002D).",
            RowColor.red,
        )

    def _try_apply_calibration_temp_comp_zero_trim_air_zero_adjust(self):
        """Цель функции в завершении автоподстройки zero trim, затем она вычисляет значение по фактическому уровню DID 0x0018 и отправляет DID 0x002D."""
        if not bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
            return

        empty_period = self._calibration_temp_comp_zero_trim_air_zero_adjust_empty_period
        full_period = self._calibration_temp_comp_zero_trim_air_zero_adjust_full_period
        level_samples = list(self._calibration_temp_comp_zero_trim_air_zero_adjust_level_samples)
        current_zero_trim = self._calibration_temp_comp_zero_trim_air_zero_adjust_current_zero_trim
        if (
            empty_period is None
            or full_period is None
            or len(level_samples) <= 0
            or current_zero_trim is None
        ):
            return

        current_level_x10 = int(round(sum(level_samples) / float(len(level_samples))))
        sample_spread_x10 = int(max(level_samples) - min(level_samples))
        stability_threshold_x10 = max(0, int(self._calibration_temp_comp_zero_trim_air_zero_adjust_stability_threshold_x10))
        if sample_spread_x10 > stability_threshold_x10:
            self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
            self._set_calibration_temp_comp_operation_status(
                (
                    "Автоподстройка zero trim остановлена: сигнал уровня нестабилен, "
                    f"разброс {float(sample_spread_x10) / 10.0:.1f}% выше порога {float(stability_threshold_x10) / 10.0:.1f}%."
                ),
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            self._append_log(
                (
                    "Калибровка: автоподстройка zero trim прервана из-за нестабильного уровня "
                    f"(n={len(level_samples)}, разброс={float(sample_spread_x10) / 10.0:.1f}%)."
                ),
                RowColor.yellow,
            )
            return

        span = int(full_period) - int(empty_period)
        if span <= 0:
            self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
            self._set_calibration_temp_comp_operation_status(
                "Автоподстройка zero trim остановлена: некорректные границы 0%/100% (DID 0x0012/0x0013).",
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        current_zero_trim_int = int(current_zero_trim)
        delta_zero_trim, target_zero_trim, residual_x10 = self._calculate_zero_trim_adjustment(
            span_count=int(span),
            current_level_x10=int(current_level_x10),
            current_zero_trim=current_zero_trim_int,
        )
        normalized_level_x10 = max(-1000, min(1000, int(current_level_x10)))
        self._calibration_temp_comp_zero_trim_count_recommended = int(target_zero_trim)
        self._calibration_temp_comp_zero_trim_count_delta = int(delta_zero_trim)
        self._calibration_temp_comp_zero_trim_count_next = int(target_zero_trim)
        self._calibration_temp_comp_zero_trim_residual_x10 = int(residual_x10)
        self._set_calibration_temp_comp_zero_trim_last_report(
            old_zero_trim=int(current_zero_trim_int),
            new_zero_trim=int(target_zero_trim),
            residual_x10=int(residual_x10),
            status_text="расчет выполнен, запись отправлена",
        )
        self._reset_calibration_temp_comp_zero_trim_air_zero_adjust_state()
        self._append_log(
            (
                "Калибровка: автоподстройка zero trim по фактическому уровню МК: "
                f"level={int(normalized_level_x10) / 10.0:.1f}%, span={int(span)} count, "
                f"dTrim={int(delta_zero_trim):+d}, TrimNew={int(current_zero_trim_int)}+{int(delta_zero_trim):+d}={int(target_zero_trim)}, "
                f"остаток={float(residual_x10) / 10.0:+.1f}%, n={len(level_samples)}, разброс={float(sample_spread_x10) / 10.0:.1f}%."
            ),
            RowColor.blue,
        )

        pending_key = int(UdsData.fuel_zero_trim_count.pid)
        pending_before = pending_key in self._calibration_write_verify_pending
        self.writeCalibrationTempCompZeroTrim(str(int(target_zero_trim)))
        pending_after = pending_key in self._calibration_write_verify_pending

        if (not pending_before) and pending_after:
            self._calibration_temp_comp_zero_trim_verify_retries_left = max(
                0,
                int(self._calibration_temp_comp_zero_trim_verify_retries_max),
            )
            self._calibration_temp_comp_zero_trim_verify_pending = True
            self._calibration_temp_comp_zero_trim_verify_timeout_timer.start(
                max(500, int(self._calibration_temp_comp_zero_trim_verify_timeout_ms))
            )
            self._set_calibration_temp_comp_operation_status(
                (
                    "Автоподстройка zero trim: рассчитано и отправлено значение "
                    f"{int(target_zero_trim)} в DID 0x002D. Ожидается автопроверка."
                ),
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        self._reset_calibration_temp_comp_zero_trim_verify_state()
        self._set_calibration_temp_comp_operation_status(
            "Автоподстройка zero trim завершена с ошибкой: запись DID 0x002D не подтверждена.",
            busy=False,
            progress_percent=100,
            determinate=True,
        )

    def _on_calibration_temp_comp_k0_air_zero_adjust_timeout(self):
        """Цель функции в защите от зависания автоподстройки K0, затем она завершает операцию по таймауту шага чтения DID."""
        if not bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
            return
        self._reset_calibration_temp_comp_k0_air_zero_adjust_state()
        self._set_calibration_temp_comp_operation_status(
            "Автоподстройка K0 остановлена по таймауту чтения DID. Повторите операцию.",
            busy=False,
            progress_percent=100,
            determinate=True,
        )
        self._append_log(
            "Калибровка: автоподстройка K0 остановлена по таймауту (ожидание DID 0x0012/0x0013/0x0018/0x001C).",
            RowColor.red,
        )

    def _try_apply_calibration_temp_comp_k0_air_zero_adjust(self):
        """Цель функции в завершении автоподстройки K0, затем она вычисляет K0 по фактическому уровню DID 0x0018 и отправляет DID 0x001C."""
        if not bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
            return

        empty_period = self._calibration_temp_comp_k0_air_zero_adjust_empty_period
        full_period = self._calibration_temp_comp_k0_air_zero_adjust_full_period
        current_level_x10 = self._calibration_temp_comp_k0_air_zero_adjust_level_x10
        current_k0 = self._calibration_temp_comp_k0_air_zero_adjust_current_k0
        if (
            empty_period is None
            or full_period is None
            or current_level_x10 is None
            or current_k0 is None
        ):
            return

        span = int(full_period) - int(empty_period)
        if span <= 0:
            self._reset_calibration_temp_comp_k0_air_zero_adjust_state()
            self._set_calibration_temp_comp_operation_status(
                "Автоподстройка K0 остановлена: некорректные границы 0%/100% (DID 0x0012/0x0013).",
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        normalized_level_x10 = max(0, min(1000, int(current_level_x10)))
        delta_k0 = self._c_trunc_div((-int(normalized_level_x10)) * int(span), 1000)
        target_k0 = self._saturate_int16(int(current_k0) + int(delta_k0))
        self._reset_calibration_temp_comp_k0_air_zero_adjust_state()
        self._append_log(
            (
                "Калибровка: автоподстройка K0 по фактическому уровню МК: "
                f"level={int(normalized_level_x10) / 10.0:.1f}%, span={int(span)} count, "
                f"dK0={int(delta_k0)}, K0new={int(current_k0)}+{int(delta_k0)}={int(target_k0)}."
            ),
            RowColor.blue,
        )

        pending_key = int(UdsData.fuel_temp_comp_k0_count.pid)
        pending_before = pending_key in self._calibration_write_verify_pending
        self.writeCalibrationTempCompK0(str(int(target_k0)))
        pending_after = pending_key in self._calibration_write_verify_pending

        if (not pending_before) and pending_after:
            self._set_calibration_temp_comp_operation_status(
                (
                    "Автоподстройка K0: рассчитано и отправлено значение "
                    f"{int(target_k0)} в DID 0x001C. Ожидается автопроверка."
                ),
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        self._set_calibration_temp_comp_operation_status(
            "Автоподстройка K0 завершена с ошибкой: запись DID 0x001C не подтверждена.",
            busy=False,
            progress_percent=100,
            determinate=True,
        )

    def _request_calibration_temp_comp_k1_read(self) -> bool:
        """Цель функции в чтении текущего K1 через UDS, затем она отправляет запрос DID 0x001B в выбранный узел."""
        if not self._can.is_connect or not self._can.is_trace:
            return False
        self._configure_calibration_uds_services()
        return bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                UdsData.fuel_temp_comp_k1_x100,
            )
        )

    def _request_calibration_temp_comp_k0_read(self) -> bool:
        """Цель функции в чтении текущего K0 через UDS, затем она отправляет запрос DID 0x001C в выбранный узел."""
        if not self._can.is_connect or not self._can.is_trace:
            return False
        self._configure_calibration_uds_services()
        return bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                UdsData.fuel_temp_comp_k0_count,
            )
        )

    def _request_calibration_temp_comp_zero_trim_read(self) -> bool:
        """Цель функции в чтении текущего zero trim через UDS, затем она отправляет запрос DID 0x002D в выбранный узел."""
        if not self._can.is_connect or not self._can.is_trace:
            return False
        self._configure_calibration_uds_services()
        return bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                UdsData.fuel_zero_trim_count,
            )
        )

    def _request_calibration_temp_comp_raw_level_read(self) -> bool:
        """Цель функции в чтении текущего сырого уровня для автопроверки, затем она отправляет запрос DID 0x0018 в выбранный узел."""
        if not self._can.is_connect or not self._can.is_trace:
            return False
        self._configure_calibration_uds_services()
        return bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                UdsData.raw_fuel_level,
            )
        )

    def _on_calibration_temp_comp_zero_trim_verify_timeout(self):
        """Цель функции в завершении автопроверки zero trim по таймауту, затем она сообщает оператору о необходимости повторить проверку."""
        if not bool(self._calibration_temp_comp_zero_trim_verify_pending):
            return
        retries_left = max(0, int(self._calibration_temp_comp_zero_trim_verify_retries_left))
        if retries_left > 0:
            self._calibration_temp_comp_zero_trim_verify_retries_left = retries_left - 1
            if self._request_calibration_temp_comp_raw_level_read():
                self._calibration_temp_comp_zero_trim_verify_timeout_timer.start(
                    max(500, int(self._calibration_temp_comp_zero_trim_verify_timeout_ms))
                )
                self._set_calibration_temp_comp_operation_status(
                    (
                        "Автопроверка zero trim: DID 0x0018 не получен, выполняется повторный запрос "
                        f"(осталось попыток: {int(self._calibration_temp_comp_zero_trim_verify_retries_left)})."
                    ),
                    busy=True,
                    progress_percent=90,
                    determinate=True,
                )
                self._append_log(
                    "Калибровка: автопроверка zero trim — повторный запрос DID 0x0018 после таймаута.",
                    RowColor.yellow,
                )
                return

        self._set_calibration_temp_comp_zero_trim_last_report(
            old_zero_trim=self._calibration_temp_comp_zero_trim_count_current,
            new_zero_trim=self._calibration_temp_comp_zero_trim_count_next,
            residual_x10=self._calibration_temp_comp_zero_trim_residual_x10,
            status_text="таймаут автопроверки, нужен повтор",
            write_csv=True,
        )
        self._reset_calibration_temp_comp_zero_trim_verify_state()
        self._set_calibration_temp_comp_operation_status(
            "Автопроверка zero trim не завершена: не получен DID 0x0018. Повторите чтение уровня.",
            busy=False,
            progress_percent=100,
            determinate=True,
        )
        self._append_log(
            "Калибровка: автопроверка zero trim остановлена по таймауту ожидания DID 0x0018.",
            RowColor.yellow,
        )

    def _request_calibration_temp_comp_advanced_read(self, field_key: str) -> bool:
        """Цель функции в чтении одного параметра компенсации, затем она отправляет DID по ключу поля."""
        if not self._can.is_connect or not self._can.is_trace:
            return False

        field = self._temp_comp_read_field_by_key(field_key)
        if field is None:
            return False

        field_var = field.get("var")
        if field_var is None:
            return False

        self._configure_calibration_uds_services()
        return bool(
            self._calibration_read_service.read_data_by_identifier(
                self._build_calibration_tx_identifier(),
                field_var,
            )
        )

    def _calibration_temp_comp_advanced_read_field_keys(self) -> list[str]:
        """Цель функции в формировании списка параметров для чтения, затем она возвращает ключи DID 0x001D..0x002C в фиксированном порядке."""
        result: list[str] = []
        for field in self._temp_comp_advanced_fields():
            key = str(field.get("key", "")).strip()
            if not key:
                continue
            result.append(key)
        return result

    @classmethod
    def _calibration_temp_comp_advanced_read_field_keys_for_mode(
        cls,
        mode_value: int,
        *,
        include_mode: bool,
    ) -> list[str]:
        """Цель функции в выборе набора DID для режима компенсации, затем она возвращает только связанные с mode поля."""
        normalized_mode = int(mode_value)
        if (
            normalized_mode < int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)
            or normalized_mode > int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)
        ):
            normalized_mode = int(cls._TEMP_COMP_MODE_SINGLE_LINEAR)

        result: list[str] = []
        if include_mode:
            result.append("mode")

        if normalized_mode == int(cls._TEMP_COMP_MODE_SINGLE_LINEAR):
            return result

        result.extend(cls._TEMP_COMP_ADV_SEGMENT_BORDER_KEYS)
        result.extend(cls._TEMP_COMP_ADV_COOLING_KEYS)

        if normalized_mode == int(cls._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL):
            result.insert(1 if include_mode else 0, "dir_hyst_x10")
            result.extend(cls._TEMP_COMP_ADV_HEATING_KEYS)
        return result

    def _set_calibration_temp_comp_operation_status(
        self,
        text: str,
        *,
        busy: bool,
        progress_percent: int | None = None,
        determinate: bool | None = None,
    ):
        """Цель функции в обновлении статуса операций температурной компенсации, затем она синхронизирует текст, занятость и прогресс для UI."""
        normalized_text = str(text or "").strip()
        if not normalized_text:
            normalized_text = "Ожидание операций."
        normalized_busy = bool(busy)
        normalized_determinate = (
            bool(progress_percent is not None)
            if determinate is None
            else bool(determinate)
        )
        normalized_progress_percent = 0
        if normalized_determinate:
            if progress_percent is None:
                normalized_progress_percent = 100 if not normalized_busy else 0
            else:
                normalized_progress_percent = max(0, min(100, int(progress_percent)))

        changed = (
            str(self._calibration_temp_comp_operation_text) != normalized_text
            or bool(self._calibration_temp_comp_operation_busy) != normalized_busy
            or int(self._calibration_temp_comp_operation_progress_percent) != int(normalized_progress_percent)
            or bool(self._calibration_temp_comp_operation_progress_determinate) != bool(normalized_determinate)
        )
        self._calibration_temp_comp_operation_text = normalized_text
        self._calibration_temp_comp_operation_busy = normalized_busy
        self._calibration_temp_comp_operation_progress_percent = int(normalized_progress_percent)
        self._calibration_temp_comp_operation_progress_determinate = bool(normalized_determinate)
        if changed:
            self.calibrationTempCompChanged.emit()

    def _set_calibration_temp_comp_preview_status(
        self,
        text: str,
        *,
        busy: bool,
        progress_percent: int | None = None,
        determinate: bool | None = None,
    ):
        """Цель функции в обновлении статуса локального превью, затем она синхронизирует отдельный текст и прогресс блока рядом с графиком."""
        normalized_text = str(text or "").strip()
        if not normalized_text:
            normalized_text = "Ожидание превью."
        normalized_busy = bool(busy)
        normalized_determinate = (
            bool(progress_percent is not None)
            if determinate is None
            else bool(determinate)
        )
        normalized_progress_percent = 0
        if normalized_determinate:
            if progress_percent is None:
                normalized_progress_percent = 100 if not normalized_busy else 0
            else:
                normalized_progress_percent = max(0, min(100, int(progress_percent)))

        changed = (
            str(self._calibration_temp_comp_preview_status) != normalized_text
            or bool(self._calibration_temp_comp_preview_busy) != normalized_busy
            or int(self._calibration_temp_comp_preview_progress_percent) != int(normalized_progress_percent)
            or bool(self._calibration_temp_comp_preview_progress_determinate) != bool(normalized_determinate)
        )
        self._calibration_temp_comp_preview_status = normalized_text
        self._calibration_temp_comp_preview_busy = normalized_busy
        self._calibration_temp_comp_preview_progress_percent = int(normalized_progress_percent)
        self._calibration_temp_comp_preview_progress_determinate = bool(normalized_determinate)
        if changed:
            self.calibrationTempCompChanged.emit()

    def _build_calibration_temp_comp_advanced_read_progress_text(self, *, prefix: str) -> str:
        """Цель функции в формировании человеко-понятного прогресса чтения DID, затем она возвращает строку статуса для панели калибровки."""
        total = int(self._calibration_temp_comp_adv_read_total_count)
        success = int(self._calibration_temp_comp_adv_read_success_count)
        inflight_key = self._calibration_temp_comp_adv_read_inflight_key
        waiting_count = len(self._calibration_temp_comp_adv_read_queue) + (1 if inflight_key is not None else 0)
        waiting_text = f", в очереди: {int(waiting_count)}"
        if inflight_key is not None:
            inflight_field = self._temp_comp_read_field_by_key(str(inflight_key))
            if inflight_field is not None:
                waiting_text += f", читается: {self._temp_comp_field_display_name(inflight_field)}"
        return f"{str(prefix).strip()} {success}/{total}{waiting_text}."

    def _calibration_temp_comp_advanced_read_progress_percent(self) -> int:
        """Цель функции в расчете прогресса чтения DID, затем она возвращает процент завершения очереди независимо от успешности отдельных ответов."""
        total = max(0, int(self._calibration_temp_comp_adv_read_total_count))
        if total <= 0:
            return 0
        waiting_count = len(self._calibration_temp_comp_adv_read_queue) + (
            1 if self._calibration_temp_comp_adv_read_inflight_key is not None else 0
        )
        processed_count = max(0, total - int(waiting_count))
        return max(0, min(100, int(round((float(processed_count) / float(total)) * 100.0))))

    def _stop_calibration_temp_comp_advanced_read_sequence(self):
        """Цель функции в безопасной остановке очереди чтения, затем она сбрасывает состояние и таймеры последовательного опроса."""
        had_active_sequence = bool(self._calibration_temp_comp_adv_read_active) or int(self._calibration_temp_comp_adv_read_total_count) > 0
        pending_recompute = bool(self._calibration_temp_comp_adv_read_recompute_pending)
        if self._calibration_temp_comp_adv_read_timeout_timer.isActive():
            self._calibration_temp_comp_adv_read_timeout_timer.stop()
        if self._calibration_temp_comp_adv_read_delay_timer.isActive():
            self._calibration_temp_comp_adv_read_delay_timer.stop()
        self._calibration_temp_comp_adv_read_active = False
        self._calibration_temp_comp_adv_read_queue = []
        self._calibration_temp_comp_adv_read_inflight_key = None
        self._calibration_temp_comp_adv_read_total_count = 0
        self._calibration_temp_comp_adv_read_success_count = 0
        self._calibration_temp_comp_adv_read_mode_aware = False
        self._calibration_temp_comp_adv_read_expected_mode = None
        self._calibration_temp_comp_adv_read_recompute_pending = False
        if pending_recompute:
            self._recompute_calibration_temp_comp_metrics()
        if had_active_sequence:
            self._set_calibration_temp_comp_operation_status(
                "Чтение параметров из МК остановлено.",
                busy=False,
                progress_percent=0,
                determinate=False,
            )

    def _start_calibration_temp_comp_advanced_read_sequence(
        self,
        field_keys: list[str],
        *,
        mode_aware: bool = False,
        expected_mode: int | None = None,
    ) -> tuple[int, int]:
        """Цель функции в запуске последовательного чтения DID, затем она активирует очередь запросов с контролем таймаутов."""
        if not self._can.is_connect or not self._can.is_trace:
            return 0, len(field_keys)

        unique_keys: list[str] = []
        used_keys: set[str] = set()
        for raw_key in field_keys:
            field_key = str(raw_key or "").strip()
            if not field_key or field_key in used_keys:
                continue
            if self._temp_comp_read_field_by_key(field_key) is None:
                continue
            used_keys.add(field_key)
            unique_keys.append(field_key)

        total_count = len(unique_keys)
        if total_count <= 0:
            self._stop_calibration_temp_comp_advanced_read_sequence()
            return 0, 0

        self._stop_calibration_temp_comp_advanced_read_sequence()
        self._calibration_temp_comp_adv_read_active = True
        self._calibration_temp_comp_adv_read_queue = list(unique_keys)
        self._calibration_temp_comp_adv_read_inflight_key = None
        self._calibration_temp_comp_adv_read_total_count = total_count
        self._calibration_temp_comp_adv_read_success_count = 0
        self._calibration_temp_comp_adv_read_mode_aware = bool(mode_aware)
        if expected_mode is None:
            self._calibration_temp_comp_adv_read_expected_mode = None
        else:
            self._calibration_temp_comp_adv_read_expected_mode = int(expected_mode)
        self._set_calibration_temp_comp_operation_status(
            self._build_calibration_temp_comp_advanced_read_progress_text(
                prefix="Чтение параметров из МК:",
            ),
            busy=True,
            progress_percent=self._calibration_temp_comp_advanced_read_progress_percent(),
            determinate=True,
        )
        self._calibration_temp_comp_adv_read_delay_timer.start(0)
        return total_count, total_count

    def _adapt_calibration_temp_comp_mode_aware_queue(self, actual_mode: int):
        """Цель функции в адаптации очереди к фактическому mode, затем она оставляет только релевантные DID для выбранного режима."""
        if not self._calibration_temp_comp_adv_read_mode_aware:
            return
        if int(self._calibration_temp_comp_adv_read_success_count) <= 0:
            return

        normalized_mode = int(actual_mode)
        if (
            normalized_mode < int(self._TEMP_COMP_MODE_SINGLE_LINEAR)
            or normalized_mode > int(self._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)
        ):
            normalized_mode = int(self._TEMP_COMP_MODE_SINGLE_LINEAR)

        expected_mode = self._calibration_temp_comp_adv_read_expected_mode
        if expected_mode is not None and int(expected_mode) == normalized_mode:
            return

        self._calibration_temp_comp_adv_read_expected_mode = int(normalized_mode)
        remaining_keys = self._calibration_temp_comp_advanced_read_field_keys_for_mode(
            normalized_mode,
            include_mode=False,
        )
        self._calibration_temp_comp_adv_read_queue = list(remaining_keys)
        self._calibration_temp_comp_adv_read_total_count = int(self._calibration_temp_comp_adv_read_success_count) + len(remaining_keys)

    def _schedule_next_calibration_temp_comp_advanced_read(self, *, immediate: bool):
        """Цель функции в планировании следующего запроса из очереди, затем она запускает таймер задержки с нужным интервалом."""
        if not self._calibration_temp_comp_adv_read_active:
            return
        if self._calibration_temp_comp_adv_read_timeout_timer.isActive():
            self._calibration_temp_comp_adv_read_timeout_timer.stop()
        delay_ms = 0 if immediate else int(self._calibration_temp_comp_adv_read_delay_ms)
        self._calibration_temp_comp_adv_read_delay_timer.start(max(0, delay_ms))

    def _on_calibration_temp_comp_advanced_read_delay_timeout(self):
        """Цель функции в отправке следующего запроса DID, затем она переводит очередь в ожидание ответа от МК."""
        if not self._calibration_temp_comp_adv_read_active:
            return
        if self._calibration_temp_comp_adv_read_inflight_key is not None:
            return

        if len(self._calibration_temp_comp_adv_read_queue) <= 0:
            total = int(self._calibration_temp_comp_adv_read_total_count)
            success = int(self._calibration_temp_comp_adv_read_success_count)
            self._append_log(
                f"Калибровка: чтение расширенных DID завершено ({success}/{total}).",
                RowColor.green if success == total else RowColor.yellow,
            )
            self._calibration_temp_comp_adv_read_active = False
            self._calibration_temp_comp_adv_read_mode_aware = False
            self._calibration_temp_comp_adv_read_expected_mode = None
            pending_recompute = bool(self._calibration_temp_comp_adv_read_recompute_pending)
            self._calibration_temp_comp_adv_read_recompute_pending = False
            if pending_recompute:
                self._recompute_calibration_temp_comp_metrics()
            self._set_calibration_temp_comp_operation_status(
                f"Чтение параметров из МК завершено: {success}/{total}.",
                busy=False,
                progress_percent=100,
                determinate=True,
            )
            return

        field_key = str(self._calibration_temp_comp_adv_read_queue.pop(0))
        field = self._temp_comp_read_field_by_key(field_key)
        if field is None:
            self._schedule_next_calibration_temp_comp_advanced_read(immediate=True)
            return

        if not self._request_calibration_temp_comp_advanced_read(field_key):
            self._append_log(
                f"Калибровка: не удалось отправить чтение {self._temp_comp_field_display_name(field)}.",
                RowColor.red,
            )
            self._schedule_next_calibration_temp_comp_advanced_read(immediate=False)
            return

        self._calibration_temp_comp_adv_read_inflight_key = field_key
        self._calibration_temp_comp_adv_read_timeout_timer.start(
            max(200, int(self._calibration_temp_comp_adv_read_timeout_ms))
        )
        self._set_calibration_temp_comp_operation_status(
            self._build_calibration_temp_comp_advanced_read_progress_text(
                prefix="Чтение параметров из МК:",
            ),
            busy=True,
            progress_percent=self._calibration_temp_comp_advanced_read_progress_percent(),
            determinate=True,
        )

    def _on_calibration_temp_comp_advanced_read_timeout(self):
        """Цель функции в обработке таймаута чтения DID, затем она фиксирует пропуск и продолжает очередь."""
        if not self._calibration_temp_comp_adv_read_active:
            return

        inflight_key = self._calibration_temp_comp_adv_read_inflight_key
        self._calibration_temp_comp_adv_read_inflight_key = None
        if inflight_key is not None:
            field = self._temp_comp_read_field_by_key(inflight_key)
            label = "параметра"
            if field is not None:
                label = self._temp_comp_field_display_name(field)
            self._append_log(
                f"Калибровка: таймаут чтения {label}. Переход к следующему DID.",
                RowColor.yellow,
            )
            self._set_calibration_temp_comp_operation_status(
                self._build_calibration_temp_comp_advanced_read_progress_text(
                    prefix="Чтение параметров из МК (таймаут, продолжаем):",
                ),
                busy=True,
                progress_percent=self._calibration_temp_comp_advanced_read_progress_percent(),
                determinate=True,
            )

        self._schedule_next_calibration_temp_comp_advanced_read(immediate=False)

    def _request_calibration_temp_comp_advanced_read_all(self) -> tuple[int, int]:
        """Цель функции в запуске очереди чтения расширенной компенсации, затем она выполняет DID 0x001D..0x002C по одному."""
        field_keys = self._calibration_temp_comp_advanced_read_field_keys()
        return self._start_calibration_temp_comp_advanced_read_sequence(
            field_keys,
            mode_aware=False,
            expected_mode=None,
        )

    def _request_calibration_temp_comp_advanced_read_for_mode(
        self,
        mode_value: int,
        *,
        include_base: bool = False,
    ) -> tuple[int, int]:
        """Цель функции в запуске чтения параметров режима, затем она собирает очередь DID и выполняет их строго последовательно."""
        normalized_mode = int(mode_value)
        if (
            normalized_mode < int(self._TEMP_COMP_MODE_SINGLE_LINEAR)
            or normalized_mode > int(self._TEMP_COMP_MODE_SEGMENTED_HEAT_COOL)
        ):
            normalized_mode = int(self._TEMP_COMP_MODE_SINGLE_LINEAR)

        field_keys = self._calibration_temp_comp_advanced_read_field_keys_for_mode(
            normalized_mode,
            include_mode=True,
        )
        if include_base:
            field_keys = list(self._TEMP_COMP_BASE_READ_KEYS) + list(field_keys)
        self._calibration_temp_comp_adv_read_recompute_pending = False
        return self._start_calibration_temp_comp_advanced_read_sequence(
            field_keys,
            mode_aware=True,
            expected_mode=normalized_mode,
        )

    def _try_complete_calibration_temp_comp_advanced_read_step(
        self,
        *,
        did: int,
        value: int,
    ) -> bool:
        """Цель функции в подтверждении шага очереди чтения, затем она сравнивает DID ответа с ожидаемым и запускает следующий запрос."""
        if not self._calibration_temp_comp_adv_read_active:
            return False

        inflight_key = self._calibration_temp_comp_adv_read_inflight_key
        if inflight_key is None:
            return False

        inflight_field = self._temp_comp_read_field_by_key(str(inflight_key))
        if inflight_field is None:
            return False

        inflight_var = inflight_field.get("var")
        if inflight_var is None:
            return False

        expected_did = int(inflight_var.pid) & 0xFFFF
        if expected_did != (int(did) & 0xFFFF):
            return False

        if self._calibration_temp_comp_adv_read_timeout_timer.isActive():
            self._calibration_temp_comp_adv_read_timeout_timer.stop()
        self._calibration_temp_comp_adv_read_inflight_key = None
        self._calibration_temp_comp_adv_read_success_count += 1
        self._set_calibration_temp_comp_operation_status(
            self._build_calibration_temp_comp_advanced_read_progress_text(
                prefix="Чтение параметров из МК:",
            ),
            busy=True,
            progress_percent=self._calibration_temp_comp_advanced_read_progress_percent(),
            determinate=True,
        )
        if str(inflight_key) == "mode":
            self._adapt_calibration_temp_comp_mode_aware_queue(int(value))
        self._schedule_next_calibration_temp_comp_advanced_read(immediate=False)
        return True

    def _reset_calibration_temp_comp_recommendation_apply_queue(self):
        """Цель функции в остановке цепочки записи рекомендаций, затем она очищает очередь шагов K1/K0."""
        self._calibration_temp_comp_recommendation_apply_queue = []

    def _continue_calibration_temp_comp_recommendation_apply_queue(self):
        """Цель функции в поэтапной записи рекомендаций, затем она запускает следующий шаг после завершения автопроверки."""
        if len(self._calibration_temp_comp_recommendation_apply_queue) <= 0:
            return
        if len(self._calibration_write_verify_pending) > 0:
            return

        while len(self._calibration_temp_comp_recommendation_apply_queue) > 0:
            if len(self._calibration_write_verify_pending) > 0:
                return

            step = str(self._calibration_temp_comp_recommendation_apply_queue.pop(0)).strip().lower()
            pending_before = len(self._calibration_write_verify_pending)

            if step == "k1":
                value = self._calibration_temp_comp_k1_x100_next
                if value is None:
                    self._append_log(
                        "Калибровка: шаг рекомендаций K1 пропущен, значение не рассчитано.",
                        RowColor.yellow,
                    )
                    continue
                self._append_log(f"Калибровка: запись рекомендованного K1 = {int(value)}.", RowColor.blue)
                self.writeCalibrationTempCompK1(str(int(value)))
            elif step == "k0":
                value = self._calibration_temp_comp_k0_count_next
                if value is None:
                    self._append_log(
                        "Калибровка: шаг рекомендаций K0 пропущен, значение не рассчитано.",
                        RowColor.yellow,
                    )
                    continue
                self._append_log(f"Калибровка: запись рекомендованного K0 = {int(value)}.", RowColor.blue)
                self.writeCalibrationTempCompK0(str(int(value)))
            elif step.startswith("adv:"):
                field_key = str(step[4:]).strip()
                if not field_key:
                    continue
                field = self._temp_comp_advanced_field_by_key(field_key)
                if field is None:
                    self._append_log(
                        f"Калибровка: шаг рекомендаций {field_key} пропущен, параметр не найден в карте DID.",
                        RowColor.yellow,
                    )
                    continue
                value = self._calibration_temp_comp_advanced_recommended_values.get(field_key)
                if value is None:
                    self._append_log(
                        f"Калибровка: шаг рекомендаций {field_key} пропущен, значение не рассчитано.",
                        RowColor.yellow,
                    )
                    continue
                self._append_log(
                    f"Калибровка: запись рекомендации {self._temp_comp_field_display_name(field)} = {int(value)}.",
                    RowColor.blue,
                )
                self.writeCalibrationTempCompAdvancedParam(field_key, str(int(value)))
            else:
                continue

            if len(self._calibration_write_verify_pending) > pending_before:
                return

            self._append_log(
                "Калибровка: цепочка записи рекомендаций остановлена, запрос не был принят МК.",
                RowColor.red,
            )
            self._reset_calibration_temp_comp_recommendation_apply_queue()
            return

    def _handle_calibration_frame(self, identifier: int, payload: list[int]):
        if len(payload) < 2:
            return

        self._try_bind_calibration_runtime_target_from_session_response(identifier, payload)
        if not self._is_calibration_response_identifier(identifier):
            return

        if payload[1] == 0x7F and len(payload) >= 4:
            original_sid = int(payload[2]) & 0xFF
            nrc = int(payload[3]) & 0xFF
            nrc_text = self._uds_nrc_description(nrc)

            if self._calibration_waiting_session and original_sid == 0x10:
                current_action = str(self._calibration_sequence_waiting_action or "")
                if current_action == "deactivate_session":
                    self._append_log(
                        f"Калибровка: ошибка возврата в default-сессию, NRC 0x{nrc:02X} ({nrc_text}).",
                        RowColor.red,
                    )
                    self._finish_calibration_deactivation(
                        f"Калибровка: default-сессия не подтверждена, NRC 0x{nrc:02X} ({nrc_text})."
                    )
                    return

                self._fail_calibration_activation(
                    f"Калибровка: ошибка смены сессии, NRC 0x{nrc:02X} ({nrc_text})."
                )
                return

            if original_sid == 0x27:
                target_sa = self._resolve_calibration_target_sa()
                self._set_calibration_service_access_state(
                    busy=False,
                    pending_action="",
                    unlocked=False,
                    target_sa=target_sa,
                    status=f"Калибровка: отказ Security Access, NRC 0x{nrc:02X} ({nrc_text}).",
                )
                self._fail_calibration_activation(
                    f"Калибровка: Security Access отклонён, NRC 0x{nrc:02X} ({nrc_text})."
                )
                return

            if original_sid == 0x2E:
                did = self._pending_calibration_write_did()
                if (did is None) and (not self._calibration_restore_active):
                    # Игнорируем чужие ответы 0x2E (например, запись DID из других карточек UI).
                    return

                self._reset_calibration_temp_comp_recommendation_apply_queue()
                did_label = "параметра"
                extra_hint = ""
                expected_value = None

                if did is not None:
                    expected_value = self._calibration_write_verify_pending.get(int(did))
                    did_label = self._calibration_did_label(did)
                    self._calibration_write_verify_pending.pop(int(did), None)
                    if int(did) == int(UdsData.empty_fuel_tank.pid):
                        self._calibration_level0_written = False
                        self._calibration_verify0_ok = False
                    elif int(did) == int(UdsData.full_fuel_tank.pid):
                        self._calibration_level100_written = False
                        self._calibration_verify100_ok = False
                    self.calibrationVerificationChanged.emit()

                if self._calibration_restore_active:
                    self._calibration_restore_active = False
                    self._calibration_restore_current_did = None

                if nrc == 0x33:
                    target_sa = self._resolve_calibration_target_sa()
                    if self._service_access_target_sa is None:
                        extra_hint = f" Откройте Security Access 0x27 для узла 0x{target_sa:02X}."
                    else:
                        access_target_sa = int(self._service_access_target_sa) & 0xFF
                        if access_target_sa != target_sa:
                            extra_hint = (
                                f" Сейчас 0x27 открыт для узла 0x{access_target_sa:02X}, "
                                f"а запись идёт в узел 0x{target_sa:02X}."
                            )
                        else:
                            extra_hint = f" Повторно выполните Security Access 0x27 для узла 0x{target_sa:02X}."
                elif nrc == 0x22 and did is not None:
                    extra_hint = (
                        " Проверьте условия записи на стороне МК: "
                        "активная сессия, Security Access, состояние приложения/загрузчика и внутренние блокировки DID."
                    )
                elif nrc in (0x7E, 0x7F):
                    extra_hint = (
                        " На МК неактивна требуемая диагностическая сессия. "
                        "Повторно запустите калибровку (0x10/0x27) и повторите запись."
                    )

                self._append_log(
                    f"Калибровка: запись {did_label} отклонена, NRC 0x{nrc:02X} ({nrc_text}).{extra_hint}",
                    RowColor.red,
                )
                if nrc in (0x7E, 0x7F):
                    self._invalidate_calibration_session_after_nrc(
                        "Калибровка: сессия на МК завершена или недоступна. Перезапустите калибровку перед записью параметров."
                    )
                self._recompute_calibration_wizard_state()
                return

            if original_sid == 0x22:
                if (
                    (not bool(self._calibration_dump_capture_active))
                    and (not (self._calibration_temp_comp_adv_read_active and self._calibration_temp_comp_adv_read_inflight_key is not None))
                    and (str(self._calibration_sequence_waiting_action or "") not in ("read_level_0", "read_level_100"))
                ):
                    # Игнорируем чужие ответы 0x22 (например, чтение DID из других карточек UI).
                    return

                if bool(self._calibration_dump_capture_active):
                    did_text = "-"
                    if self._calibration_dump_capture_current_did is not None:
                        did_text = f"0x{int(self._calibration_dump_capture_current_did) & 0xFFFF:04X}"
                    self._finish_calibration_dump_capture(
                        False,
                        f"Калибровка: чтение дампа отклонено, DID {did_text}, NRC 0x{nrc:02X} ({nrc_text}).",
                    )
                    return

                if self._calibration_temp_comp_adv_read_active and self._calibration_temp_comp_adv_read_inflight_key is not None:
                    inflight_key = str(self._calibration_temp_comp_adv_read_inflight_key)
                    self._calibration_temp_comp_adv_read_inflight_key = None
                    if self._calibration_temp_comp_adv_read_timeout_timer.isActive():
                        self._calibration_temp_comp_adv_read_timeout_timer.stop()
                    inflight_field = self._temp_comp_read_field_by_key(inflight_key)
                    if inflight_field is not None:
                        self._append_log(
                            (
                                f"Калибровка: чтение {self._temp_comp_field_display_name(inflight_field)} "
                                f"отклонено, NRC 0x{nrc:02X} ({nrc_text})."
                            ),
                            RowColor.yellow,
                        )
                    self._set_calibration_temp_comp_operation_status(
                        self._build_calibration_temp_comp_advanced_read_progress_text(
                            prefix="Чтение параметров из МК (NRC, продолжаем):",
                        ),
                        busy=True,
                        progress_percent=self._calibration_temp_comp_advanced_read_progress_percent(),
                        determinate=True,
                    )
                    if nrc in (0x7E, 0x7F):
                        self._invalidate_calibration_session_after_nrc(
                            "Калибровка: сессия на МК завершена или недоступна. Перезапустите калибровку перед чтением параметров."
                        )
                        return
                    self._schedule_next_calibration_temp_comp_advanced_read(immediate=False)
                    return

                if self._calibration_sequence_waiting_action in ("read_level_0", "read_level_100"):
                    self._fail_calibration_activation(
                        f"Калибровка: чтение исходных уровней отклонено, NRC 0x{nrc:02X} ({nrc_text})."
                    )
                    return
                self._append_log(
                    f"Калибровка: чтение параметра отклонено, NRC 0x{nrc:02X} ({nrc_text}).",
                    RowColor.red,
                )
                if nrc in (0x7E, 0x7F):
                    self._invalidate_calibration_session_after_nrc(
                        "Калибровка: сессия на МК завершена или недоступна. Перезапустите калибровку перед чтением параметров."
                    )
                return

        if payload[1] == 0x67 and len(payload) >= 3:
            sub_function = int(payload[2]) & 0xFF
            target_sa = self._resolve_calibration_target_sa()

            if sub_function == 0x01 and self._calibration_sequence_waiting_action == "security_seed":
                if self._service_security_access_service.verify_answer_request_seed(payload):
                    self._finish_calibration_sequence_wait("security_seed")
                    self._set_calibration_service_access_state(
                        busy=True,
                        pending_action="calibration_security_key",
                        unlocked=False,
                        target_sa=target_sa,
                        status=f"Калибровка: seed получен, подготовка key для SA 0x{target_sa:02X}...",
                    )
                    self._append_log(
                        (
                            f"Калибровка: seed=0x{int(self._service_security_access_service.seed) & 0xFFFF:04X}, "
                            f"key=0x{int(self._service_security_access_service.key) & 0xFFFF:04X}, "
                            f"узел 0x{target_sa:02X}."
                        ),
                        RowColor.blue,
                    )
                    self._schedule_calibration_sequence_action("send_security_key")
                    return

            if sub_function == 0x02 and self._calibration_sequence_waiting_action == "security_key":
                if self._service_security_access_service.verify_answer_request_check_key(payload):
                    self._finish_calibration_sequence_wait("security_key")
                    self._set_calibration_service_access_state(
                        busy=False,
                        pending_action="",
                        unlocked=True,
                        target_sa=target_sa,
                        status=f"Калибровка: Security Access открыт для SA 0x{target_sa:02X}.",
                    )
                    self._append_log(
                        f"Калибровка: Security Access подтверждён для узла 0x{target_sa:02X}.",
                        RowColor.green,
                    )
                    self._schedule_calibration_sequence_action("read_level_0")
                    return

        if self._calibration_waiting_session:
            if self._calibration_session_service.verify_answer(payload):
                current_action = str(self._calibration_sequence_waiting_action or "")
                self._finish_calibration_sequence_wait(current_action if current_action else None)
                self._calibration_waiting_session = False
                self._calibration_session_ready = False
                if current_action == "deactivate_session":
                    self._append_log("Калибровка: возврат в default-сессию выполнен.", RowColor.green)
                    self._finish_calibration_deactivation("Калибровка завершена. Security Access закрыт, активна default-сессия.")
                    return

                self._append_log("Калибровка: расширенная сессия активирована.", RowColor.green)
                self._schedule_calibration_sequence_action("request_security_seed")
                self._recompute_calibration_wizard_state()
                return

            if payload[1] == 0x7F and len(payload) >= 4 and payload[2] == 0x10:
                self._calibration_waiting_session = False
                self._stop_calibration_poll_timer()
                self._calibration_session_ready = False
                self._calibration_write_verify_pending = {}
                self.calibrationVerificationChanged.emit()
                if self._calibration_active:
                    self._calibration_active = False
                    self.calibrationStateChanged.emit()
                self._append_log(f"Калибровка: ошибка смены сессии (NRC=0x{payload[3]:02X}).", RowColor.red)
                self._recompute_calibration_wizard_state()
                return

        if (int(payload[1]) & 0xFF) == int(self._calibration_read_service.success_sid) and len(payload) >= 4:
            did = (int(payload[2]) << 8) | int(payload[3])
            raw_value = int(self._calibration_read_service.parse_data_field(payload))
            value = int(raw_value)
            changed = False
            emit_temp_comp_signal = False
            recompute_temp_comp = False

            if did == int(UdsData.curr_fuel_tank.pid):
                if self._calibration_current_level != value:
                    self._calibration_current_level = value
                    changed = True
                self._add_calibration_recent_sample(value)
                if self._calibration_temp_comp_last_period != value:
                    self._calibration_temp_comp_last_period = int(value)
                    if len(self._calibration_temp_comp_samples) <= 0:
                        emit_temp_comp_signal = True
            elif did == int(UdsData.empty_fuel_tank.pid):
                self._calibration_level_0 = value
                self._calibration_level_0_known = True
                if bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
                    self._calibration_temp_comp_k0_air_zero_adjust_empty_period = int(value)
                if bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
                    self._calibration_temp_comp_zero_trim_air_zero_adjust_empty_period = int(value)
                changed = True
                recompute_temp_comp = True
                self._append_log(f"Калибровка: считан уровень 0% = {value}.", RowColor.green)
            elif did == int(UdsData.full_fuel_tank.pid):
                self._calibration_level_100 = value
                self._calibration_level_100_known = True
                if bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
                    self._calibration_temp_comp_k0_air_zero_adjust_full_period = int(value)
                if bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
                    self._calibration_temp_comp_zero_trim_air_zero_adjust_full_period = int(value)
                changed = True
                recompute_temp_comp = True
                self._append_log(f"Калибровка: считан уровень 100% = {value}.", RowColor.green)
            elif did == int(UdsData.raw_fuel_level.pid):
                bits = max(8, int(UdsData.raw_fuel_level.size) * 8)
                signed_level = self._decode_signed_value(raw_value, bits)
                value = int(signed_level)
                if bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
                    self._calibration_temp_comp_k0_air_zero_adjust_level_x10 = int(signed_level)
                if bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
                    self._calibration_temp_comp_zero_trim_air_zero_adjust_level_x10 = int(signed_level)
                    samples = list(self._calibration_temp_comp_zero_trim_air_zero_adjust_level_samples)
                    samples.append(int(signed_level))
                    max_samples = max(1, int(self._calibration_temp_comp_zero_trim_air_zero_adjust_required_samples))
                    if len(samples) > max_samples:
                        samples = samples[-max_samples:]
                    self._calibration_temp_comp_zero_trim_air_zero_adjust_level_samples = list(samples)
                if bool(self._calibration_temp_comp_zero_trim_verify_pending):
                    self._reset_calibration_temp_comp_zero_trim_verify_state()
                    self._calibration_temp_comp_zero_trim_residual_x10 = int(signed_level)
                    tolerance_x10 = max(0, int(self._calibration_temp_comp_zero_trim_verify_tolerance_x10))
                    repeat_threshold_x10 = max(
                        tolerance_x10,
                        int(self._calibration_temp_comp_zero_trim_verify_repeat_threshold_x10),
                    )
                    residual_text = f"{float(int(signed_level)) / 10.0:+.1f}%"
                    verification_state = self._classify_zero_trim_verification_result(
                        residual_x10=int(signed_level),
                        tolerance_x10=int(tolerance_x10),
                        repeat_threshold_x10=int(repeat_threshold_x10),
                    )
                    if verification_state == "success":
                        self._set_calibration_temp_comp_zero_trim_last_report(
                            old_zero_trim=self._calibration_temp_comp_zero_trim_count_current,
                            new_zero_trim=self._calibration_temp_comp_zero_trim_count_next,
                            residual_x10=int(signed_level),
                            status_text="успех",
                            write_csv=True,
                        )
                        self._set_calibration_temp_comp_operation_status(
                            f"Автопроверка zero trim: успех, остаток {residual_text}.",
                            busy=False,
                            progress_percent=100,
                            determinate=True,
                        )
                        self._append_log(
                            f"Калибровка: автопроверка zero trim успешна, остаток уровня {residual_text}.",
                            RowColor.green,
                        )
                    elif verification_state == "repeat":
                        self._set_calibration_temp_comp_zero_trim_last_report(
                            old_zero_trim=self._calibration_temp_comp_zero_trim_count_current,
                            new_zero_trim=self._calibration_temp_comp_zero_trim_count_next,
                            residual_x10=int(signed_level),
                            status_text="нужен повтор",
                            write_csv=True,
                        )
                        self._set_calibration_temp_comp_operation_status(
                            f"Автопроверка zero trim: остаток {residual_text} выше допуска, рекомендуется повторить подгонку.",
                            busy=False,
                            progress_percent=100,
                            determinate=True,
                        )
                        self._append_log(
                            f"Калибровка: автопроверка zero trim дала остаток {residual_text}, рекомендуется повторить подгонку.",
                            RowColor.yellow,
                        )
                    else:
                        self._set_calibration_temp_comp_zero_trim_last_report(
                            old_zero_trim=self._calibration_temp_comp_zero_trim_count_current,
                            new_zero_trim=self._calibration_temp_comp_zero_trim_count_next,
                            residual_x10=int(signed_level),
                            status_text="подозрение на механику/датчик",
                            write_csv=True,
                        )
                        self._set_calibration_temp_comp_operation_status(
                            f"Автопроверка zero trim: остаток {residual_text} слишком большой, проверьте механику/датчик.",
                            busy=False,
                            progress_percent=100,
                            determinate=True,
                        )
                        self._append_log(
                            f"Калибровка: автопроверка zero trim выявила большой остаток {residual_text} — требуется проверка механики/датчика.",
                            RowColor.red,
                        )
                    recompute_temp_comp = True
            elif did == int(UdsData.raw_temperature.pid):
                bits = max(8, int(UdsData.raw_temperature.size) * 8)
                signed_temperature = self._decode_signed_value(raw_value, bits)
                value = int(signed_temperature)
                self._calibration_temp_comp_last_temperature_x10 = int(signed_temperature)
                self._calibration_temp_comp_last_temperature_c = float(signed_temperature) / 10.0
                if len(self._calibration_temp_comp_samples) <= 0:
                    emit_temp_comp_signal = True
            elif did == int(UdsData.fuel_temp_comp_k1_x100.pid):
                bits = max(8, int(UdsData.fuel_temp_comp_k1_x100.size) * 8)
                signed_k1 = self._decode_signed_value(raw_value, bits)
                value = int(signed_k1)
                if self._calibration_temp_comp_k1_x100_current != signed_k1:
                    self._calibration_temp_comp_k1_x100_current = int(signed_k1)
                    self._append_log(f"Калибровка: считан коэффициент K1 = {int(signed_k1)}.", RowColor.green)
                if self._calibration_temp_comp_adv_read_active:
                    self._calibration_temp_comp_adv_read_recompute_pending = True
                else:
                    recompute_temp_comp = True
            elif did == int(UdsData.fuel_temp_comp_k0_count.pid):
                bits = max(8, int(UdsData.fuel_temp_comp_k0_count.size) * 8)
                signed_k0 = self._decode_signed_value(raw_value, bits)
                value = int(signed_k0)
                if self._calibration_temp_comp_k0_count_current != signed_k0:
                    self._calibration_temp_comp_k0_count_current = int(signed_k0)
                    self._append_log(f"Калибровка: считан коэффициент K0 = {int(signed_k0)}.", RowColor.green)
                if bool(self._calibration_temp_comp_k0_air_zero_adjust_active):
                    self._calibration_temp_comp_k0_air_zero_adjust_current_k0 = int(signed_k0)
                if self._calibration_temp_comp_adv_read_active:
                    self._calibration_temp_comp_adv_read_recompute_pending = True
                else:
                    recompute_temp_comp = True
            elif did == int(UdsData.fuel_zero_trim_count.pid):
                bits = max(8, int(UdsData.fuel_zero_trim_count.size) * 8)
                signed_zero_trim = self._decode_signed_value(raw_value, bits)
                value = int(signed_zero_trim)
                previous_zero_trim = self._calibration_temp_comp_zero_trim_count_current
                self._calibration_temp_comp_zero_trim_count_current = int(signed_zero_trim)
                if previous_zero_trim != signed_zero_trim:
                    self._append_log(
                        f"Калибровка: считана коррекция zero trim = {int(signed_zero_trim)}.",
                        RowColor.green,
                    )
                else:
                    self._append_log(
                        f"Калибровка: подтверждено значение zero trim = {int(signed_zero_trim)} (без изменений).",
                        RowColor.blue,
                    )
                if bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active):
                    self._calibration_temp_comp_zero_trim_air_zero_adjust_current_zero_trim = int(signed_zero_trim)
                if self._calibration_temp_comp_adv_read_active:
                    self._calibration_temp_comp_adv_read_recompute_pending = True
                else:
                    recompute_temp_comp = True
            else:
                advanced_field = self._temp_comp_advanced_field_by_did(did)
                if advanced_field is not None:
                    field_var = advanced_field.get("var")
                    field_key = str(advanced_field.get("key", ""))
                    bits = 16
                    if field_var is not None:
                        bits = max(8, int(field_var.size) * 8)

                    if bool(advanced_field.get("signed", False)):
                        parsed_value = self._decode_signed_value(raw_value, bits)
                    else:
                        parsed_value = int(raw_value) & ((1 << bits) - 1)
                    value = int(parsed_value)

                    previous_value = self._calibration_temp_comp_advanced_values.get(field_key)
                    advanced_value_changed = previous_value != value
                    if previous_value != value:
                        self._calibration_temp_comp_advanced_values[field_key] = int(value)
                        label_text = str(advanced_field.get("label", "Параметр"))
                        if field_key == "mode":
                            formatted = self._temp_comp_mode_text(value)
                            self._append_log(f"Калибровка: считан {label_text} = {formatted}.", RowColor.green)
                        else:
                            self._append_log(f"Калибровка: считан {label_text} = {int(value)}.", RowColor.green)
                    if advanced_value_changed:
                        if self._calibration_temp_comp_adv_read_active:
                            self._calibration_temp_comp_adv_read_recompute_pending = True
                        else:
                            recompute_temp_comp = True

            self._try_complete_calibration_temp_comp_advanced_read_step(
                did=int(did),
                value=int(value),
            )

            if bool(self._calibration_dump_capture_active):
                required_dids = set(int(item) & 0xFFFF for item in self._calibration_dump_required_dids())
                if (int(did) & 0xFFFF) in required_dids:
                    self._calibration_dump_capture_values[int(did) & 0xFFFF] = int(value)
                    if self._calibration_dump_capture_timeout_timer.isActive():
                        self._calibration_dump_capture_timeout_timer.stop()
                    self._calibration_dump_capture_current_did = None
                    self._request_next_calibration_dump_capture_did()

            if self._calibration_sequence_waiting_action == "read_level_0" and did == int(UdsData.empty_fuel_tank.pid):
                self._finish_calibration_sequence_wait("read_level_0")
                self._schedule_calibration_sequence_action("read_level_100")
            elif self._calibration_sequence_waiting_action == "read_level_100" and did == int(UdsData.full_fuel_tank.pid):
                self._finish_calibration_sequence_wait("read_level_100")
                self._calibration_session_ready = True
                self._start_calibration_poll_timer()
                self._request_calibration_runtime_snapshot()
                self._append_log(
                    "Калибровка: extended-сессия и Security Access активны, исходные уровни считаны. Параметры температурной компенсации читаются только вручную.",
                    RowColor.green,
                )
                self._recompute_calibration_wizard_state()

            if bool(self._calibration_backup_all_nodes_active):
                required_dids = self._calibration_backup_all_nodes_required_dids()
                if did in required_dids:
                    current_sa = self._calibration_backup_all_nodes_current_sa
                    if current_sa is not None:
                        node_key = int(current_sa) & 0xFF
                        node_values = self._calibration_backup_all_nodes_values_by_sa.setdefault(node_key, {})
                        node_values[int(did)] = int(value)
                        if all((required_did in node_values) for required_did in required_dids):
                            if self._calibration_backup_all_nodes_step_timer.isActive():
                                self._calibration_backup_all_nodes_step_timer.stop()
                            self._request_next_calibration_backup_all_nodes_node()

            if (not bool(self._calibration_backup_all_nodes_active)) and self._calibration_backup_pending and did in (int(UdsData.empty_fuel_tank.pid), int(UdsData.full_fuel_tank.pid)):
                self._calibration_backup_values_pending[did] = value
                if (
                    int(UdsData.empty_fuel_tank.pid) in self._calibration_backup_values_pending
                    and int(UdsData.full_fuel_tank.pid) in self._calibration_backup_values_pending
                ):
                    self._calibration_backup_pending = False
                    self._calibration_backup_level_0 = int(self._calibration_backup_values_pending[int(UdsData.empty_fuel_tank.pid)])
                    self._calibration_backup_level_100 = int(self._calibration_backup_values_pending[int(UdsData.full_fuel_tank.pid)])
                    self._calibration_backup_values_pending = {}
                    self._calibration_backup_available = True
                    self.calibrationBackupChanged.emit()
                    self._append_log("Калибровка: резервная копия параметров сохранена.", RowColor.green)

            expected_value = self._calibration_write_verify_pending.get(did)
            if expected_value is not None:
                diff = abs(int(value) - int(expected_value))
                if diff <= int(self._calibration_verify_tolerance):
                    if did == int(UdsData.empty_fuel_tank.pid):
                        self._calibration_verify0_ok = True
                    elif did == int(UdsData.full_fuel_tank.pid):
                        self._calibration_verify100_ok = True
                    self._append_log(
                        f"Калибровка: автопроверка DID 0x{did:04X} успешна (ожидалось {expected_value}, факт {value}).",
                        RowColor.green,
                    )
                else:
                    if did == int(UdsData.empty_fuel_tank.pid):
                        self._calibration_verify0_ok = False
                    elif did == int(UdsData.full_fuel_tank.pid):
                        self._calibration_verify100_ok = False
                    self._append_log(
                        f"Калибровка: автопроверка DID 0x{did:04X} НЕ пройдена (ожидалось {expected_value}, факт {value}).",
                        RowColor.red,
                    )
                self._calibration_write_verify_pending.pop(did, None)
                self.calibrationVerificationChanged.emit()
                self._recompute_calibration_wizard_state()
                self._continue_calibration_temp_comp_recommendation_apply_queue()

            k0_air_zero_adjust_dids = {
                int(UdsData.empty_fuel_tank.pid),
                int(UdsData.full_fuel_tank.pid),
                int(UdsData.raw_fuel_level.pid),
                int(UdsData.fuel_temp_comp_k0_count.pid),
            }
            zero_trim_air_zero_adjust_dids = {
                int(UdsData.empty_fuel_tank.pid),
                int(UdsData.full_fuel_tank.pid),
                int(UdsData.raw_fuel_level.pid),
                int(UdsData.fuel_zero_trim_count.pid),
            }

            if bool(self._calibration_temp_comp_k0_air_zero_adjust_active) and did in k0_air_zero_adjust_dids:
                self._continue_calibration_temp_comp_k0_air_zero_adjust()
            if bool(self._calibration_temp_comp_zero_trim_air_zero_adjust_active) and did in zero_trim_air_zero_adjust_dids:
                self._continue_calibration_temp_comp_zero_trim_air_zero_adjust()

            if changed:
                self.calibrationValuesChanged.emit()
            if recompute_temp_comp:
                self._recompute_calibration_temp_comp_metrics()
            elif emit_temp_comp_signal:
                self.calibrationTempCompChanged.emit()
            return

        if self._calibration_write_service.verify_answer_write_data(payload):
            did = int(self._calibration_write_service.parse_pid_field(payload))
            if did == int(UdsData.empty_fuel_tank.pid):
                self._append_log("Калибровка: уровень 0% успешно сохранен.", RowColor.green)
                self.readCalibrationLevel0()
                self._append_log(
                    "Калибровка: после изменения 0% рекомендуется выполнить эксплуатационную подгонку zero trim.",
                    RowColor.yellow,
                )
            elif did == int(UdsData.full_fuel_tank.pid):
                self._append_log("Калибровка: уровень 100% успешно сохранен.", RowColor.green)
                self.readCalibrationLevel100()
                self._append_log(
                    "Калибровка: после изменения 100% рекомендуется выполнить эксплуатационную подгонку zero trim.",
                    RowColor.yellow,
                )
            elif did == int(UdsData.fuel_temp_comp_k1_x100.pid):
                self._append_log("Калибровка: коэффициент K1 успешно сохранен.", RowColor.green)
                self._request_calibration_temp_comp_k1_read()
            elif did == int(UdsData.fuel_temp_comp_k0_count.pid):
                self._append_log("Калибровка: коэффициент K0 успешно сохранен.", RowColor.green)
                self._request_calibration_temp_comp_k0_read()
            elif did == int(UdsData.fuel_zero_trim_count.pid):
                self._append_log("Калибровка: коррекция zero trim успешно сохранена.", RowColor.green)
                self._request_calibration_temp_comp_zero_trim_read()
                if bool(self._calibration_temp_comp_zero_trim_verify_pending):
                    if not self._request_calibration_temp_comp_raw_level_read():
                        self._reset_calibration_temp_comp_zero_trim_verify_state()
                        self._set_calibration_temp_comp_operation_status(
                            "Автопроверка zero trim не запущена: не удалось отправить DID 0x0018.",
                            busy=False,
                            progress_percent=100,
                            determinate=True,
                        )
                        self._append_log(
                            "Калибровка: после записи zero trim не удалось запросить DID 0x0018 для автопроверки.",
                            RowColor.yellow,
                        )
            else:
                advanced_field = self._temp_comp_advanced_field_by_did(did)
                if advanced_field is not None:
                    field_key = str(advanced_field.get("key", ""))
                    label_text = str(advanced_field.get("label", "параметр"))
                    self._append_log(f"Калибровка: {label_text} успешно сохранен.", RowColor.green)
                    self._request_calibration_temp_comp_advanced_read(field_key)

            if self._calibration_restore_active and self._calibration_restore_current_did == did:
                self._send_next_calibration_restore_write()

    def _on_calibration_poll_tick(self):
        if not self._calibration_active:
            self._stop_calibration_poll_timer()
            return
        if not self._can.is_connect or not self._can.is_trace:
            return
        if self._source_address_busy:
            return
        if self._programming_active:
            return
        if self._calibration_temp_comp_adv_read_active:
            return
        self._request_calibration_runtime_snapshot()

    def _start_calibration_poll_timer(self):
        if self._calibration_poll_timer.interval() != self._calibration_poll_interval_ms:
            self._calibration_poll_timer.setInterval(self._calibration_poll_interval_ms)
        if not self._calibration_poll_timer.isActive():
            self._calibration_poll_timer.start()

    def _stop_calibration_poll_timer(self):
        if self._calibration_poll_timer.isActive():
            self._calibration_poll_timer.stop()

    def _send_next_calibration_restore_write(self):
        if not self._calibration_restore_active:
            return

        if len(self._calibration_restore_queue) == 0:
            self._calibration_restore_active = False
            self._calibration_restore_current_did = None
            self._append_log("Калибровка: восстановление из резервной копии завершено.", RowColor.green)
            return

        did, value = self._calibration_restore_queue.pop(0)
        self._calibration_restore_current_did = int(did)
        target_var = UdsData.get_var_by_pid(int(did))
        if target_var is None:
            self._calibration_restore_active = False
            self._calibration_restore_current_did = None
            self._append_log(f"Калибровка: восстановление остановлено, DID 0x{int(did) & 0xFFFF:04X} не найден.", RowColor.red)
            return
        self._configure_calibration_uds_services()
        if self._calibration_write_service.write_data(
            target_var,
            int(value),
            tx_identifier=self._build_calibration_tx_identifier(),
        ):
            self._calibration_write_verify_pending[int(did)] = int(value)
            if int(did) == int(UdsData.empty_fuel_tank.pid):
                self._calibration_level0_written = True
                self._calibration_verify0_ok = False
            elif int(did) == int(UdsData.full_fuel_tank.pid):
                self._calibration_level100_written = True
                self._calibration_verify100_ok = False
            self.calibrationVerificationChanged.emit()
            self._recompute_calibration_wizard_state()
            self._append_log(f"Калибровка: восстановление DID 0x{int(did):04X} = {int(value)}.", RowColor.blue)
        else:
            self._calibration_restore_active = False
            self._calibration_restore_current_did = None
            self._append_log(f"Калибровка: ошибка восстановления DID 0x{int(did):04X}.", RowColor.red)

    def _reset_calibration_wizard_state(self):
        self._calibration_session_ready = False
        self._calibration_level0_written = False
        self._calibration_level100_written = False
        self._calibration_verify0_ok = False
        self._calibration_verify100_ok = False
        self.calibrationVerificationChanged.emit()
        self._recompute_calibration_wizard_state()

    def _recompute_calibration_wizard_state(self):
        if not self._calibration_active:
            stage = 0
            hint = "Запустите калибровку, чтобы начать пошаговый процесс."
        elif self._calibration_verify0_ok and self._calibration_verify100_ok:
            stage = 4
            hint = "Проверка значений 0% и 100% успешно завершена."
        elif self._calibration_level100_written:
            stage = 3
            hint = "Эталон 100% записан. Дождитесь автопроверки."
        elif self._calibration_level0_written:
            stage = 2
            hint = "Эталон 0% записан. Запишите 100%."
        elif self._calibration_session_ready:
            stage = 1
            hint = "Сессия открыта. Сохраните эталон 0%, затем 100%."
        else:
            stage = 0
            hint = "Ожидание подтверждения сессии UDS."

        changed = False
        if self._calibration_wizard_stage != stage:
            self._calibration_wizard_stage = stage
            changed = True
        if self._calibration_wizard_hint != hint:
            self._calibration_wizard_hint = hint
            changed = True

        if changed:
            self.calibrationWizardChanged.emit()

    def _refresh_calibration_node_options(self):
        previous_values = list(self._calibration_node_values)
        previous_options = list(self._calibration_node_options)
        previous_selected = int(self._selected_calibration_node_index)
        selected_sa = self._calibration_target_node_sa

        new_values: list[int | None] = [None]
        new_options: list[str] = ["Авто (по текущим UDS ID)"]

        merged_candidates = set(int(value) & 0xFF for value in self._observed_candidate_values)

        for sa in sorted(merged_candidates):
            new_values.append(sa)
            # Статичные подписи: без live-счетчиков, чтобы выбор не "прыгал" при обновлении трафика.
            new_options.append(f"Узел 0x{sa:02X}")

        new_selected = 0
        if selected_sa is not None:
            for index, value in enumerate(new_values):
                if value is not None and int(value) == int(selected_sa):
                    new_selected = index
                    break
            else:
                self._calibration_target_node_sa = None
        else:
            # При режиме "Авто" всегда держим индекс 0, чтобы отображение не расходилось с реальной логикой target SA.
            new_selected = 0

        self._calibration_node_values = new_values
        self._calibration_node_options = new_options
        self._selected_calibration_node_index = new_selected

        if (
            previous_values != self._calibration_node_values
            or previous_options != self._calibration_node_options
            or previous_selected != self._selected_calibration_node_index
        ):
            self.calibrationNodeSelectionChanged.emit()

    @staticmethod
    def _resolve_calibration_write_value(text, fallback_value: int) -> int:
        raw = str(text).strip()
        if not raw:
            return int(fallback_value)

        base = 16 if raw.lower().startswith("0x") else 10
        try:
            value = int(raw, base)
        except ValueError as exc:
            raise ValueError("Некорректное значение калибровки. Используйте десятичный или hex-формат.") from exc

        if value < 0 or value > 0xFFFF:
            raise ValueError("Значение калибровки вне диапазона 0..65535.")
        return int(value)

    @classmethod
    def _resolve_calibration_signed_int16_value(
        cls,
        text,
        fallback_value: int | None,
        value_label: str,
        did_hint: str,
    ) -> int:
        """Цель функции в едином разборе signed int16, затем она валидирует dec/hex и подготавливает значение для UDS 0x2E."""
        raw = str(text).strip()
        if not raw:
            if fallback_value is None:
                raise ValueError(
                    f"Текущее значение {value_label} неизвестно. "
                    f"Сначала прочитайте DID {did_hint} или введите число вручную."
                )
            return cls._saturate_int16(int(fallback_value))

        base = 16 if raw.lower().startswith(("0x", "-0x", "+0x")) else 10
        try:
            value = int(raw, base)
        except ValueError as exc:
            raise ValueError(
                f"Некорректное значение {value_label}. Используйте signed dec или 0xHEX."
            ) from exc

        if base == 16 and value >= 0:
            if value > 0xFFFF:
                raise ValueError(f"HEX-значение {value_label} должно быть в диапазоне 0x0000..0xFFFF.")
            if value > cls._INT16_MAX:
                value -= 0x10000

        if value < cls._INT16_MIN or value > cls._INT16_MAX:
            raise ValueError(f"Значение {value_label} вне диапазона int16 (-32768..32767).")

        return int(value)

    @classmethod
    def _resolve_calibration_k1_write_value(cls, text, fallback_value: int | None) -> int:
        """Цель функции в удобном вводе K1 из UI, затем она парсит dec/hex и приводит значение к int16."""
        return cls._resolve_calibration_signed_int16_value(text, fallback_value, "K1", "0x001B")

    @classmethod
    def _resolve_calibration_k0_write_value(cls, text, fallback_value: int | None) -> int:
        """Цель функции в удобном вводе K0 из UI, затем она парсит dec/hex и приводит значение к int16."""
        return cls._resolve_calibration_signed_int16_value(text, fallback_value, "K0", "0x001C")

    @classmethod
    def _resolve_calibration_zero_trim_write_value(cls, text, fallback_value: int | None) -> int:
        """Цель функции в удобном вводе zero trim из UI, затем она парсит dec/hex и приводит значение к int16."""
        return cls._resolve_calibration_signed_int16_value(text, fallback_value, "zero trim", "0x002D")

    @classmethod
    def _resolve_calibration_temp_comp_advanced_write_value(
        cls,
        field: dict[str, object],
        text,
        fallback_value: int | None,
    ) -> int:
        """Цель функции в валидации расширенного параметра компенсации, затем она готовит корректное payload-значение для DID."""
        field_label = str(field.get("label", "параметр"))
        field_var = field.get("var")
        did_hint = "----" if field_var is None else f"0x{int(field_var.pid) & 0xFFFF:04X}"

        if bool(field.get("signed", False)):
            return cls._resolve_calibration_signed_int16_value(text, fallback_value, field_label, did_hint)

        min_value = int(field.get("min", 0))
        max_value = int(field.get("max", 0xFFFF))
        raw = str(text).strip()
        if not raw:
            if fallback_value is None:
                raise ValueError(
                    f"Текущее значение {field_label} неизвестно. "
                    f"Сначала прочитайте DID {did_hint} или введите число вручную."
                )
            fallback_int = int(fallback_value)
            if fallback_int < min_value or fallback_int > max_value:
                raise ValueError(
                    f"Текущее значение {field_label} вне диапазона {min_value}..{max_value}. "
                    "Введите новое значение вручную."
                )
            return fallback_int

        base = 16 if raw.lower().startswith("0x") else 10
        try:
            value = int(raw, base)
        except ValueError as exc:
            raise ValueError(
                f"Некорректное значение {field_label}. Используйте dec или 0xHEX."
            ) from exc

        if value < min_value or value > max_value:
            raise ValueError(
                f"Значение {field_label} вне диапазона {min_value}..{max_value}."
            )

        return int(value)
